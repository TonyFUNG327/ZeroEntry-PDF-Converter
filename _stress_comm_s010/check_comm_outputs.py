from pathlib import Path

from openpyxl import load_workbook


output_dir = Path("_stress_comm_s010") / "output"
issues = []
bad_desc_markers = [
    "TOTAL TRANSACTION",
    "NO.OF TRANSACTION",
    "CONSOLIDATED STATEMENT",
    "DEPOSIT INTEREST RATE",
    "IN CASE OF ERROR",
    "TO BE CONTINUED",
    "END OF STATEMENT",
    "Customer Services",
    "PRINTED ON",
    "PAGE",
]

for path in sorted(output_dir.glob("*.xlsx"), key=lambda p: p.name.lower()):
    workbook = load_workbook(path, data_only=False)
    print("====", path.name, workbook.sheetnames)
    expected_sheets = {"COMM HKD Savings", "COMM USD Savings", "COMM HKD Current"}
    if set(workbook.sheetnames) != expected_sheets:
        issues.append((path.name, "unexpected sheets", workbook.sheetnames))
    for sheet in workbook.worksheets:
        running = None
        mismatches = []
        bad_dates = []
        bad_descriptions = []
        bad_formats = []
        for row_idx in range(2, sheet.max_row + 1):
            date_value = sheet.cell(row_idx, 2).value
            description = str(sheet.cell(row_idx, 3).value or "")
            deposit = sheet.cell(row_idx, 4).value or 0
            withdrawal = sheet.cell(row_idx, 5).value or 0
            balance = sheet.cell(row_idx, 6).value
            if date_value and not (2023 <= date_value.year <= 2024):
                bad_dates.append((row_idx, date_value))
            if any(marker in description for marker in bad_desc_markers):
                bad_descriptions.append((row_idx, description[:120]))
            if sheet.cell(row_idx, 5).number_format != '(#,##0.00);(#,##0.00);"-"':
                bad_formats.append((row_idx, sheet.cell(row_idx, 5).number_format))
            if deposit == 0 and withdrawal == 0 and balance is not None:
                running = balance
            else:
                if running is None:
                    running = balance or 0
                running += deposit
                running -= withdrawal
            if balance is not None and abs(balance - running) > 0.01:
                mismatches.append((row_idx, balance, round(running, 2)))
        print(
            sheet.title,
            "rows",
            sheet.max_row - 1,
            "final",
            sheet.cell(sheet.max_row, 6).value,
            "mismatches",
            len(mismatches),
            "bad_dates",
            len(bad_dates),
            "bad_desc",
            len(bad_descriptions),
            "bad_formats",
            len(bad_formats),
        )
        if mismatches or bad_dates or bad_descriptions or bad_formats:
            issues.append((path.name, sheet.title, mismatches[:5], bad_dates[:5], bad_descriptions[:5], bad_formats[:5]))

print("ISSUES", len(issues))
for issue in issues[:20]:
    print(issue)
