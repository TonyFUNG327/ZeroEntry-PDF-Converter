from pathlib import Path

from openpyxl import load_workbook


output_dir = Path("_stress_unknown_h026") / "output"
issues = []
summary = []

bad_desc_markers = [
    "TotalNo.",
    "TotalDepositAmount",
    "TotalWithdrawalAmount",
    "PortfolioSummary",
    "The Hongkong",
    "Thank you",
    "SpecialPrivileges",
    "ExchangeRate",
    "Commercial Tariffs",
]

for path in sorted(output_dir.glob("*.xlsx"), key=lambda p: p.name):
    workbook = load_workbook(path, data_only=False)
    print("====", path.name, workbook.sheetnames)
    for sheet in workbook.worksheets:
        running = None
        mismatches = []
        bad_descriptions = []
        bad_dates = []
        bad_formats = []
        for row_idx in range(2, sheet.max_row + 1):
            date_value = sheet.cell(row_idx, 2).value
            description = str(sheet.cell(row_idx, 3).value or "")
            deposit = sheet.cell(row_idx, 4).value or 0
            withdrawal = sheet.cell(row_idx, 5).value or 0
            balance = sheet.cell(row_idx, 6).value
            control_formula = str(sheet.cell(row_idx, 7).value or "")

            if date_value and not (2024 <= date_value.year <= 2026):
                bad_dates.append((row_idx, date_value))
            if any(marker in description for marker in bad_desc_markers):
                bad_descriptions.append((row_idx, description[:100]))
            if sheet.cell(row_idx, 5).number_format != '(#,##0.00);(#,##0.00);"-"':
                bad_formats.append((row_idx, sheet.cell(row_idx, 5).number_format))
            if not control_formula.startswith("=IF("):
                issues.append((path.name, sheet.title, row_idx, "missing control formula"))

            if deposit == 0 and withdrawal == 0 and balance is not None:
                running = balance
            else:
                if running is None:
                    running = balance or 0
                running += deposit
                running -= withdrawal
            if balance is not None and abs(balance - running) > 0.01:
                mismatches.append((row_idx, balance, round(running, 2)))

        first = [sheet.cell(2, col).value for col in range(1, 7)] if sheet.max_row >= 2 else []
        last = [sheet.cell(sheet.max_row, col).value for col in range(1, 7)] if sheet.max_row >= 2 else []
        print(sheet.title, "rows", sheet.max_row - 1, "first", first, "last", last)
        print("mismatches", mismatches[:5], "bad_dates", bad_dates[:5], "bad_desc", bad_descriptions[:5], "bad_formats", bad_formats[:5])
        summary.append((path.name, sheet.title, sheet.max_row - 1, len(mismatches), len(bad_dates), len(bad_descriptions), len(bad_formats)))
        if mismatches or bad_dates or bad_descriptions or bad_formats:
            issues.append((path.name, sheet.title, mismatches[:5], bad_dates[:5], bad_descriptions[:5], bad_formats[:5]))

print("SUMMARY")
for item in summary:
    print(item)
print("ISSUES", len(issues))
for issue in issues[:20]:
    print(issue)
