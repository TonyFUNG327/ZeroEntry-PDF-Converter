from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
BANK_PATH = BASE / "001-S240-Bank record-31.12.24.converted.xlsx"
REPORT_PATH = BASE / "S240-TB,GL,BS,PL-31.12.24.xlsx"
OUT = BASE / "accounting_mapping_notes.md"


def fmt(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def row_values(ws, row, start_col, end_col):
    return [ws.cell(row, c).value for c in range(start_col, end_col + 1)]


def bank_sheet_notes(ws):
    lines = [f"## Bank record sheet: {ws.title}", ""]
    lines.append(f"- Bank/account label: {fmt(ws['C3'].value)}")
    lines.append("- Core input columns:")
    core_headers = []
    for c in range(1, 13):
        h = ws.cell(7, c).value or ws.cell(8, c).value
        if h not in (None, ""):
            core_headers.append(f"{get_column_letter(c)}={h}")
    lines.append("  " + ", ".join(core_headers))

    categories = []
    for c in range(13, ws.max_column + 1):
        label = ws.cell(7, c).value
        if label not in (None, ""):
            categories.append((get_column_letter(c), label))
    lines.append("- Accounting classification columns from row 7:")
    lines.append("  " + "; ".join(f"{col}={label}" for col, label in categories))

    lines.append("- Sample transaction rows:")
    count = 0
    for r in range(9, ws.max_row + 1):
        date = ws.cell(r, 2).value
        particular = ws.cell(r, 4).value
        deposit = ws.cell(r, 8).value
        withdrawal = ws.cell(r, 9).value
        balance = ws.cell(r, 10).value
        control = ws.cell(r, 12).value
        if any(v not in (None, "") for v in [date, particular, deposit, withdrawal, balance]):
            active = []
            for c, label in categories:
                value = ws[f"{c}{r}"].value
                if isinstance(value, (int, float)) and abs(value) > 0.00001:
                    active.append(f"{label}={value}")
            lines.append(
                f"  - row {r}: date={fmt(date)}, particular={fmt(particular)}, "
                f"deposit={fmt(deposit)}, withdrawal={fmt(withdrawal)}, balance={fmt(balance)}, "
                f"control={fmt(control)}, classified=[{'; '.join(active)}]"
            )
            count += 1
            if count >= 12:
                break
    return "\n".join(lines)


def report_sheet_notes(ws, max_rows=80):
    lines = [f"## Report sheet: {ws.title}", ""]
    for r in range(1, min(ws.max_row, max_rows) + 1):
        vals = row_values(ws, r, 1, min(ws.max_column, 9))
        if any(v not in (None, "") for v in vals):
            lines.append(f"- row {r}: " + " | ".join(fmt(v) for v in vals))
    return "\n".join(lines)


def main():
    bank_wb = load_workbook(BANK_PATH, data_only=False)
    report_wb_formula = load_workbook(REPORT_PATH, data_only=False)
    report_wb_values = load_workbook(REPORT_PATH, data_only=True)

    sections = ["# Accounting Mapping Notes", ""]
    for ws in bank_wb.worksheets:
        if ws.title.lower() != "instruction":
            sections.append(bank_sheet_notes(ws))
            sections.append("")

    for ws in report_wb_formula.worksheets:
        sections.append(report_sheet_notes(ws))
        sections.append("")

    sections.append("# Report Values Snapshot")
    sections.append("")
    for ws in report_wb_values.worksheets:
        sections.append(report_sheet_notes(ws, max_rows=60))
        sections.append("")

    OUT.write_text("\n".join(sections), encoding="utf-8")
    print(OUT)


if __name__ == "__main__":
    main()
