from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
PATH = BASE / "001-S240-Bank record-31.12.24.converted.xlsx"


def fmt(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def main():
    wb_formula = load_workbook(PATH, data_only=False)
    wb_values = load_workbook(PATH, data_only=True)
    for ws in wb_formula.worksheets:
        if ws.title.lower() == "instruction":
            continue
        vws = wb_values[ws.title]
        print(f"SHEET {ws.title}")
        categories = []
        for c in range(12, ws.max_column + 1):
            label = ws.cell(7, c).value
            if label not in (None, ""):
                categories.append((c, label))
        found = 0
        for r in range(9, ws.max_row + 1):
            active = []
            for c, label in categories:
                value = vws.cell(r, c).value
                if isinstance(value, (int, float)) and abs(value) > 0.00001:
                    active.append(f"{get_column_letter(c)}:{label}={value}")
            if active:
                print(
                    f"row {r}: date={fmt(vws.cell(r, 2).value)} "
                    f"particular={fmt(vws.cell(r, 4).value or vws.cell(r, 3).value)} "
                    f"deposit={fmt(vws.cell(r, 8).value)} withdrawal={fmt(vws.cell(r, 9).value)} "
                    f"classified=[{'; '.join(active)}]"
                )
                found += 1
                if found >= 30:
                    break
        print(f"found={found}\n")


if __name__ == "__main__":
    main()
