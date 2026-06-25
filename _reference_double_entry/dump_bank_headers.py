from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
PATH = BASE / "001-S240-Bank record-31.12.24.converted.xlsx"


def main():
    wb = load_workbook(PATH, data_only=False)
    for sheet_name in ["HSBC HKD CA", "USD SA"]:
        ws = wb[sheet_name]
        print(f"SHEET {sheet_name}")
        for r in range(1, 9):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    vals.append(f"{get_column_letter(c)}={v}")
            print(f"row {r}: " + "; ".join(vals))
        print()


if __name__ == "__main__":
    main()
