from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
FILES = {
    "bank_record": BASE / "001-S240-Bank record-31.12.24.converted.xlsx",
    "reports": BASE / "S240-TB,GL,BS,PL-31.12.24.xlsx",
}


def cell_value_for_dump(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def used_bounds(ws):
    min_row, min_col = None, None
    max_row, max_col = 0, 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if min_row is None:
        return None
    return {"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col}


def sample_rows(ws, max_rows=20, max_cols=16):
    rows = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, max_rows),
        min_col=1,
        max_col=min(ws.max_column, max_cols),
        values_only=True,
    ):
        rows.append([cell_value_for_dump(v) for v in row])
    return rows


def formula_samples(ws, limit=25):
    out = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                out.append({"cell": cell.coordinate, "formula": cell.value})
                if len(out) >= limit:
                    return out
    return out


def nonempty_row_summaries(ws, limit=40):
    out = []
    bounds = used_bounds(ws)
    if not bounds:
        return out
    for r in range(bounds["min_row"], bounds["max_row"] + 1):
        vals = []
        for c in range(bounds["min_col"], min(bounds["max_col"], bounds["min_col"] + 11) + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                vals.append(cell_value_for_dump(v))
        if vals:
            out.append({"row": r, "values": vals})
            if len(out) >= limit:
                break
    return out


def workbook_summary(path):
    wb_formula = load_workbook(path, data_only=False)
    wb_values = load_workbook(path, data_only=True)
    result = {"path": str(path), "sheets": []}
    for ws in wb_formula.worksheets:
        ws_values = wb_values[ws.title]
        formulas = formula_samples(ws)
        result["sheets"].append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "used_bounds": used_bounds(ws),
                "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:20]],
                "formula_count_sampled": len(formulas),
                "formula_samples": formulas,
                "sample_rows_formulas": sample_rows(ws),
                "sample_rows_values": sample_rows(ws_values),
                "nonempty_rows": nonempty_row_summaries(ws),
            }
        )
    return result


def main():
    summaries = {name: workbook_summary(path) for name, path in FILES.items()}
    out_path = BASE / "workbook_structure_summary.json"
    out_path.write_text(json.dumps(summaries, indent=2, ensure_ascii=False), encoding="utf-8")
    print(out_path)
    for key, summary in summaries.items():
        print(f"\n## {key}: {Path(summary['path']).name}")
        for sheet in summary["sheets"]:
            bounds = sheet["used_bounds"]
            print(
                f"- {sheet['name']}: rows={sheet['max_row']} cols={sheet['max_col']} "
                f"used={bounds} formulas_sampled={sheet['formula_count_sampled']}"
            )


if __name__ == "__main__":
    main()
