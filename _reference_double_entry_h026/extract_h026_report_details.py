from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
PATH = BASE / "999-H026-TB,GL,BS,IS-31.12.25.xlsx"


def fmt(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def print_sheet_rows(sheet_name, start, end):
    wb_formula = load_workbook(PATH, data_only=False)
    wb_values = load_workbook(PATH, data_only=True)
    ws_f = wb_formula[sheet_name]
    ws_v = wb_values[sheet_name]
    print(f"\nSHEET {sheet_name} rows {start}-{end}")
    for r in range(start, min(end, ws_f.max_row) + 1):
        formula = [fmt(ws_f.cell(r, c).value) for c in range(1, ws_f.max_column + 1)]
        values = [fmt(ws_v.cell(r, c).value) for c in range(1, ws_v.max_column + 1)]
        if any(formula):
            if formula == values:
                print(f"{r}: " + " | ".join(formula))
            else:
                print(f"{r}: F " + " | ".join(formula))
                print(f"   V " + " | ".join(values))


def main():
    print_sheet_rows("GL-31.12.25", 60, 108)
    print_sheet_rows("IS-31.12.25", 1, 27)
    print_sheet_rows("BS-31.12.25", 1, 45)


if __name__ == "__main__":
    main()
