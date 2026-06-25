from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
BANK_PATH = BASE / "001-H026-Bank record-31.12.25.converted.xlsx"
REPORT_PATH = BASE / "999-H026-TB,GL,BS,IS-31.12.25.xlsx"
JSON_OUT = BASE / "h026_structure_summary.json"
MD_OUT = BASE / "h026_accounting_mapping_notes.md"


def fmt(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def used_bounds(ws):
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if min_row is None:
        return None
    return {"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col}


def formula_samples(ws, limit=30):
    out = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                out.append({"cell": cell.coordinate, "formula": cell.value})
                if len(out) >= limit:
                    return out
    return out


def workbook_summary(path):
    wb = load_workbook(path, data_only=False)
    result = {"path": str(path), "sheets": []}
    for ws in wb.worksheets:
        result["sheets"].append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "used_bounds": used_bounds(ws),
                "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:30]],
                "formula_samples": formula_samples(ws),
            }
        )
    return result


def nonempty_row(ws, row, max_col=None):
    vals = []
    end = max_col or ws.max_column
    for c in range(1, end + 1):
        v = ws.cell(row, c).value
        if v not in (None, ""):
            vals.append(f"{get_column_letter(c)}={fmt(v)}")
    return vals


def bank_sheet_notes(ws_formula, ws_values):
    lines = [f"## Bank record sheet: {ws_formula.title}", ""]
    lines.append(f"- Company: {fmt(ws_values['C1'].value)}")
    lines.append(f"- Period: {fmt(ws_values['C2'].value)}")
    lines.append(f"- Account label: {fmt(ws_values['C3'].value)}")
    for r in range(1, min(9, ws_formula.max_row + 1)):
        vals = nonempty_row(ws_formula, r)
        if vals:
            lines.append(f"- header row {r}: " + "; ".join(vals))

    likely_class_cols = []
    for c in range(1, ws_formula.max_column + 1):
        label = ws_formula.cell(7, c).value
        if label not in (None, ""):
            likely_class_cols.append((c, label))
    lines.append("- Row 7 labels:")
    lines.append("  " + "; ".join(f"{get_column_letter(c)}={label}" for c, label in likely_class_cols))

    active_by_label = defaultdict(float)
    active_rows = []
    for r in range(1, ws_values.max_row + 1):
        active = []
        for c, label in likely_class_cols:
            value = ws_values.cell(r, c).value
            if isinstance(value, (int, float)) and abs(value) > 0.00001:
                active.append((c, label, value))
                active_by_label[str(label)] += value
        if active:
            active_rows.append((r, active))

    lines.append("- First active classification/control rows:")
    for r, active in active_rows[:30]:
        date = ws_values.cell(r, 2).value
        particular = ws_values.cell(r, 4).value or ws_values.cell(r, 3).value
        amount_bits = []
        for c in range(1, min(ws_values.max_column, 16) + 1):
            label = ws_formula.cell(7, c).value or ws_formula.cell(8, c).value
            val = ws_values.cell(r, c).value
            if label in ("Deposit", "Withdrawal", "Balance", "Control") and val not in (None, ""):
                amount_bits.append(f"{get_column_letter(c)}:{label}={fmt(val)}")
        active_bits = "; ".join(f"{get_column_letter(c)}:{label}={value}" for c, label, value in active)
        lines.append(
            f"  - row {r}: date={fmt(date)}, particular={fmt(particular)}, "
            f"amounts=[{'; '.join(amount_bits)}], active=[{active_bits}]"
        )

    lines.append("- Non-zero totals by row 7 label:")
    for label, total in sorted(active_by_label.items()):
        if abs(total) > 0.00001:
            lines.append(f"  - {label}: {total}")
    return "\n".join(lines)


def report_sheet_notes(ws_formula, ws_values, max_rows=120):
    lines = [f"## Report sheet: {ws_formula.title}", ""]
    for r in range(1, min(ws_formula.max_row, max_rows) + 1):
        vals_formula = []
        vals_values = []
        for c in range(1, min(ws_formula.max_column, 12) + 1):
            fv = ws_formula.cell(r, c).value
            vv = ws_values.cell(r, c).value
            if fv not in (None, "") or vv not in (None, ""):
                vals_formula.append(fmt(fv))
                vals_values.append(fmt(vv))
        if vals_formula:
            if vals_formula == vals_values:
                lines.append(f"- row {r}: " + " | ".join(vals_formula))
            else:
                lines.append(f"- row {r}: formula=[" + " | ".join(vals_formula) + "]")
                lines.append(f"          values =[" + " | ".join(vals_values) + "]")
    return "\n".join(lines)


def main():
    summaries = {
        "bank_record": workbook_summary(BANK_PATH),
        "reports": workbook_summary(REPORT_PATH),
    }
    JSON_OUT.write_text(json.dumps(summaries, indent=2, ensure_ascii=False), encoding="utf-8")

    bank_formula = load_workbook(BANK_PATH, data_only=False)
    bank_values = load_workbook(BANK_PATH, data_only=True)
    report_formula = load_workbook(REPORT_PATH, data_only=False)
    report_values = load_workbook(REPORT_PATH, data_only=True)

    sections = ["# H026 Accounting Mapping Notes", ""]
    sections.append("## Workbook Overview")
    for key, summary in summaries.items():
        sections.append(f"- {key}: {Path(summary['path']).name}")
        for sheet in summary["sheets"]:
            sections.append(
                f"  - {sheet['name']}: rows={sheet['max_row']}, cols={sheet['max_col']}, "
                f"used={sheet['used_bounds']}, formulas_sampled={len(sheet['formula_samples'])}"
            )
    sections.append("")

    for ws in bank_formula.worksheets:
        if ws.title.lower() != "instruction":
            sections.append(bank_sheet_notes(ws, bank_values[ws.title]))
            sections.append("")

    for ws in report_formula.worksheets:
        sections.append(report_sheet_notes(ws, report_values[ws.title]))
        sections.append("")

    MD_OUT.write_text("\n".join(sections), encoding="utf-8")
    print(JSON_OUT)
    print(MD_OUT)


if __name__ == "__main__":
    main()
