from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
PATH = BASE / "001-H026-Bank record-31.12.25.converted.xlsx"


def fmt(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def main():
    wb_formula = load_workbook(PATH, data_only=False)
    wb_values = load_workbook(PATH, data_only=True)
    for ws in wb_formula.worksheets:
        if ws.title == "Chart of Accounts":
            continue
        vws = wb_values[ws.title]
        print(f"\nSHEET {ws.title}")
        for r in range(1, min(ws.max_row, 15) + 1):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    vals.append(f"{get_column_letter(c)}={fmt(v)}")
            if vals:
                print(f"row {r}: " + "; ".join(vals))

        print("data/classification rows:")
        for r in range(1, ws.max_row + 1):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = vws.cell(r, c).value
                if isinstance(v, (int, float)) and abs(v) > 0.00001:
                    header = ws.cell(7, c).value or ws.cell(8, c).value or ws.cell(9, c).value
                    vals.append(f"{get_column_letter(c)}:{fmt(header)}={v}")
            if vals:
                context = " | ".join(fmt(vws.cell(r, c).value) for c in range(1, min(8, ws.max_column) + 1))
                print(f"row {r}: {context} || " + "; ".join(vals))

    coa = wb_values["Chart of Accounts"]
    print("\nCHART OF ACCOUNTS first 80:")
    for r in range(1, min(coa.max_row, 80) + 1):
        a, b = coa.cell(r, 1).value, coa.cell(r, 2).value
        if a not in (None, "") or b not in (None, ""):
            print(f"{r}: {fmt(a)} => {fmt(b)}")


if __name__ == "__main__":
    main()
