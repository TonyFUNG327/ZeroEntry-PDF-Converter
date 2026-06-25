from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SHEETS = {
    "ICBC Asia USD Current",
    "ICBC Asia HKD Current",
    "ICBC Asia CNY Current",
}


def main():
    issues = []
    saw_negative_balance = False
    saw_frn_continuation = False
    for path in sorted(Path("_reference_icbc_asia/output").glob("*.xlsx")):
        workbook = load_workbook(path, data_only=False)
        print("====", path.name, workbook.sheetnames)
        missing = EXPECTED_SHEETS.difference(workbook.sheetnames)
        if missing:
            issues.append((path.name, "missing sheets", sorted(missing)))
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            deposits = 0.0
            withdrawals = 0.0
            control = None
            mismatches = 0
            bad_desc = 0
            bad_format = 0
            for row_idx in range(2, sheet.max_row + 1):
                desc = str(sheet.cell(row_idx, 3).value or "")
                deposit = sheet.cell(row_idx, 4).value
                withdrawal = sheet.cell(row_idx, 5).value
                balance = sheet.cell(row_idx, 6).value
                withdrawal_format = sheet.cell(row_idx, 5).number_format
                if balance is not None and balance < 0:
                    saw_negative_balance = True
                if "FRN" in desc:
                    saw_frn_continuation = True
                if withdrawal and "(" not in withdrawal_format:
                    bad_format += 1
                if any(
                    marker in desc
                    for marker in [
                        "中國工商銀行",
                        "INDUSTRIAL AND COMMERCIAL",
                        "更改地址通知书",
                        "各种交易代码",
                        "客户服务热线",
                        "备注：",
                    ]
                ):
                    bad_desc += 1
                if deposit is None and withdrawal is None:
                    control = balance
                else:
                    control = round((control or 0) + (deposit or 0) - (withdrawal or 0), 2)
                if round(balance or 0, 2) != round(control or 0, 2):
                    mismatches += 1
                deposits += deposit or 0
                withdrawals += withdrawal or 0
            print(
                sheet_name,
                "rows", sheet.max_row - 1,
                "deposit", round(deposits, 2),
                "withdrawal", round(withdrawals, 2),
                "mismatches", mismatches,
                "bad_desc", bad_desc,
                "bad_format", bad_format,
            )
            if mismatches or bad_desc or bad_format:
                issues.append((path.name, sheet_name, mismatches, bad_desc, bad_format))
    if not saw_negative_balance:
        issues.append(("all", "negative DR balance not observed"))
    if not saw_frn_continuation:
        issues.append(("all", "FRN continuation not observed"))
    print("ISSUES", len(issues))
    for issue in issues:
        print(issue)
    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
