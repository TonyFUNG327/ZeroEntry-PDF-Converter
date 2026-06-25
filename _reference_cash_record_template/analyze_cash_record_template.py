from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
PATH = BASE / "002-Template-Cash record-dd.mm.yy.converted.xlsx"
JSON_OUT = BASE / "cash_record_template_structure.json"
MD_OUT = BASE / "cash_record_template_notes.md"


def fmt(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "date"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def used_bounds(ws) -> dict[str, int] | None:
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if min_row is None:
        return None
    return {"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col}


def detect_header_row(ws) -> int | None:
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [str(ws.cell(row, col).value or "").strip().lower() for col in range(1, ws.max_column + 1)]
        joined = "|".join(values)
        if "no." in joined and "date" in joined and ("particular" in joined or "nature" in joined):
            return row
    return None


def formula_samples(ws, limit: int = 80) -> list[dict[str, str]]:
    samples = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                samples.append({"cell": cell.coordinate, "formula": cell.value})
                if len(samples) >= limit:
                    return samples
    return samples


def nonempty_rows(ws, max_rows: int = 80, max_cols: int = 80) -> list[dict[str, Any]]:
    rows = []
    for row in range(1, min(ws.max_row, max_rows) + 1):
        cells = []
        for col in range(1, min(ws.max_column, max_cols) + 1):
            value = ws.cell(row, col).value
            if value not in (None, ""):
                cells.append({"cell": f"{get_column_letter(col)}{row}", "value": fmt(value)})
        if cells:
            rows.append({"row": row, "cells": cells})
    return rows


def labels_at(ws, row: int | None) -> dict[str, str]:
    if not row:
        return {}
    return {
        get_column_letter(col): fmt(ws.cell(row, col).value)
        for col in range(1, ws.max_column + 1)
        if ws.cell(row, col).value not in (None, "")
    }


def main() -> None:
    wb_formula = load_workbook(PATH, data_only=False)
    wb_values = load_workbook(PATH, data_only=True)
    summary = {"path": str(PATH), "sheets": []}
    notes = ["# Cash Record Template Notes", ""]
    notes.append("## Workbook Overview")
    for ws in wb_formula.worksheets:
        header_row = detect_header_row(ws)
        item = {
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "used_bounds": used_bounds(ws),
            "header_row": header_row,
            "header_labels": labels_at(ws, header_row),
            "subheader_labels": labels_at(ws, header_row + 1 if header_row else None),
            "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:30]],
            "formula_samples": formula_samples(ws),
            "nonempty_rows_formula": nonempty_rows(ws),
            "nonempty_rows_values": nonempty_rows(wb_values[ws.title]),
        }
        summary["sheets"].append(item)
        notes.append(
            f"- {ws.title}: rows={ws.max_row}, cols={ws.max_column}, "
            f"used={item['used_bounds']}, header_row={header_row}, formulas={len(item['formula_samples'])}"
        )
    notes.append("")

    for item in summary["sheets"]:
        notes.append(f"## Sheet: {item['name']}")
        notes.append("")
        notes.append(f"- Used bounds: {item['used_bounds']}")
        notes.append(f"- Header row: {item['header_row']}")
        if item["merged_ranges"]:
            notes.append("- Merged ranges: " + "; ".join(item["merged_ranges"]))
        if item["header_labels"]:
            notes.append("- Header labels:")
            notes.append("  " + "; ".join(f"{k}={v}" for k, v in item["header_labels"].items()))
        if item["subheader_labels"]:
            notes.append("- Subheader labels:")
            notes.append("  " + "; ".join(f"{k}={v}" for k, v in item["subheader_labels"].items()))
        if item["formula_samples"]:
            notes.append("- Formula samples:")
            for sample in item["formula_samples"][:25]:
                notes.append(f"  - {sample['cell']}: {sample['formula']}")
        notes.append("- Non-empty value rows:")
        for row in item["nonempty_rows_values"][:35]:
            notes.append(
                f"  - row {row['row']}: "
                + "; ".join(f"{cell['cell']}={cell['value']}" for cell in row["cells"][:30])
            )
        notes.append("")

    JSON_OUT.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    MD_OUT.write_text("\n".join(notes), encoding="utf-8")
    print(JSON_OUT)
    print(MD_OUT)


if __name__ == "__main__":
    main()
