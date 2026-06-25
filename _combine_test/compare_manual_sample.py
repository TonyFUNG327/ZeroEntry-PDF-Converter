from pathlib import Path

from openpyxl import load_workbook


manual = load_workbook(Path("_reference_combine") / "HSBC 2501 - 2024 (手動合并).xlsx", data_only=False)
generated = load_workbook(Path("_combine_test") / "BB3_2501_2504" / "HSBC 2501 - HSBC 2504 combined.xlsx", data_only=False)

issues = []
for sheet_name in manual.sheetnames:
    if sheet_name not in generated.sheetnames:
        issues.append((sheet_name, "missing sheet"))
        continue
    ws_m = manual[sheet_name]
    ws_g = generated[sheet_name]
    print("====", sheet_name, "manual rows", ws_m.max_row, "generated rows", ws_g.max_row)
    if ws_m.max_row != ws_g.max_row:
        issues.append((sheet_name, "row count", ws_m.max_row, ws_g.max_row))
    for row_idx in range(1, min(ws_m.max_row, ws_g.max_row) + 1):
        manual_values = [ws_m.cell(row_idx, col).value for col in range(1, 7)]
        generated_values = [ws_g.cell(row_idx, col).value for col in range(1, 7)]
        if manual_values != generated_values:
            issues.append((sheet_name, row_idx, manual_values, generated_values))
            if len(issues) > 20:
                break
    print("last manual", [ws_m.cell(ws_m.max_row, col).value for col in range(1, 8)])
    print("last generated", [ws_g.cell(ws_g.max_row, col).value for col in range(1, 8)])

print("issues", len(issues))
for issue in issues[:20]:
    print(issue)
