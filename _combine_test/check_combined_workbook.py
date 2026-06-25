from pathlib import Path

from openpyxl import load_workbook


path = Path("_combine_test") / "BB3_full" / "HSBC 2501 - HSBC 2601 combined.xlsx"
workbook = load_workbook(path, data_only=False)
for sheet in workbook.worksheets:
    print("====", sheet.title, "rows", sheet.max_row)
    mismatches = []
    running = None
    for row_idx in range(2, sheet.max_row + 1):
        deposit = sheet.cell(row_idx, 4).value or 0
        withdrawal = sheet.cell(row_idx, 5).value or 0
        balance = sheet.cell(row_idx, 6).value
        if deposit == 0 and withdrawal == 0 and balance is not None:
            running = balance
        else:
            if running is None:
                running = balance or 0
            running += deposit
            running -= withdrawal
        if balance is not None and abs(balance - running) > 0.01:
            mismatches.append((row_idx, balance, round(running, 2)))
    print("first", [sheet.cell(2, col).value for col in range(1, 7)])
    print("last", [sheet.cell(sheet.max_row, col).value for col in range(1, 7)])
    print("mismatches", mismatches[:10])
