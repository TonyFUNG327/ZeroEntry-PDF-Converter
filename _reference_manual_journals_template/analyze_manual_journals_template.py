from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
PATH = BASE / "009-Template-AR,AP, Accruals schedule-dd.mm.yy.xlsx"
JSON_OUT = BASE / "manual_journals_template_structure.json"
MD_OUT = BASE / "manual_journals_template_notes.md"


def fmt(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def used_bounds(ws):
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if min_row is None:
        return None
    return {"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col}


def formula_samples(ws, limit=40):
    out = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                out.append({"cell": cell.coordinate, "formula": cell.value})
                if len(out) >= limit:
                    return out
    return out


def nonempty_rows(ws, max_rows=80, max_cols=30):
    rows = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        vals = []
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                vals.append(f"{get_column_letter(c)}={fmt(v)}")
        if vals:
            rows.append((r, vals))
    return rows


def detect_header_candidates(ws):
    candidates = []
    key_terms = ("date", "description", "account", "amount", "customer", "supplier", "invoice", "reference", "dr", "cr")
    for r in range(1, min(ws.max_row, 50) + 1):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        score = sum(any(term in v for v in vals) for term in key_terms)
        filled = sum(1 for v in vals if v)
        if score >= 2 and filled >= 3:
            candidates.append({"row": r, "score": score, "filled": filled})
    return candidates[:10]


def sheet_notes(ws_formula, ws_values):
    lines = [f"## Sheet: {ws_formula.title}", ""]
    lines.append(f"- Used bounds: {used_bounds(ws_formula)}")
    lines.append(f"- Header candidates: {detect_header_candidates(ws_formula)}")
    merged = [str(rng) for rng in list(ws_formula.merged_cells.ranges)[:20]]
    if merged:
        lines.append("- Merged ranges: " + "; ".join(merged))
    formulas = formula_samples(ws_formula)
    if formulas:
        lines.append("- Formula samples:")
        for item in formulas[:20]:
            lines.append(f"  - {item['cell']}: {item['formula']}")
    lines.append("- Non-empty row samples:")
    for r, vals in nonempty_rows(ws_formula):
        lines.append(f"  - row {r}: " + "; ".join(vals))
    lines.append("- Value row samples:")
    for r, vals in nonempty_rows(ws_values):
        lines.append(f"  - row {r}: " + "; ".join(vals))
    return "\n".join(lines)


def main():
    wb_formula = load_workbook(PATH, data_only=False)
    wb_values = load_workbook(PATH, data_only=True)
    summary = {
        "path": str(PATH),
        "sheets": [
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "used_bounds": used_bounds(ws),
                "header_candidates": detect_header_candidates(ws),
                "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:20]],
                "formula_samples": formula_samples(ws),
            }
            for ws in wb_formula.worksheets
        ],
    }
    JSON_OUT.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

    sections = ["# Manual Journals Template Notes", ""]
    sections.append("## Workbook Overview")
    for sheet in summary["sheets"]:
        sections.append(
            f"- {sheet['name']}: rows={sheet['max_row']}, cols={sheet['max_col']}, "
            f"used={sheet['used_bounds']}, formulas_sampled={len(sheet['formula_samples'])}"
        )
    sections.append("")
    for ws in wb_formula.worksheets:
        sections.append(sheet_notes(ws, wb_values[ws.title]))
        sections.append("")
    MD_OUT.write_text("\n".join(sections), encoding="utf-8")
    print(JSON_OUT)
    print(MD_OUT)


if __name__ == "__main__":
    main()
