from pathlib import Path

from openpyxl import load_workbook


def show(path):
    workbook = load_workbook(path, data_only=False)
    print("====", path.name, workbook.sheetnames)
    for sheet in workbook.worksheets:
        print("SHEET", sheet.title, "rows", sheet.max_row, "cols", sheet.max_column)
        for row_idx in range(1, min(sheet.max_row, 8) + 1):
            print(row_idx, [sheet.cell(row_idx, col).value for col in range(1, 8)])
        print("...")
        for row_idx in range(max(1, sheet.max_row - 5), sheet.max_row + 1):
            print(row_idx, [sheet.cell(row_idx, col).value for col in range(1, 8)])
        print("formats", sheet["D2"].number_format, sheet["E2"].number_format, sheet["F2"].number_format, sheet["G2"].number_format)


show(Path("_reference_combine") / "HSBC 2501 - 2024 (手動合并).xlsx")
for name in ["HSBC 2501.xlsx", "HSBC 2502.xlsx", "HSBC 2503.xlsx", "HSBC 2504.xlsx"]:
    show(Path("_stress_hsbc_w013") / "output" / name)
