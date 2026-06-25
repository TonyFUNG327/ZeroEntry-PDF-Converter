from pathlib import Path

from openpyxl import load_workbook


combined_dir = Path("_stress_hsb_h001") / "combined"
workbook_path = next(combined_dir.glob("*.xlsx"))
workbook = load_workbook(workbook_path, data_only=False)
print("workbook", workbook_path.name)

issues = []
for sheet in workbook.worksheets:
    running = None
    mismatches = []
    bad_formats = []
    dates = []
    for row_idx in range(2, sheet.max_row + 1):
        date_value = sheet.cell(row_idx, 2).value
        deposit = sheet.cell(row_idx, 4).value or 0
        withdrawal = sheet.cell(row_idx, 5).value or 0
        balance = sheet.cell(row_idx, 6).value
        if date_value:
            dates.append(date_value)
        if sheet.cell(row_idx, 5).number_format != '(#,##0.00);(#,##0.00);"-"':
            bad_formats.append((row_idx, sheet.cell(row_idx, 5).number_format))
        if deposit == 0 and withdrawal == 0 and balance is not None:
            running = balance
        else:
            if running is None:
                running = balance or 0
            running += deposit
            running -= withdrawal
        if balance is not None and abs(balance - running) > 0.01:
            mismatches.append((row_idx, balance, round(running, 2)))
    print("====", sheet.title)
    print("rows", sheet.max_row - 1)
    print("date_range", min(dates), max(dates))
    print("first", [sheet.cell(2, col).value for col in range(1, 7)])
    print("last", [sheet.cell(sheet.max_row, col).value for col in range(1, 7)])
    print("mismatches", mismatches[:10], "bad_formats", bad_formats[:10])
    if mismatches or bad_formats:
        issues.append((sheet.title, mismatches[:10], bad_formats[:10]))

print("issues", issues)
