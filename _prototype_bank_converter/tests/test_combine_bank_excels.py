import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from combine_bank_excels import collect_excel_paths, first_brought_forward_balance, is_brought_forward


def make_workbook(path, description):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Account"
    sheet.append(["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"])
    sheet.append(["A", datetime(2025, 1, 1), description, None, None, 100.0, 100.0])
    workbook.save(path)


class TestCombineBankExcels(unittest.TestCase):
    def test_is_brought_forward_aliases(self):
        self.assertTrue(is_brought_forward(["A", None, "B/F BALANCE"]))
        self.assertTrue(is_brought_forward(["A", None, "BAL B/F"]))
        self.assertTrue(is_brought_forward(["A", None, "BALANCE BROUGHT FORWARD"]))
        self.assertFalse(is_brought_forward(["A", None, "Balance Carried Forward"]))

    def test_collect_excel_paths_ignores_excel_temp_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            normal = tmp_path / "normal.xlsx"
            temp = tmp_path / "~$normal.xlsx"
            csv = tmp_path / "normal.csv"
            make_workbook(normal, "B/F BALANCE")
            temp.write_text("temp")
            csv.write_text("csv")

            self.assertEqual(collect_excel_paths(tmp_path), [normal])

    def test_first_brought_forward_balance_supports_aliases(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            for description in ["B/F BALANCE", "BAL B/F", "BALANCE BROUGHT FORWARD"]:
                path = tmp_path / f"{description.replace('/', '_')}.xlsx"
                make_workbook(path, description)
                workbook = load_workbook(path)
                self.assertEqual(first_brought_forward_balance(workbook.active), 100.0)
                workbook.close()


if __name__ == "__main__":
    unittest.main()
