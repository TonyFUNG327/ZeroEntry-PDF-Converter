from pathlib import Path

from openpyxl import load_workbook


def main():
    issues = []
    for path in sorted(Path("_reference_ncb/output").glob("*.xlsx")):
        workbook = load_workbook(path, data_only=False)
        print("====", path.name, workbook.sheetnames)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            deposits = 0.0
            withdrawals = 0.0
            control = None
            mismatches = 0
            rows = []
            for row_idx in range(2, sheet.max_row + 1):
                date = sheet.cell(row_idx, 2).value
                desc = sheet.cell(row_idx, 3).value
                deposit = sheet.cell(row_idx, 4).value
                withdrawal = sheet.cell(row_idx, 5).value
                balance = sheet.cell(row_idx, 6).value
                withdrawal_format = sheet.cell(row_idx, 5).number_format
                rows.append((date, desc, deposit, withdrawal, balance))
                if withdrawal and "(" not in withdrawal_format:
                    issues.append((path.name, sheet_name, row_idx, "bad withdrawal format", withdrawal_format))
                if desc and any(marker in str(desc) for marker in ["重要事項", "請查看", "南商提醒", "地址："]):
                    issues.append((path.name, sheet_name, row_idx, "footer leaked into description", desc))
                if deposit is None and withdrawal is None:
                    control = balance
                else:
                    control = round((control or 0) + (deposit or 0) - (withdrawal or 0), 2)
                if round(balance or 0, 2) != round(control or 0, 2):
                    mismatches += 1
                deposits += deposit or 0
                withdrawals += withdrawal or 0
            print(sheet_name, "rows", len(rows), "deposit", round(deposits, 2), "withdrawal", round(withdrawals, 2), "mismatches", mismatches)
            for row in rows:
                print(" ", row)
            if mismatches:
                issues.append((path.name, sheet_name, "mismatches", mismatches))
    print("ISSUES", len(issues))
    for issue in issues:
        print(issue)
    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
