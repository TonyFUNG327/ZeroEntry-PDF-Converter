from pathlib import Path

from openpyxl import load_workbook


EXPECTED_SHEETS = {
    "NCB HKD Savings Account",
    "NCB HKD MCY Savings Account",
    "NCB HKD Current Account",
}


def main():
    issues = []
    for path in sorted(Path("_reference_ncb_integrated/output").glob("*.xlsx")):
        workbook = load_workbook(path, data_only=False)
        print("====", path.name, workbook.sheetnames)
        missing = EXPECTED_SHEETS.difference(workbook.sheetnames)
        if missing:
            issues.append((path.name, "missing sheets", sorted(missing)))
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            deposits = 0.0
            withdrawals = 0.0
            control = None
            mismatches = 0
            bad_desc = 0
            bad_format = 0
            for row_idx in range(2, sheet.max_row + 1):
                desc = str(sheet.cell(row_idx, 3).value or "")
                deposit = sheet.cell(row_idx, 4).value
                withdrawal = sheet.cell(row_idx, 5).value
                balance = sheet.cell(row_idx, 6).value
                withdrawal_format = sheet.cell(row_idx, 5).number_format
                if withdrawal and "(" not in withdrawal_format:
                    bad_format += 1
                if any(marker in desc for marker in ["Important Notes", "Address:", "Please examine", "Customer Service"]):
                    bad_desc += 1
                if deposit is None and withdrawal is None:
                    control = balance
                else:
                    control = round((control or 0) + (deposit or 0) - (withdrawal or 0), 2)
                if round(balance or 0, 2) != round(control or 0, 2):
                    mismatches += 1
                deposits += deposit or 0
                withdrawals += withdrawal or 0
            print(
                sheet_name,
                "rows", sheet.max_row - 1,
                "deposit", round(deposits, 2),
                "withdrawal", round(withdrawals, 2),
                "mismatches", mismatches,
                "bad_desc", bad_desc,
                "bad_format", bad_format,
            )
            if mismatches or bad_desc or bad_format:
                issues.append((path.name, sheet_name, mismatches, bad_desc, bad_format))
    print("ISSUES", len(issues))
    for issue in issues:
        print(issue)
    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
