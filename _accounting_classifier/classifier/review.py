from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .engine import OUTPUT_COLUMNS
from .io import write_json, write_workbook
from .rules import text


MANUAL_REVIEW_COLUMNS = [
    "Manual_Category",
    "Manual_Account_Code",
    "Manual_Account_Name",
    "Manual_Tax_Type",
    "Manual_Counterparty",
    "Manual_Review_Status",
    "Manual_Notes",
]
REVIEW_COLUMNS = OUTPUT_COLUMNS + MANUAL_REVIEW_COLUMNS
VALID_REVIEW_STATUSES = {"Pending", "Confirmed", "Corrected", "Ignore", "Need_Advice"}
STATUS_ALIASES = {
    "pending": "Pending",
    "confirmed": "Confirmed",
    "corrected": "Corrected",
    "ignore": "Ignore",
    "ignored": "Ignore",
    "need_advice": "Need_Advice",
    "need advice": "Need_Advice",
}


def read_review_rows(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if path.suffix.casefold() == ".csv":
        return read_review_csv(path)
    if path.suffix.casefold() == ".xlsx":
        return read_review_xlsx(path)
    raise ValueError(f"Unsupported review file type: {path.suffix}")


def read_review_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return normalize_rows(reader.fieldnames or [], list(reader))


def read_review_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook.active
        headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
        header_names = [str(header) if header is not None else "" for header in headers]
        rows = []
        for row_idx in range(2, sheet.max_row + 1):
            row = {
                header_names[col_idx - 1]: sheet.cell(row_idx, col_idx).value
                for col_idx in range(1, len(header_names) + 1)
                if header_names[col_idx - 1]
            }
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        return normalize_rows(header_names, rows)
    finally:
        workbook.close()


def normalize_rows(headers: list[Any], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    header_set = {str(header) for header in headers if header}
    validate_classified_headers(header_set)
    return [
        {column: row.get(column, "") for column in REVIEW_COLUMNS if column in header_set or column in OUTPUT_COLUMNS}
        for row in rows
    ]


def validate_classified_headers(headers: set[str]) -> None:
    missing = [column for column in OUTPUT_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Reviewed input missing A.1.2 output columns: {', '.join(missing)}")


def validate_review_headers(headers: set[str]) -> None:
    validate_classified_headers(headers)
    missing = [column for column in MANUAL_REVIEW_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Reviewed input missing manual review columns: {', '.join(missing)}")


def add_manual_review_columns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for row in rows:
        normalized.append({**{column: row.get(column, "") for column in OUTPUT_COLUMNS}, **{column: "" for column in MANUAL_REVIEW_COLUMNS}})
    return normalized


def append_note(existing: Any, manual_note: Any) -> str:
    base = text(existing)
    note = text(manual_note)
    if not note:
        return base
    if not base:
        return note
    return f"{base}; {note}"


def normalize_review_status(value: Any, row_number: int) -> str:
    status = text(value)
    if not status:
        return ""
    normalized = STATUS_ALIASES.get(status.casefold())
    if normalized:
        return normalized
    raise ValueError(f"Review row {row_number}: invalid Manual_Review_Status '{status}'")


def validate_review_row(row: dict[str, Any], row_number: int) -> str:
    status = normalize_review_status(row.get("Manual_Review_Status"), row_number)
    if not status:
        return ""
    if status == "Corrected" and not text(row.get("Manual_Category")):
        raise ValueError(f"Review row {row_number}: Manual_Category is required for Corrected")
    return status


def manual_or_existing(row: dict[str, Any], manual_column: str, output_column: str) -> Any:
    manual_value = text(row.get(manual_column))
    return manual_value if manual_value else row.get(output_column, "")


def apply_review_row(row: dict[str, Any], row_number: int) -> dict[str, Any]:
    status = validate_review_row(row, row_number)
    reviewed = {column: row.get(column, "") for column in REVIEW_COLUMNS}
    reviewed["Manual_Review_Status"] = status

    if not status:
        return reviewed
    if status == "Pending":
        reviewed["Review_Needed"] = "Yes"
        reviewed["Notes"] = append_note(reviewed.get("Notes"), "manual review pending")
        return reviewed
    if status == "Confirmed":
        reviewed["Review_Needed"] = "No"
        reviewed["Classification_Source"] = "manual_confirmed"
        reviewed["Notes"] = append_note(reviewed.get("Notes"), reviewed.get("Manual_Notes"))
        return reviewed
    if status == "Corrected":
        reviewed["Category"] = text(reviewed.get("Manual_Category"))
        reviewed["Account_Code"] = manual_or_existing(reviewed, "Manual_Account_Code", "Account_Code")
        reviewed["Account_Name"] = manual_or_existing(reviewed, "Manual_Account_Name", "Account_Name")
        reviewed["Tax_Type"] = manual_or_existing(reviewed, "Manual_Tax_Type", "Tax_Type")
        reviewed["Counterparty"] = manual_or_existing(reviewed, "Manual_Counterparty", "Counterparty")
        reviewed["Review_Needed"] = "No"
        reviewed["Classification_Source"] = "manual_corrected"
        reviewed["Confidence"] = 1.0
        reviewed["Notes"] = append_note(reviewed.get("Notes"), reviewed.get("Manual_Notes"))
        return reviewed
    if status == "Ignore":
        reviewed["Category"] = "Ignored"
        reviewed["Review_Needed"] = "No"
        reviewed["Classification_Source"] = "manual_ignored"
        reviewed["Confidence"] = 1.0
        reviewed["Notes"] = append_note(reviewed.get("Notes"), reviewed.get("Manual_Notes"))
        return reviewed
    if status == "Need_Advice":
        reviewed["Review_Needed"] = "Yes"
        reviewed["Classification_Source"] = "manual_need_advice"
        reviewed["Notes"] = append_note(reviewed.get("Notes"), reviewed.get("Manual_Notes"))
        return reviewed
    return reviewed


def apply_manual_review(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if rows:
        validate_review_headers(set(rows[0].keys()))
    return [apply_review_row(row, idx) for idx, row in enumerate(rows, start=2)]


def summarize_review(rows: list[dict[str, Any]]) -> dict[str, Any]:
    status_counts = Counter(text(row.get("Manual_Review_Status")) for row in rows)
    source_counts = Counter(text(row.get("Classification_Source")) or "unknown" for row in rows)
    category_counts = Counter(text(row.get("Category")) or "Unclassified" for row in rows)
    corrected_category_counts = Counter(
        text(row.get("Category")) or "Unclassified"
        for row in rows
        if text(row.get("Manual_Review_Status")) == "Corrected"
    )
    need_advice = [
        text(row.get("Description")) or "(blank)"
        for row in rows
        if text(row.get("Manual_Review_Status")) == "Need_Advice"
    ]

    completed = sum(
        status_counts.get(status, 0)
        for status in ["Confirmed", "Corrected", "Ignore"]
    )
    actioned = completed + status_counts.get("Need_Advice", 0)
    review_needed = sum(1 for row in rows if text(row.get("Review_Needed")).casefold() == "yes")
    return {
        "transaction_count": len(rows),
        "manual_pending_count": status_counts.get("Pending", 0),
        "manual_confirmed_count": status_counts.get("Confirmed", 0),
        "manual_corrected_count": status_counts.get("Corrected", 0),
        "manual_ignored_count": status_counts.get("Ignore", 0),
        "manual_need_advice_count": status_counts.get("Need_Advice", 0),
        "manual_blank_status_count": status_counts.get("", 0),
        "manual_actioned_count": actioned,
        "review_completed_count": completed,
        "review_completed_ratio": round(completed / len(rows), 4) if rows else 0,
        "review_needed_count": review_needed,
        "review_needed_ratio": round(review_needed / len(rows), 4) if rows else 0,
        "classification_source_counts": dict(sorted(source_counts.items())),
        "category_counts": dict(sorted(category_counts.items())),
        "corrected_category_counts": dict(sorted(corrected_category_counts.items())),
        "need_advice_descriptions": need_advice,
        "ignored_count": status_counts.get("Ignore", 0),
    }


def write_review_summary_text(path: str | Path, summary: dict[str, Any]) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"transaction_count: {summary['transaction_count']}",
        f"manual_pending_count: {summary['manual_pending_count']}",
        f"manual_confirmed_count: {summary['manual_confirmed_count']}",
        f"manual_corrected_count: {summary['manual_corrected_count']}",
        f"manual_ignored_count: {summary['manual_ignored_count']}",
        f"manual_need_advice_count: {summary['manual_need_advice_count']}",
        f"manual_blank_status_count: {summary['manual_blank_status_count']}",
        f"manual_actioned_count: {summary['manual_actioned_count']}",
        f"review_completed_count: {summary['review_completed_count']}",
        f"review_completed_ratio: {summary['review_completed_ratio']:.4f}",
        f"review_needed_count: {summary['review_needed_count']}",
        f"review_needed_ratio: {summary['review_needed_ratio']:.4f}",
        "",
        "classification_source_counts:",
    ]
    lines.extend(f"- {key}: {value}" for key, value in summary["classification_source_counts"].items())
    lines.append("")
    lines.append("category_counts:")
    lines.extend(f"- {key}: {value}" for key, value in summary["category_counts"].items())
    lines.append("")
    lines.append("corrected_category_counts:")
    lines.extend(f"- {key}: {value}" for key, value in summary["corrected_category_counts"].items())
    lines.append("")
    lines.append("need_advice_descriptions:")
    lines.extend(f"- {description}" for description in summary["need_advice_descriptions"])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_review_outputs(output_path: str | Path, rows: list[dict[str, Any]], summary_json: str | Path, summary_txt: str | Path) -> dict[str, Path | dict]:
    output_path = write_workbook(output_path, rows, REVIEW_COLUMNS)
    summary = summarize_review(rows)
    json_path = write_json(summary_json, summary)
    text_path = write_review_summary_text(summary_txt, summary)
    return {"output": output_path, "summary_json": json_path, "summary_txt": text_path, "summary": summary}
