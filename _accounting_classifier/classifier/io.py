from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .engine import INPUT_COLUMNS, OUTPUT_COLUMNS


def json_default(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def read_transactions(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if path.suffix.casefold() == ".csv":
        return read_csv(path)
    if path.suffix.casefold() == ".xlsx":
        return read_xlsx(path)
    raise ValueError(f"Unsupported input file type: {path.suffix}")


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        validate_input_headers(reader.fieldnames or [])
        return [{column: row.get(column, "") for column in INPUT_COLUMNS} for row in reader]


def read_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            if sheet.max_row < 2:
                continue
            headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
            if not set(INPUT_COLUMNS).issubset({str(header) for header in headers if header is not None}):
                continue
            index = {str(header): col for col, header in enumerate(headers, start=1) if header is not None}
            for row_idx in range(2, sheet.max_row + 1):
                row = {column: sheet.cell(row_idx, index[column]).value for column in INPUT_COLUMNS}
                if any(value not in (None, "") for value in row.values()):
                    rows.append(row)
    finally:
        workbook.close()
    if not rows:
        raise ValueError(f"No transaction rows with required headers found in: {path}")
    return rows


def validate_input_headers(headers: list[Any]) -> None:
    header_set = {str(header) for header in headers}
    missing = [column for column in INPUT_COLUMNS if column not in header_set]
    if missing:
        raise ValueError(f"Input missing columns: {', '.join(missing)}")


def write_workbook(path: str | Path, rows: list[dict[str, Any]], columns: list[str] | None = None) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = columns or OUTPUT_COLUMNS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Classified"
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    sheet.freeze_panes = "A2"
    for idx, column in enumerate(columns, start=1):
        sheet.column_dimensions[sheet.cell(1, idx).column_letter].width = min(max(len(column) + 2, 12), 28)
    workbook.save(path)
    return path


def write_json(path: str | Path, payload: dict[str, Any]) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=json_default), encoding="utf-8")
    return path


def write_summary_text(path: str | Path, summary: dict[str, Any]) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"transaction_count: {summary['transaction_count']}",
        f"classified_count: {summary['classified_count']}",
        f"unclassified_count: {summary['unclassified_count']}",
        f"review_needed_count: {summary['review_needed_count']}",
        f"unclassified_ratio: {summary['unclassified_ratio']:.4f}",
        "",
        "category_counts:",
    ]
    lines.extend(f"- {key}: {value}" for key, value in summary["category_counts"].items())
    lines.append("")
    lines.append("source_counts:")
    lines.extend(f"- {key}: {value}" for key, value in summary["source_counts"].items())
    lines.append("")
    lines.append("rule_hit_counts:")
    lines.extend(f"- {key}: {value}" for key, value in summary["rule_hit_counts"].items())
    lines.append("")
    lines.append("anomaly_counts:")
    lines.extend(f"- {key}: {value}" for key, value in summary["anomaly_counts"].items())
    lines.append("")
    lines.append("top_unclassified_descriptions:")
    lines.extend(
        f"- {item['description']}: {item['count']}"
        for item in summary["top_unclassified_descriptions"]
    )
    lines.append("")
    lines.append("direction_amounts:")
    lines.extend(f"- {key}: {value:,.2f}" for key, value in summary["direction_amounts"].items())
    lines.append("")
    lines.append("category_amounts:")
    lines.extend(f"- {key}: {value:,.2f}" for key, value in summary["category_amounts"].items())
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path
