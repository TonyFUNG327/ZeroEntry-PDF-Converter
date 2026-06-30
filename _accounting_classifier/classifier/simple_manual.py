from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .engine import number
from .mappings import extract_confirmed_mappings
from .review import REVIEW_COLUMNS, normalize_review_status
from .rules import text


SIMPLE_MANUAL_COLUMNS = [
    "Bank_Account",
    "Date",
    "Description",
    "Deposit",
    "Withdrawal",
    "Manual_Category",
    "Manual_Account_Code",
    "Manual_Account_Name",
    "Manual_Tax_Type",
    "Manual_Counterparty",
    "Manual_Review_Status",
    "Manual_Notes",
]


def validate_simple_headers(headers: list[Any]) -> None:
    normalized = [str(header) if header is not None else "" for header in headers]
    if normalized != SIMPLE_MANUAL_COLUMNS:
        raise ValueError("Invalid simple manual template header")


def read_simple_manual_rows(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if path.suffix.casefold() == ".csv":
        return read_simple_manual_csv(path)
    if path.suffix.casefold() == ".xlsx":
        return read_simple_manual_xlsx(path)
    raise ValueError(f"Unsupported simple manual template file type: {path.suffix}")


def read_simple_manual_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        validate_simple_headers(reader.fieldnames or [])
        return [{column: row.get(column, "") for column in SIMPLE_MANUAL_COLUMNS} for row in reader]


def read_simple_manual_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook.active
        headers = [sheet.cell(1, col).value for col in range(1, len(SIMPLE_MANUAL_COLUMNS) + 1)]
        validate_simple_headers(headers)
        rows = []
        for row_idx in range(2, sheet.max_row + 1):
            row = {
                column: sheet.cell(row_idx, col_idx).value
                for col_idx, column in enumerate(SIMPLE_MANUAL_COLUMNS, start=1)
            }
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def infer_direction(row: dict[str, Any]) -> str:
    deposit = number(row.get("Deposit"))
    withdrawal = number(row.get("Withdrawal"))
    if deposit != 0 and withdrawal == 0:
        return "Deposit"
    if withdrawal != 0 and deposit == 0:
        return "Withdrawal"
    return "Unknown"


def infer_amount(row: dict[str, Any], direction: str) -> float:
    if direction == "Deposit":
        return abs(number(row.get("Deposit")))
    if direction == "Withdrawal":
        return abs(number(row.get("Withdrawal")))
    return 0.0


def simple_manual_rows_to_reviewed_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reviewed_rows = []
    for idx, row in enumerate(rows, start=2):
        direction = infer_direction(row)
        status = normalize_review_status(row.get("Manual_Review_Status"), idx)
        reviewed = {
            "Bank_Account": row.get("Bank_Account", ""),
            "Date": row.get("Date", ""),
            "Description": row.get("Description", ""),
            "Deposit": row.get("Deposit", ""),
            "Withdrawal": row.get("Withdrawal", ""),
            "Balance": "",
            "Control": "",
            "Direction": direction,
            "Amount": infer_amount(row, direction),
            "Category": row.get("Manual_Category", ""),
            "Account_Code": row.get("Manual_Account_Code", ""),
            "Account_Name": row.get("Manual_Account_Name", ""),
            "Tax_Type": row.get("Manual_Tax_Type", ""),
            "Counterparty": row.get("Manual_Counterparty", ""),
            "Confidence": 1.0,
            "Rule_ID": "",
            "Review_Needed": "No" if status in {"Confirmed", "Corrected"} else "Yes",
            "Classification_Source": "manual_simple",
            "Notes": row.get("Manual_Notes", ""),
            "Manual_Category": row.get("Manual_Category", ""),
            "Manual_Account_Code": row.get("Manual_Account_Code", ""),
            "Manual_Account_Name": row.get("Manual_Account_Name", ""),
            "Manual_Tax_Type": row.get("Manual_Tax_Type", ""),
            "Manual_Counterparty": row.get("Manual_Counterparty", ""),
            "Manual_Review_Status": status,
            "Manual_Notes": row.get("Manual_Notes", ""),
        }
        reviewed_rows.append({column: reviewed.get(column, "") for column in REVIEW_COLUMNS})
    return reviewed_rows


def extract_mappings_from_simple_manual_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return extract_confirmed_mappings(simple_manual_rows_to_reviewed_rows(rows))
