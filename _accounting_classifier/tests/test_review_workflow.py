import csv
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from classifier.engine import OUTPUT_COLUMNS
from classifier.io import write_workbook
from classifier.review import (
    MANUAL_REVIEW_COLUMNS,
    REVIEW_COLUMNS,
    add_manual_review_columns,
    apply_manual_review,
    read_review_rows,
    summarize_review,
)
from review_bank_transactions import build_arg_parser, run


FIXTURES = APP_ROOT / "tests" / "fixtures"


def read_csv_rows(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


class TestManualReviewWorkflowA20(unittest.TestCase):
    def test_add_manual_review_columns(self):
        row = {column: "" for column in OUTPUT_COLUMNS}
        row.update({"Description": "Synthetic row", "Category": "Unclassified"})
        templated = add_manual_review_columns([row])
        self.assertEqual(list(templated[0].keys()), REVIEW_COLUMNS)
        for column in MANUAL_REVIEW_COLUMNS:
            self.assertEqual(templated[0][column], "")

    def test_validate_required_a12_output_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bad.csv"
            path.write_text("Description,Manual_Review_Status\nSynthetic,Pending\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "missing A.1.2 output columns"):
                read_review_rows(path)

    def test_invalid_manual_review_status_raises_value_error(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[0]["Manual_Review_Status"] = "Done"
        with self.assertRaisesRegex(ValueError, "invalid Manual_Review_Status"):
            apply_manual_review(rows)

    def test_lowercase_status_normalized(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[1]["Manual_Review_Status"] = "confirmed"
        reviewed = apply_manual_review(rows)
        self.assertEqual(reviewed[1]["Manual_Review_Status"], "Confirmed")
        self.assertEqual(reviewed[1]["Classification_Source"], "manual_confirmed")

    def test_uppercase_status_normalized(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[2]["Manual_Review_Status"] = "CORRECTED"
        reviewed = apply_manual_review(rows)
        self.assertEqual(reviewed[2]["Manual_Review_Status"], "Corrected")
        self.assertEqual(reviewed[2]["Classification_Source"], "manual_corrected")

    def test_need_advice_variants_normalized(self):
        for value in ["Need Advice", "need advice", "NEED_ADVICE"]:
            rows = read_review_rows(FIXTURES / "review_input.csv")
            rows[4]["Manual_Review_Status"] = value
            reviewed = apply_manual_review(rows)
            self.assertEqual(reviewed[4]["Manual_Review_Status"], "Need_Advice")
            self.assertEqual(reviewed[4]["Classification_Source"], "manual_need_advice")

    def test_ignored_normalized_to_ignore(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[3]["Manual_Review_Status"] = "ignored"
        reviewed = apply_manual_review(rows)
        self.assertEqual(reviewed[3]["Manual_Review_Status"], "Ignore")
        self.assertEqual(reviewed[3]["Classification_Source"], "manual_ignored")

    def test_pending_behavior(self):
        reviewed = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))
        row = reviewed[0]
        self.assertEqual(row["Category"], "Unclassified")
        self.assertEqual(row["Review_Needed"], "Yes")
        self.assertEqual(row["Classification_Source"], "unclassified")
        self.assertIn("manual review pending", row["Notes"])

    def test_confirmed_behavior(self):
        row = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))[1]
        self.assertEqual(row["Category"], "Bank Charges")
        self.assertEqual(row["Review_Needed"], "No")
        self.assertEqual(row["Classification_Source"], "manual_confirmed")

    def test_corrected_behavior(self):
        row = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))[2]
        self.assertEqual(row["Category"], "Rent Expense")
        self.assertEqual(row["Account_Code"], "6100")
        self.assertEqual(row["Account_Name"], "Rent Expense")
        self.assertEqual(row["Tax_Type"], "Review")
        self.assertEqual(row["Counterparty"], "Synthetic Landlord")
        self.assertEqual(row["Confidence"], 1.0)
        self.assertEqual(row["Review_Needed"], "No")
        self.assertEqual(row["Classification_Source"], "manual_corrected")

    def test_corrected_missing_manual_category_raises_value_error(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[2]["Manual_Category"] = ""
        with self.assertRaisesRegex(ValueError, "Manual_Category"):
            apply_manual_review(rows)

    def test_manual_notes_optional_for_final_statuses(self):
        expected_sources = {
            1: "manual_confirmed",
            2: "manual_corrected",
            3: "manual_ignored",
            4: "manual_need_advice",
        }
        for index, source in expected_sources.items():
            rows = read_review_rows(FIXTURES / "review_input.csv")
            rows[index]["Manual_Notes"] = ""
            reviewed = apply_manual_review(rows)
            self.assertEqual(reviewed[index]["Classification_Source"], source)

    def test_corrected_blank_manual_account_code_preserves_original(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[2]["Account_Code"] = "6999"
        rows[2]["Manual_Account_Code"] = ""
        reviewed = apply_manual_review(rows)
        self.assertEqual(reviewed[2]["Account_Code"], "6999")

    def test_corrected_blank_manual_account_name_preserves_original(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[2]["Account_Name"] = "Original Expense"
        rows[2]["Manual_Account_Name"] = ""
        reviewed = apply_manual_review(rows)
        self.assertEqual(reviewed[2]["Account_Name"], "Original Expense")

    def test_corrected_blank_manual_tax_type_preserves_original(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[2]["Tax_Type"] = "Original Tax"
        rows[2]["Manual_Tax_Type"] = ""
        reviewed = apply_manual_review(rows)
        self.assertEqual(reviewed[2]["Tax_Type"], "Original Tax")

    def test_corrected_blank_manual_counterparty_preserves_original(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        rows[2]["Counterparty"] = "Original Counterparty"
        rows[2]["Manual_Counterparty"] = ""
        reviewed = apply_manual_review(rows)
        self.assertEqual(reviewed[2]["Counterparty"], "Original Counterparty")

    def test_ignore_behavior(self):
        row = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))[3]
        self.assertEqual(row["Category"], "Ignored")
        self.assertEqual(row["Review_Needed"], "No")
        self.assertEqual(row["Classification_Source"], "manual_ignored")
        self.assertEqual(row["Confidence"], 1.0)

    def test_need_advice_behavior(self):
        row = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))[4]
        self.assertEqual(row["Category"], "Sales Receipt")
        self.assertEqual(row["Review_Needed"], "Yes")
        self.assertEqual(row["Classification_Source"], "manual_need_advice")

    def test_manual_notes_appended_to_notes(self):
        reviewed = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))
        self.assertIn("Reviewer corrected classification", reviewed[2]["Notes"])
        self.assertTrue(reviewed[2]["Notes"].startswith("No matching enabled rule;"))

    def test_review_summary_counts(self):
        reviewed = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))
        summary = summarize_review(reviewed)
        self.assertEqual(summary["transaction_count"], 6)
        self.assertEqual(summary["manual_pending_count"], 1)
        self.assertEqual(summary["manual_confirmed_count"], 1)
        self.assertEqual(summary["manual_corrected_count"], 1)
        self.assertEqual(summary["manual_ignored_count"], 1)
        self.assertEqual(summary["manual_need_advice_count"], 1)
        self.assertEqual(summary["manual_blank_status_count"], 1)
        self.assertEqual(summary["manual_actioned_count"], 4)
        self.assertEqual(summary["review_completed_count"], 3)
        self.assertEqual(summary["review_completed_ratio"], 0.5)
        self.assertEqual(summary["review_needed_count"], 2)
        self.assertEqual(summary["review_needed_ratio"], 0.3333)
        self.assertEqual(summary["ignored_count"], 1)
        self.assertEqual(summary["corrected_category_counts"], {"Rent Expense": 1})
        self.assertEqual(summary["need_advice_descriptions"], ["Needs advice receipt"])

    def test_review_summary_text_contains_a21_ratio_fields(self):
        from classifier.review import write_review_summary_text

        reviewed = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))
        summary = summarize_review(reviewed)
        with tempfile.TemporaryDirectory() as tmp:
            path = write_review_summary_text(Path(tmp) / "review.summary.txt", summary)
            content = path.read_text(encoding="utf-8")
        self.assertIn("manual_actioned_count:", content)
        self.assertIn("review_completed_ratio:", content)
        self.assertIn("review_needed_ratio:", content)

    def test_reviewed_output_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = write_workbook(Path(tmp) / "reviewed.xlsx", [], REVIEW_COLUMNS)
            workbook = load_workbook(output)
            try:
                headers = [workbook.active.cell(1, col).value for col in range(1, workbook.active.max_column + 1)]
            finally:
                workbook.close()
            self.assertEqual(headers, REVIEW_COLUMNS)

    def test_baseline_review_expected_results(self):
        reviewed = apply_manual_review(read_review_rows(FIXTURES / "review_input.csv"))
        expected = read_csv_rows(FIXTURES / "review_expected.csv")
        actual = [
            {
                "Description": row["Description"],
                "Category": row["Category"],
                "Account_Code": str(row["Account_Code"]),
                "Account_Name": row["Account_Name"],
                "Tax_Type": row["Tax_Type"],
                "Counterparty": row["Counterparty"],
                "Confidence": str(row["Confidence"]),
                "Review_Needed": row["Review_Needed"],
                "Classification_Source": row["Classification_Source"],
                "Notes": row["Notes"],
            }
            for row in reviewed
        ]
        self.assertEqual(actual, expected)

    def test_cli_add_template_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "classified.xlsx"
            row = {column: "" for column in OUTPUT_COLUMNS}
            row.update({"Description": "Synthetic row", "Category": "Unclassified"})
            write_workbook(input_path, [row], OUTPUT_COLUMNS)
            args = build_arg_parser().parse_args([
                str(input_path),
                "--output",
                str(Path(tmp) / "template.xlsx"),
                "--add-template-columns",
            ])
            result = run(args)
            self.assertTrue(result["output"].exists())
            self.assertEqual(result["summary"]["transaction_count"], 1)

    def test_cli_apply_review(self):
        with tempfile.TemporaryDirectory() as tmp:
            args = build_arg_parser().parse_args([
                str(FIXTURES / "review_input.csv"),
                "--output",
                str(Path(tmp) / "reviewed.xlsx"),
            ])
            result = run(args)
            self.assertTrue(result["output"].exists())
            self.assertTrue(result["summary_json"].exists())
            self.assertTrue(result["summary_txt"].exists())
            payload = json.loads(result["summary_json"].read_text(encoding="utf-8"))
            self.assertEqual(payload["manual_corrected_count"], 1)

    def test_csv_reviewed_input(self):
        rows = read_review_rows(FIXTURES / "review_input.csv")
        self.assertEqual(len(rows), 6)
        self.assertEqual(rows[0]["Manual_Review_Status"], "Pending")

    def test_xlsx_reviewed_input_where_practical(self):
        with tempfile.TemporaryDirectory() as tmp:
            rows = read_review_rows(FIXTURES / "review_input.csv")
            path = write_workbook(Path(tmp) / "review.xlsx", rows, REVIEW_COLUMNS)
            read_back = read_review_rows(path)
            self.assertEqual(len(read_back), 6)
            reviewed = apply_manual_review(read_back)
            self.assertEqual(reviewed[2]["Classification_Source"], "manual_corrected")


if __name__ == "__main__":
    unittest.main()
