import csv
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from classifier.engine import (
    OUTPUT_COLUMNS,
    classify_transactions,
    detect_direction,
    summarize_classification,
    unclassified_rows,
)
from classifier.io import read_transactions, write_summary_text, write_workbook
from classifier.rules import ClassificationRule, RULE_FIELDS, load_rules
from classify_bank_transactions import DEFAULT_RULES, build_arg_parser, run


FIXTURES = APP_ROOT / "tests" / "fixtures"


def write_rules(path, rows, fieldnames=None):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames or RULE_FIELDS)
        writer.writeheader()
        for row in rows:
            base = {field: "" for field in (fieldnames or RULE_FIELDS)}
            base.update(row)
            writer.writerow({field: base.get(field, "") for field in (fieldnames or RULE_FIELDS)})


def read_csv_rows(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def make_rule(rule_id, keyword, **overrides):
    row = {
        "rule_id": rule_id,
        "priority": "10",
        "enabled": "Yes",
        "match_type": "contains",
        "keyword": keyword,
        "direction": "Any",
        "category": rule_id,
        "account_code": "1000",
        "account_name": rule_id.title(),
        "tax_type": "Review",
        "counterparty": "Synthetic Counterparty",
        "confidence": "0.8",
        "review_needed": "No",
        "notes": "Synthetic rule",
    }
    row.update(overrides)
    return row


class TestRuleClassifierA111(unittest.TestCase):
    def test_default_rules_csv_exists_and_loads(self):
        self.assertTrue(DEFAULT_RULES.exists())
        rules = load_rules(DEFAULT_RULES)
        self.assertEqual([rule.rule_id for rule in rules], [
            "BANK_CHARGE",
            "RENT_PAYMENT",
            "CUSTOMER_RECEIPT",
            "TRANSFER",
            "INTEREST",
        ])
        self.assertEqual(set(RULE_FIELDS), set(read_csv_rows(DEFAULT_RULES)[0].keys()))

    def test_required_rule_columns_validation(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bad_rules.csv"
            fields = [field for field in RULE_FIELDS if field != "notes"]
            write_rules(path, [make_rule("FEE", "fee")], fieldnames=fields)
            with self.assertRaisesRegex(ValueError, "notes"):
                load_rules(path)

    def test_duplicate_rule_id_raises_value_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("DUP", "one"), make_rule("DUP", "two")])
            with self.assertRaisesRegex(ValueError, "Duplicate rule_id"):
                load_rules(path)

    def test_invalid_confidence_raises_value_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("BAD", "fee", confidence="high")])
            with self.assertRaisesRegex(ValueError, "confidence must be numeric"):
                load_rules(path)

    def test_confidence_greater_than_one_raises_value_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("BAD", "fee", confidence="1.1")])
            with self.assertRaisesRegex(ValueError, "between 0 and 1"):
                load_rules(path)

    def test_confidence_less_than_zero_raises_value_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("BAD", "fee", confidence="-0.1")])
            with self.assertRaisesRegex(ValueError, "between 0 and 1"):
                load_rules(path)

    def test_amount_min_greater_than_amount_max_raises_value_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("BAD", "fee", amount_min="20", amount_max="10")])
            with self.assertRaisesRegex(ValueError, "amount_min"):
                load_rules(path)

    def test_invalid_enabled_value_raises_value_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("BAD", "fee", enabled="Maybe")])
            with self.assertRaisesRegex(ValueError, "invalid enabled"):
                load_rules(path)

    def test_contains_keyword_and_deposit_direction_matching(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("RECEIPT", "customer", direction="Deposit")])
            classified = classify_transactions(
                [{"Description": "Customer payment", "Deposit": 100, "Withdrawal": ""}],
                load_rules(path),
            )
            self.assertEqual(classified[0]["Rule_ID"], "RECEIPT")
            self.assertEqual(classified[0]["Direction"], "Deposit")

    def test_withdrawal_direction_matching(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("FEE", "fee", direction="Withdrawal")])
            classified = classify_transactions(
                [{"Description": "Monthly fee", "Deposit": "", "Withdrawal": 10}],
                load_rules(path),
            )
            self.assertEqual(classified[0]["Rule_ID"], "FEE")
            self.assertEqual(classified[0]["Direction"], "Withdrawal")

    def test_any_direction_matches_deposit_and_withdrawal_but_not_unknown(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("TRANSFER", "transfer", direction="Any")])
            classified = classify_transactions(
                [
                    {"Description": "Transfer in", "Deposit": 100, "Withdrawal": ""},
                    {"Description": "Transfer out", "Deposit": "", "Withdrawal": 100},
                    {"Description": "Transfer unclear", "Deposit": "", "Withdrawal": ""},
                ],
                load_rules(path),
            )
            self.assertEqual(classified[0]["Rule_ID"], "TRANSFER")
            self.assertEqual(classified[1]["Rule_ID"], "TRANSFER")
            self.assertEqual(classified[2]["Category"], "Unclassified")
            self.assertIn("Unable to determine transaction direction", classified[2]["Notes"])

    def test_unknown_direction_rule_matches_returns_false(self):
        rule = ClassificationRule(
            rule_id="ANY",
            priority=1,
            enabled=True,
            match_type="contains",
            keyword="transfer",
            direction="Any",
            amount_min=None,
            amount_max=None,
            category="Transfer",
            account_code="",
            account_name="",
            tax_type="",
            counterparty="",
            confidence=0.5,
            review_needed="Yes",
            notes="Synthetic rule",
        )
        self.assertFalse(rule.matches("Transfer unclear", "Unknown", 0))

    def test_priority_lower_number_wins_and_disabled_rule_does_not_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(
                path,
                [
                    make_rule("LOW", "fee", priority="50"),
                    make_rule("HIGH_DISABLED", "fee", priority="1", enabled="No"),
                    make_rule("HIGH", "fee", priority="2"),
                ],
            )
            classified = classify_transactions(
                [{"Description": "Monthly fee", "Deposit": "", "Withdrawal": 10}],
                load_rules(path),
            )
            self.assertEqual(classified[0]["Rule_ID"], "HIGH")

    def test_amount_min_and_amount_max_filters(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.csv"
            write_rules(path, [make_rule("SMALL_FEE", "fee", amount_min="1", amount_max="20")])
            classified = classify_transactions(
                [
                    {"Description": "Service fee", "Deposit": "", "Withdrawal": 10},
                    {"Description": "Service fee", "Deposit": "", "Withdrawal": 30},
                    {"Description": "Service fee", "Deposit": "", "Withdrawal": 0.5},
                ],
                load_rules(path),
            )
            self.assertEqual(classified[0]["Rule_ID"], "SMALL_FEE")
            self.assertEqual(classified[1]["Classification_Source"], "unclassified")
            self.assertEqual(classified[2]["Classification_Source"], "unclassified")

    def test_unclassified_transaction_defaults(self):
        classified = classify_transactions(
            [{"Description": "No matching keyword", "Deposit": 50, "Withdrawal": ""}],
            load_rules(DEFAULT_RULES),
        )
        self.assertEqual(classified[0]["Category"], "Unclassified")
        self.assertEqual(classified[0]["Confidence"], 0)
        self.assertEqual(classified[0]["Review_Needed"], "Yes")
        self.assertEqual(classified[0]["Classification_Source"], "unclassified")

    def test_unknown_direction_transaction_defaults(self):
        rules = load_rules(DEFAULT_RULES)
        for row in [
            {"Description": "Transfer unclear", "Deposit": "", "Withdrawal": ""},
            {"Description": "Transfer unclear", "Deposit": 10, "Withdrawal": 10},
        ]:
            classified = classify_transactions([row], rules)[0]
            self.assertEqual(detect_direction(row), "Unknown")
            self.assertEqual(classified["Category"], "Unclassified")
            self.assertEqual(classified["Confidence"], 0)
            self.assertEqual(classified["Review_Needed"], "Yes")
            self.assertEqual(classified["Classification_Source"], "unclassified")
            self.assertIn("Unable to determine transaction direction", classified["Notes"])

    def test_both_deposit_and_withdrawal_contain_values_note(self):
        classified = classify_transactions(
            [{"Description": "Transfer unclear", "Deposit": 10, "Withdrawal": 10}],
            load_rules(DEFAULT_RULES),
        )[0]
        self.assertEqual(classified["Direction"], "Unknown")
        self.assertIn("Both Deposit and Withdrawal contain values", classified["Notes"])

    def test_both_deposit_and_withdrawal_blank_or_zero_note(self):
        classified = classify_transactions(
            [{"Description": "Transfer unclear", "Deposit": "", "Withdrawal": ""}],
            load_rules(DEFAULT_RULES),
        )[0]
        self.assertEqual(classified["Direction"], "Unknown")
        self.assertIn("Deposit and Withdrawal are both blank or zero", classified["Notes"])

    def test_negative_deposit_anomaly_note(self):
        classified = classify_transactions(
            [{"Description": "Customer receipt", "Deposit": -100, "Withdrawal": ""}],
            load_rules(DEFAULT_RULES),
        )[0]
        self.assertEqual(classified["Rule_ID"], "CUSTOMER_RECEIPT")
        self.assertIn("Negative Deposit amount", classified["Notes"])

    def test_negative_withdrawal_anomaly_note(self):
        classified = classify_transactions(
            [{"Description": "Bank charge", "Deposit": "", "Withdrawal": -25}],
            load_rules(DEFAULT_RULES),
        )[0]
        self.assertEqual(classified["Rule_ID"], "BANK_CHARGE")
        self.assertIn("Negative Withdrawal amount", classified["Notes"])

    def test_output_columns_are_complete_and_ordered(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = write_workbook(Path(tmp) / "classified.xlsx", [])
            workbook = load_workbook(output_path)
            try:
                headers = [workbook.active.cell(1, col).value for col in range(1, workbook.active.max_column + 1)]
            finally:
                workbook.close()
            self.assertEqual(headers, OUTPUT_COLUMNS)

    def test_summary_counts_amounts_and_unclassified_review_rows(self):
        rows = read_transactions(FIXTURES / "baseline_input.csv")
        classified = classify_transactions(rows, load_rules(DEFAULT_RULES))
        summary = summarize_classification(classified)

        self.assertEqual(summary["transaction_count"], 6)
        self.assertEqual(summary["classified_count"], 5)
        self.assertEqual(summary["unclassified_count"], 1)
        self.assertEqual(summary["review_needed_count"], 4)
        self.assertEqual(summary["unclassified_ratio"], 0.1667)
        self.assertEqual(summary["source_counts"], {"rule": 5, "unclassified": 1})
        self.assertEqual(summary["rule_hit_counts"]["BANK_CHARGE"], 1)
        self.assertEqual(summary["rule_hit_counts"]["RENT_PAYMENT"], 1)
        self.assertEqual(summary["anomaly_counts"]["unknown_direction"], 1)
        self.assertEqual(summary["anomaly_counts"]["zero_amount"], 1)
        self.assertEqual(summary["top_unclassified_descriptions"], [
            {"description": "Unknown adjustment", "count": 1}
        ])
        self.assertEqual(summary["direction_amounts"]["Deposit"], 2010)
        self.assertEqual(summary["direction_amounts"]["Withdrawal"], 1825)
        self.assertEqual(summary["category_amounts"]["Unclassified"], 0)

        review_rows = unclassified_rows(classified)
        self.assertEqual([row["Description"] for row in review_rows], [
            "Office rent payment",
            "Customer receipt",
            "Transfer to savings",
            "Unknown adjustment",
        ])

    def test_summary_text_report_contains_a111_fields(self):
        rows = read_transactions(FIXTURES / "baseline_input.csv")
        summary = summarize_classification(classify_transactions(rows, load_rules(DEFAULT_RULES)))
        with tempfile.TemporaryDirectory() as tmp:
            path = write_summary_text(Path(tmp) / "classified.summary.txt", summary)
            text = path.read_text(encoding="utf-8")
        for expected in [
            "transaction_count:",
            "classified_count:",
            "unclassified_count:",
            "review_needed_count:",
            "unclassified_ratio:",
            "category_counts:",
            "source_counts:",
            "rule_hit_counts:",
            "anomaly_counts:",
            "top_unclassified_descriptions:",
            "direction_amounts:",
            "category_amounts:",
        ]:
            self.assertIn(expected, text)

    def test_anomaly_counts_summary(self):
        rows = [
            {"Description": "Both values", "Deposit": 10, "Withdrawal": 5},
            {"Description": "Zero amount", "Deposit": "", "Withdrawal": ""},
            {"Description": "Negative deposit", "Deposit": -10, "Withdrawal": ""},
            {"Description": "Negative withdrawal", "Deposit": "", "Withdrawal": -5},
        ]
        summary = summarize_classification(classify_transactions(rows, load_rules(DEFAULT_RULES)))
        self.assertEqual(summary["anomaly_counts"]["unknown_direction"], 2)
        self.assertEqual(summary["anomaly_counts"]["both_deposit_and_withdrawal"], 1)
        self.assertEqual(summary["anomaly_counts"]["zero_amount"], 1)
        self.assertEqual(summary["anomaly_counts"]["negative_deposit"], 1)
        self.assertEqual(summary["anomaly_counts"]["negative_withdrawal"], 1)

    def test_cli_default_rules_path_is_usable(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "input.csv"
            input_path.write_text(
                "Bank_Account,Date,Description,Deposit,Withdrawal,Balance,Control\n"
                "Synthetic Bank,2025-01-01,Bank charge,,25,975,975\n",
                encoding="utf-8",
            )
            args = build_arg_parser().parse_args([str(input_path), "--output", str(tmp_path / "classified.xlsx")])
            result = run(args)
            self.assertEqual(result["summary"]["classified_count"], 1)
            self.assertEqual(args.rules, DEFAULT_RULES)

    def test_csv_input(self):
        rows = read_transactions(FIXTURES / "baseline_input.csv")
        self.assertEqual(len(rows), 6)
        self.assertEqual(rows[0]["Bank_Account"], "Synthetic Bank")

    def test_xlsx_input_where_practical(self):
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"])
            sheet.append(["Synthetic Bank", "2025-01-01", "Bank charge", "", 25, 975, 975])
            workbook.save(input_path)

            rows = read_transactions(input_path)
            classified = classify_transactions(rows, load_rules(DEFAULT_RULES))
            self.assertEqual(classified[0]["Rule_ID"], "BANK_CHARGE")

    def test_baseline_classification_expected_results(self):
        rows = read_transactions(FIXTURES / "baseline_input.csv")
        classified = classify_transactions(rows, load_rules(DEFAULT_RULES))
        expected_rows = read_csv_rows(FIXTURES / "baseline_expected.csv")

        actual = [
            {
                "Description": row["Description"],
                "Direction": row["Direction"],
                "Amount": str(int(row["Amount"])) if float(row["Amount"]).is_integer() else str(row["Amount"]),
                "Category": row["Category"],
                "Rule_ID": row["Rule_ID"],
                "Review_Needed": row["Review_Needed"],
                "Classification_Source": row["Classification_Source"],
                "Notes": row["Notes"],
            }
            for row in classified
        ]
        self.assertEqual(actual, expected_rows)

    def test_cli_run_writes_all_outputs_and_summary_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            args = build_arg_parser().parse_args(
                [
                    str(FIXTURES / "baseline_input.csv"),
                    "--output",
                    str(tmp_path / "classified.xlsx"),
                ]
            )
            result = run(args)

            self.assertTrue(result["output"].exists())
            self.assertTrue(result["review_output"].exists())
            self.assertTrue(result["summary_json"].exists())
            self.assertTrue(result["summary_txt"].exists())
            payload = json.loads(result["summary_json"].read_text(encoding="utf-8"))
            self.assertEqual(payload["transaction_count"], 6)
            self.assertEqual(payload["unclassified_count"], 1)


if __name__ == "__main__":
    unittest.main()
