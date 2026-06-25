from pathlib import Path

from openpyxl import load_workbook


issues = []
for path in sorted(Path(__file__).resolve().parent.joinpath("output").glob("*.xlsx")):
    workbook = load_workbook(path, data_only=False)
    print("====", path.name, workbook.sheetnames)
    for sheet in workbook.worksheets:
        last = [sheet.cell(sheet.max_row, col).value for col in range(1, 8)]
        bad_descriptions = []
        for row_idx in range(2, sheet.max_row + 1):
            description = str(sheet.cell(row_idx, 3).value or "")
            if any(
                marker in description
                for marker in ["TotalNo.", "TotalDeposit", "The Hongkong", "Page ", "PortfolioSummary"]
            ):
                bad_descriptions.append((row_idx, description[:120]))
            if not str(sheet.cell(row_idx, 7).value or "").startswith("=IF("):
                issues.append((path.name, sheet.title, row_idx, "missing control formula"))
        print(
            sheet.title,
            "rows",
            sheet.max_row - 1,
            "last_balance",
            last[5],
            "last_control_formula",
            last[6],
            "bad_desc",
            bad_descriptions[:3],
        )
        if bad_descriptions:
            issues.append((path.name, sheet.title, "bad_desc", bad_descriptions[:3]))

print("ISSUES", issues)
