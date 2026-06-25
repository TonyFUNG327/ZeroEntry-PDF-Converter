from pathlib import Path

from openpyxl import load_workbook


for name in ["HSBC 2501.xlsx", "HSBC 2512.xlsx", "HSBC 2601.xlsx"]:
    path = Path(__file__).resolve().parent / "output" / name
    workbook = load_workbook(path, data_only=False)
    print("====", name)
    for sheet in workbook.worksheets:
        print(sheet.title)
        for row_idx in range(2, min(sheet.max_row, 8) + 1):
            print(
                row_idx,
                sheet.cell(row_idx, 2).value,
                sheet.cell(row_idx, 3).value,
                sheet.cell(row_idx, 5).value,
                sheet.cell(row_idx, 5).number_format,
            )
