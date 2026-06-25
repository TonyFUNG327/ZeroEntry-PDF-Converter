import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = ["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"]
DEFAULT_INPUT_PATH = Path(__file__).resolve().parent / "BB"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "BB2"
AMOUNT_RE = re.compile(r"^\(?\d{1,3}(?:,\d{3})*(?:\.\d{2})\)?$|^\(?\d+\.\d{2}\)?$")
MONTHS = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}


def amount_value(text):
    sign = -1 if text.startswith("(") and text.endswith(")") else 1
    return sign * float(text.strip("()").replace(",", ""))


def normalize_date(day, month, statement_year=2024, statement_month=11):
    month_num = MONTHS[month]
    year = statement_year
    if month_num > statement_month:
        year -= 1
    return datetime(year, month_num, int(day))


def words_to_lines(words):
    rows = []
    for word in words:
        if word["top"] < 120 or word["top"] > 785:
            continue
        for row in rows:
            if abs(row["top"] - word["top"]) <= 2.2:
                row["words"].append(word)
                row["top"] = (row["top"] + word["top"]) / 2
                break
        else:
            rows.append({"top": word["top"], "words": [word]})

    rows.sort(key=lambda row: row["top"])
    for row in rows:
        row["words"].sort(key=lambda word: word["x0"])
    return rows


def classify_amount(word):
    x0, x1 = word["x0"], word["x1"]
    if 320 <= x1 <= 379:
        return "Deposit"
    if 395 <= x1 <= 456:
        return "Withdrawal"
    if x0 >= 470:
        return "Balance"
    return None


def extract_pdf(pdf_path):
    pdf_path = Path(pdf_path)
    accounts = defaultdict(list)
    current_account = None
    in_table = False
    current_date = None
    pending_desc = []

    def flush_transaction(line_desc, amounts):
        nonlocal pending_desc, current_date, current_account
        if not current_account or not current_date:
            pending_desc = []
            return

        desc_parts = pending_desc[:]
        if line_desc:
            desc_parts.append(line_desc)
        description = " ".join(part for part in desc_parts if part).strip() or "BALANCE"
        accounts[current_account].append(
            {
                "Bank_Account": current_account,
                "Date": current_date,
                "Description": description,
                "Deposit": amounts.get("Deposit"),
                "Withdrawal": amounts.get("Withdrawal"),
                "Balance": amounts.get("Balance"),
            }
        )
        pending_desc = []

    with pdfplumber.open(pdf_path) as pdf:
        statement_year = 2024
        statement_month = 11
        for page in pdf.pages:
            lines = words_to_lines(page.extract_words(x_tolerance=1, y_tolerance=3) or [])
            for line in lines:
                words = line["words"]
                text = " ".join(word["text"] for word in words)

                if "Statement Date" in text:
                    match = re.search(r"(\d{1,2})\s+([A-Z][a-z]{2})\s+(\d{4})", text)
                    if match:
                        statement_month = MONTHS[match.group(2)]
                        statement_year = int(match.group(3))

                if text.strip() == "HKD Statement Savings":
                    current_account = "Hang Seng Savings Account"
                    in_table = False
                    current_date = None
                    pending_desc = []
                    continue
                if text.strip() == "Current":
                    current_account = "Hang Seng Current Account"
                    in_table = False
                    current_date = None
                    pending_desc = []
                    continue
                if text.startswith("Date Transaction Details"):
                    in_table = True
                    pending_desc = []
                    continue
                if not in_table:
                    continue
                if text.startswith("Transaction Summary") or text.startswith("Credit Interest Accrued"):
                    in_table = False
                    pending_desc = []
                    continue
                if "Business Partner Direct" in text or "hangseng.com" in text:
                    continue

                date_words = [word for word in words if 55 <= word["x0"] <= 90]
                if len(date_words) >= 2 and date_words[0]["text"].isdigit() and date_words[1]["text"] in MONTHS:
                    current_date = normalize_date(
                        date_words[0]["text"],
                        date_words[1]["text"],
                        statement_year,
                        statement_month,
                    )

                amounts = {}
                desc_words = []
                has_balance_dr = any(word["text"] == "DR" and word["x0"] >= 520 for word in words)
                for word in words:
                    cls = classify_amount(word) if AMOUNT_RE.match(word["text"]) else None
                    if cls:
                        value = amount_value(word["text"])
                        if cls == "Balance" and has_balance_dr:
                            value = -abs(value)
                        amounts[cls] = value
                    elif 94 <= word["x0"] <= 295:
                        desc_words.append(word["text"])

                line_desc = " ".join(desc_words).strip()
                if amounts:
                    flush_transaction(line_desc, amounts)
                elif line_desc:
                    pending_desc.append(line_desc)

    for rows in accounts.values():
        control = None
        for row in rows:
            if row["Deposit"] is None and row["Withdrawal"] is None and row["Balance"] is not None:
                control = row["Balance"]
            else:
                if control is None:
                    control = row["Balance"] or 0.0
                control += row["Deposit"] or 0.0
                control -= row["Withdrawal"] or 0.0
            row["Control"] = round(control, 2) if control is not None else None

    return dict(accounts)


def write_workbook(accounts, output_path):
    output_path = Path(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    control_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    grid_side = Side(style="thin", color="D9D9D9")
    grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)
    widths = [28, 14, 48, 16, 16, 16, 16]

    for account, rows in accounts.items():
        sheet = workbook.create_sheet(account[:31])
        sheet.append(HEADERS)
        for row in rows:
            sheet.append(
                [
                    row["Bank_Account"],
                    row["Date"],
                    row["Description"],
                    row["Deposit"],
                    row["Withdrawal"],
                    row["Balance"],
                    None,
                ]
            )

        for cell in sheet[1]:
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = header_fill
            cell.border = grid_border

        for row_idx in range(2, sheet.max_row + 1):
            if row_idx == 2:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )
            else:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'G{row_idx - 1}+IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10, bold=cell.column == 7 or cell.row == 1)
                cell.border = grid_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if cell.column == 7:
                    cell.fill = control_fill

        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in sheet["B"][1:]:
            cell.number_format = "dd mmm yyyy"
        for cell in sheet["D"][1:]:
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        for cell in sheet["E"][1:]:
            cell.number_format = '(#,##0.00);(#,##0.00);"-"'
        for column in ["F", "G"]:
            for cell in sheet[column][1:]:
                cell.number_format = "#,##0.00;(#,##0.00);0.00"

        sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def validate_accounts(accounts):
    report = []
    for account, rows in accounts.items():
        deposit_total = round(sum(row["Deposit"] or 0 for row in rows), 2)
        withdrawal_total = round(sum(row["Withdrawal"] or 0 for row in rows), 2)
        final_balance = rows[-1]["Balance"] if rows else None
        final_control = rows[-1]["Control"] if rows else None
        balance_mismatches = [
            idx + 2
            for idx, row in enumerate(rows)
            if row["Balance"] is not None and abs(row["Balance"] - row["Control"]) > 0.01
        ]
        report.append(
            {
                "account": account,
                "rows": len(rows),
                "deposit_total": deposit_total,
                "withdrawal_total": withdrawal_total,
                "final_balance": final_balance,
                "final_control": final_control,
                "balance_mismatches": balance_mismatches,
            }
        )
    return report


def convert_pdf_to_excel(pdf_path, output_dir=None):
    pdf_path = Path(pdf_path)
    output_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
    accounts = extract_pdf(pdf_path)
    output_path = output_dir / f"{pdf_path.stem}.xlsx"
    return write_workbook(accounts, output_path)


def collect_pdf_paths(input_path):
    input_path = Path(input_path)
    if input_path.is_file():
        if input_path.suffix.lower() != ".pdf":
            raise ValueError(f"Input file is not a PDF: {input_path}")
        return [input_path]
    if input_path.is_dir():
        return sorted(path for path in input_path.iterdir() if path.is_file() and path.suffix.lower() == ".pdf")
    raise FileNotFoundError(f"Input path does not exist: {input_path}")


def build_arg_parser():
    parser = argparse.ArgumentParser(description="Convert Hang Seng statement PDF to Excel.")
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        default=DEFAULT_INPUT_PATH,
        help=f"PDF file or folder of PDFs. Default: {DEFAULT_INPUT_PATH}",
    )
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help=f"Output directory. Default: {DEFAULT_OUTPUT_DIR}")
    return parser


def main(argv=None):
    args = build_arg_parser().parse_args(argv)
    pdf_paths = collect_pdf_paths(args.input)
    if not pdf_paths:
        print(f"No PDF files found in: {args.input}")
        return 1

    for pdf_path in pdf_paths:
        accounts = extract_pdf(pdf_path)
        output_path = write_workbook(accounts, args.output_dir / f"{pdf_path.stem}.xlsx")
        print(f"Saved: {output_path}")
        for item in validate_accounts(accounts):
            print(
                f"{item['account']}: rows={item['rows']}, "
                f"deposit={item['deposit_total']:,.2f}, withdrawal={item['withdrawal_total']:,.2f}, "
                f"final_balance={item['final_balance']:,.2f}, final_control={item['final_control']:,.2f}, "
                f"balance_mismatches={len(item['balance_mismatches'])}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
