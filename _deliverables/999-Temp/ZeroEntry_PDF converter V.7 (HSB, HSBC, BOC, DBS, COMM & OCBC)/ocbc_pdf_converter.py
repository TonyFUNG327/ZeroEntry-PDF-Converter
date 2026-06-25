import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = ["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"]
DEFAULT_INPUT_PATH = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "generated_output"
DATE_RE = re.compile(r"^\d{2}[A-Z]{3}\d{2}$")
AMOUNT_RE = re.compile(r"^\(?\d{1,3}(?:,\d{3})*(?:\.\d{2})\)?$|^\(?\d+\.\d{2}\)?$")
SECTION_RE = re.compile(r"^(HKD|CNY|USD|EUR|NOK|GBP|AUD|CAD|CHF|JPY|SGD)\s+(.+?)\s+(?:港元|人民幣|美元|歐元|挪威克朗|結單|往來|儲蓄|$)")


def clean_text(text):
    return (
        text.replace("(cid:10)", " ")
        .replace("(cid:13)", " ")
        .replace("_____", " ")
        .strip()
    )


def amount_value(text):
    sign = -1 if text.startswith("(") and text.endswith(")") else 1
    return sign * float(text.strip("()").replace(",", ""))


def nonzero(value):
    if value is None or abs(value) < 0.005:
        return None
    return value


def parse_date(text):
    return datetime.strptime(text, "%d%b%y")


def append_description(row, extra):
    extra = clean_text(extra)
    if extra:
        row["Description"] = f"{row['Description']} {extra}".strip()


def account_name(currency, account_type):
    normalized = " ".join(account_type.title().split())
    return f"OCBC {currency} {normalized}"


def sheet_title_for_account(account):
    return account[:31]


def is_section_line(text):
    return SECTION_RE.match(text) and ("STATEMENT" in text or "CURRENT" in text or "SAVINGS" in text)


def is_ignored_line(text):
    markers = [
        "ACCOUNT ACTIVITIES",
        "ACCOUNT NO.",
        "ACCOUNT SUMMARY",
        "ACCOUNT TYPE",
        "ADDITIONAL INFORMATION",
        "AMOUNT ",
        "BALANCE EX. RATE",
        "BANK REFERENCE",
        "BRANCH",
        "CARRIED FORWARD",
        "CUSTOMER SHOULD",
        "DATE PARTICULARS",
        "DR=DEBIT",
        "END OF STATEMENT",
        "FOREIGN CURRENCY DEPOSIT",
        "HKD DEPOSIT",
        "INTEGRATED ACCOUNT",
        "ITEM(S)",
        "LATEST NEWS",
        "NET POSITION",
        "OVERDRAFT",
        "PAGE ",
        "PARTICULARS WITHDRAWAL",
        "PHOENIX RISK MANAGEMENT",
        "PORTFOLIO",
        "REMARKS",
        "SERVICE ENQUIRY",
        "TIME DEPOSITS SUBJECT",
        "STATEMENT DATE",
        "TOTAL OVERDRAFT",
        "TRANSACTION SUMMARY",
        "WITHDRAWAL DEPOSIT BALANCE",
        "交易詳情",
        "客戶",
        "今期結餘",
        "有關",
        "服務查詢",
        "結單日期",
        "結欠",
        "貨幣靈活",
        "銀行參考編號",
        "賬戶",
        "進支摘要",
        "頁數",
    ]
    upper_text = text.upper()
    return any(marker in upper_text for marker in markers)


def parse_transaction_line(text):
    parts = text.split()
    if not parts:
        return None

    date_value = None
    start_idx = 0
    if DATE_RE.match(parts[0]):
        date_value = parse_date(parts[0])
        start_idx = 1

    amount_positions = [idx for idx, part in enumerate(parts[start_idx:], start=start_idx) if AMOUNT_RE.match(part)]
    if len(amount_positions) < 3:
        return None

    first_amount_idx = amount_positions[-3]
    amount_tokens = [parts[idx] for idx in amount_positions[-3:]]
    description = " ".join(parts[start_idx:first_amount_idx]).strip()
    withdrawal, deposit, balance = [amount_value(token) for token in amount_tokens]
    return {
        "date": date_value,
        "description": description,
        "Withdrawal": nonzero(withdrawal),
        "Deposit": nonzero(deposit),
        "Balance": balance,
    }


def extract_pdf(pdf_path):
    accounts = defaultdict(list)
    current_account = None
    current_date = None
    last_row = None
    in_activities = False

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for raw_line in text.splitlines():
                line = clean_text(raw_line)
                if not line:
                    continue

                if "ACCOUNT ACTIVITIES" in line:
                    in_activities = True

                section_match = SECTION_RE.match(line)
                if in_activities and section_match and is_section_line(line):
                    currency = section_match.group(1)
                    account_type = section_match.group(2)
                    current_account = account_name(currency, account_type)
                    current_date = None
                    last_row = None
                    continue

                if not in_activities or not current_account:
                    continue

                if "CARRIED FORWARD" in line or "今期結餘" in line:
                    current_date = None
                    last_row = None
                    continue
                if is_ignored_line(line):
                    if any(marker in line.upper() for marker in ["BANK REFERENCE", "REMARKS", "ADDITIONAL INFORMATION", "END OF STATEMENT"]):
                        last_row = None
                    if any(marker in line for marker in ["銀行參考編號", "額 外 資 訊"]):
                        last_row = None
                    continue

                parsed = parse_transaction_line(line)
                if parsed:
                    if parsed["date"] is not None:
                        current_date = parsed["date"]
                    if current_date is None:
                        continue
                    description = parsed["description"] or "BALANCE"
                    if "B/F BALANCE" in description:
                        deposit = None
                        withdrawal = None
                    else:
                        deposit = parsed["Deposit"]
                        withdrawal = parsed["Withdrawal"]
                    row = {
                        "Bank_Account": current_account,
                        "Date": current_date,
                        "Description": description,
                        "Deposit": deposit,
                        "Withdrawal": withdrawal,
                        "Balance": parsed["Balance"],
                    }
                    accounts[current_account].append(row)
                    last_row = row
                elif last_row is not None:
                    append_description(last_row, line)

    for rows in accounts.values():
        control = None
        for row in rows:
            if row["Deposit"] is None and row["Withdrawal"] is None and row["Balance"] is not None:
                control = row["Balance"]
            else:
                if control is None:
                    control = row["Balance"] or 0.0
                control += row["Deposit"] or 0.0
                control -= row["Withdrawal"] or 0.0
            row["Control"] = round(control, 2) if control is not None else None

    return dict(accounts)


def write_workbook(accounts, output_path):
    output_path = Path(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    control_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    grid_side = Side(style="thin", color="D9D9D9")
    grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)
    widths = [30, 14, 58, 16, 16, 16, 16]

    for account, rows in accounts.items():
        sheet = workbook.create_sheet(sheet_title_for_account(account))
        sheet.append(HEADERS)
        for row in rows:
            sheet.append([
                row["Bank_Account"],
                row["Date"],
                row["Description"],
                row["Deposit"],
                row["Withdrawal"],
                row["Balance"],
                None,
            ])
        for cell in sheet[1]:
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = header_fill
            cell.border = grid_border
        for row_idx in range(2, sheet.max_row + 1):
            if row_idx == 2:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )
            else:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'G{row_idx - 1}+IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10, bold=cell.column == 7 or cell.row == 1)
                cell.border = grid_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if cell.column == 7:
                    cell.fill = control_fill
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in sheet["B"][1:]:
            cell.number_format = "dd mmm yyyy"
        for column in ["D", "E"]:
            for cell in sheet[column][1:]:
                cell.number_format = '#,##0.00;(#,##0.00);"-"'
        for column in ["F", "G"]:
            for cell in sheet[column][1:]:
                cell.number_format = "#,##0.00;(#,##0.00);0.00"
        sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def validate_accounts(accounts):
    report = []
    for account, rows in accounts.items():
        deposit_total = round(sum(row["Deposit"] or 0 for row in rows), 2)
        withdrawal_total = round(sum(row["Withdrawal"] or 0 for row in rows), 2)
        final_balance = rows[-1]["Balance"] if rows else None
        final_control = rows[-1]["Control"] if rows else None
        balance_mismatches = [
            idx + 2
            for idx, row in enumerate(rows)
            if row["Balance"] is not None and abs(row["Balance"] - row["Control"]) > 0.01
        ]
        report.append({
            "account": account,
            "rows": len(rows),
            "deposit_count": sum(1 for row in rows if row["Deposit"] is not None),
            "withdrawal_count": sum(1 for row in rows if row["Withdrawal"] is not None),
            "deposit_total": deposit_total,
            "withdrawal_total": withdrawal_total,
            "final_balance": final_balance,
            "final_control": final_control,
            "balance_mismatches": balance_mismatches,
        })
    return report


def collect_pdf_paths(input_path):
    input_path = Path(input_path)
    if input_path.is_file():
        return [input_path]
    return sorted(path for path in input_path.iterdir() if path.is_file() and path.suffix.lower() == ".pdf")


def main(argv=None):
    parser = argparse.ArgumentParser(description="Convert OCBC statement PDF to Excel.")
    parser.add_argument("input", type=Path, nargs="?", default=DEFAULT_INPUT_PATH)
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args(argv)
    for pdf_path in collect_pdf_paths(args.input):
        accounts = extract_pdf(pdf_path)
        output_path = write_workbook(accounts, args.output_dir / f"{pdf_path.stem}.xlsx")
        print(f"Saved: {output_path}")
        for item in validate_accounts(accounts):
            print(
                f"{item['account']}: rows={item['rows']}, deposits={item['deposit_count']} amount={item['deposit_total']:,.2f}, "
                f"withdrawals={item['withdrawal_count']} amount={item['withdrawal_total']:,.2f}, "
                f"final_balance={item['final_balance']:,.2f}, final_control={item['final_control']:,.2f}, "
                f"balance_mismatches={len(item['balance_mismatches'])}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
