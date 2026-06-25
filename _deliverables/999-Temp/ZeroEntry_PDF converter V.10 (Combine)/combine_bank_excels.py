import argparse
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_INPUT_DIR = Path(__file__).resolve().parent / "BB2"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "BB3"
HEADERS = ["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"]


def collect_excel_paths(input_dir):
    input_dir = Path(input_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input folder does not exist: {input_dir}")
    return sorted(
        path
        for path in input_dir.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")
    )


def is_brought_forward(row):
    description = str(row[2] or "").strip().upper()
    return description in {"B/F BALANCE", "BAL B/F", "BALANCE BROUGHT FORWARD"}


def row_values(sheet, row_idx):
    return [sheet.cell(row_idx, col).value for col in range(1, 8)]


def copy_cell_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def apply_control_formula(sheet, row_idx):
    if row_idx == 2:
        sheet.cell(row_idx, 7).value = (
            f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
            f'IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
        )
    else:
        sheet.cell(row_idx, 7).value = (
            f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
            f'G{row_idx - 1}+IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
        )


def copy_row(source_sheet, source_row_idx, target_sheet):
    target_row_idx = target_sheet.max_row + 1
    for col_idx in range(1, 8):
        source_cell = source_sheet.cell(source_row_idx, col_idx)
        target_cell = target_sheet.cell(target_row_idx, col_idx)
        target_cell.value = source_cell.value
        copy_cell_style(source_cell, target_cell)
    apply_control_formula(target_sheet, target_row_idx)
    return target_row_idx


def copy_header(source_sheet, target_sheet):
    for col_idx in range(1, 8):
        source_cell = source_sheet.cell(1, col_idx)
        target_cell = target_sheet.cell(1, col_idx)
        target_cell.value = source_cell.value
        copy_cell_style(source_cell, target_cell)
    for col_idx in range(1, 8):
        letter = source_sheet.cell(1, col_idx).column_letter
        target_sheet.column_dimensions[letter].width = source_sheet.column_dimensions[letter].width
    target_sheet.freeze_panes = source_sheet.freeze_panes


def last_nonblank_balance(sheet):
    for row_idx in range(sheet.max_row, 1, -1):
        value = sheet.cell(row_idx, 6).value
        if value is not None:
            return float(value)
    return None


def first_brought_forward_balance(sheet):
    if sheet.max_row < 2:
        return None
    values = row_values(sheet, 2)
    if is_brought_forward(values):
        return values[5]
    return None


def output_name(excel_paths):
    first = excel_paths[0].stem
    last = excel_paths[-1].stem
    if first == last:
        return f"{first} combined.xlsx"
    return f"{first} - {last} combined.xlsx"


def combine_excels(input_dir=DEFAULT_INPUT_DIR, output_dir=DEFAULT_OUTPUT_DIR, output_path=None, allow_mismatch=False):
    excel_paths = collect_excel_paths(input_dir)
    if not excel_paths:
        raise ValueError(f"No .xlsx files found in: {input_dir}")

    output_dir = Path(output_dir)
    output_path = Path(output_path) if output_path else output_dir / output_name(excel_paths)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    target_workbook = Workbook()
    target_workbook.remove(target_workbook.active)
    target_sheets = {}
    reports = []

    for source_path in excel_paths:
        source_workbook = load_workbook(source_path, data_only=False)
        for source_sheet in source_workbook.worksheets:
            sheet_name = source_sheet.title
            if source_sheet.max_row < 2:
                continue

            if sheet_name not in target_sheets:
                target_sheet = target_workbook.create_sheet(sheet_name)
                copy_header(source_sheet, target_sheet)
                target_sheets[sheet_name] = target_sheet
                start_row = 2
                reports.append(f"{source_path.name} [{sheet_name}]: first file, kept B/F row.")
            else:
                target_sheet = target_sheets[sheet_name]
                expected = last_nonblank_balance(target_sheet)
                actual = first_brought_forward_balance(source_sheet)
                if expected is not None and actual is not None and abs(expected - float(actual)) > 0.01:
                    message = (
                        f"{source_path.name} [{sheet_name}]: B/F mismatch. "
                        f"Previous ending balance={expected:,.2f}, current B/F={float(actual):,.2f}."
                    )
                    if not allow_mismatch:
                        raise ValueError(message)
                    reports.append("WARNING " + message)
                else:
                    reports.append(f"{source_path.name} [{sheet_name}]: B/F matched, skipped B/F row.")
                start_row = 3 if is_brought_forward(row_values(source_sheet, 2)) else 2

            for source_row_idx in range(start_row, source_sheet.max_row + 1):
                copy_row(source_sheet, source_row_idx, target_sheets[sheet_name])

    target_workbook.save(output_path)
    report_path = output_path.with_suffix(".txt")
    report_path.write_text("\n".join(reports) + "\n", encoding="utf-8")
    return output_path, report_path, reports


def build_arg_parser():
    parser = argparse.ArgumentParser(description="Combine monthly converted bank Excel files from BB2 into one workbook.")
    parser.add_argument("input", type=Path, nargs="?", default=DEFAULT_INPUT_DIR)
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--allow-mismatch", action="store_true")
    return parser


def main(argv=None):
    args = build_arg_parser().parse_args(argv)
    output_path, report_path, reports = combine_excels(
        input_dir=args.input,
        output_dir=args.output_dir,
        output_path=args.output,
        allow_mismatch=args.allow_mismatch,
    )
    print(f"Saved combined workbook: {output_path}")
    print(f"Saved combine report: {report_path}")
    for report in reports:
        print(report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
