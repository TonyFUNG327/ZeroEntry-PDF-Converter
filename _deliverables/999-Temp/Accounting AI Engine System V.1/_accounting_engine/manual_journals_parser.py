from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEDULES = {
    "Trade Receivables": {
        "source": "Trade Receivables",
        "prefix": "AR",
        "contra_role": "Trade receivable",
        "dr": "contra",
        "cr": "nature",
    },
    "Trade payables": {
        "source": "Trade payables",
        "prefix": "AP",
        "contra_role": "Trade payable",
        "dr": "nature",
        "cr": "contra",
    },
    "Accruals": {
        "source": "Accruals",
        "prefix": "JV",
        "contra_role": "Accruals",
        "dr": "nature",
        "cr": "contra",
    },
}


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def amount(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if not value:
            return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def read_master(wb) -> dict[str, Any]:
    if "Master Pg" not in wb.sheetnames:
        return {}
    ws = wb["Master Pg"]
    return {
        "client_name": ws["B1"].value,
        "period_label": ws["B2"].value,
        "file_no": ws["B3"].value,
        "currency": ws["B5"].value,
        "period_from": ws["B7"].value,
        "period_to": ws["B8"].value,
        "incorporation_cessation": ws["B9"].value or ws["C9"].value,
        "job_in_charge": ws["B11"].value,
    }


def read_chart_of_accounts(wb) -> dict[str, str]:
    if "Chart of Accounts" not in wb.sheetnames:
        return {}
    ws = wb["Chart of Accounts"]
    accounts: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        name = text(ws.cell(row, 1).value)
        code = text(ws.cell(row, 2).value)
        if name and code and name.lower() not in {"account", "description"}:
            accounts[name.casefold()] = code
    return accounts


def account_code(accounts: dict[str, str], name: str) -> str:
    return accounts.get(text(name).casefold(), "")


def make_description(nature: str, party: str, invoice: str, details: str) -> str:
    suffix = "".join(
        f" - {part}" for part in (text(invoice), text(details)) if text(part)
    )
    if party:
        return f"{nature}:{party}{suffix}"
    return f"{nature}{suffix}"


def find_first_data_row(ws) -> int:
    for row in range(1, min(ws.max_row, 50) + 1):
        headers = [text(ws.cell(row, col).value).casefold() for col in range(1, 10)]
        if "ref. no." in headers and "date" in headers and "nature" in headers:
            return row + 1
    return 8


def row_is_total_or_check(ws, row: int) -> bool:
    labels = " ".join(text(ws.cell(row, col).value).casefold() for col in range(1, 10))
    return any(marker in labels for marker in ("per tb", "check", "total"))


def signed_lines_for_row(
    *,
    schedule_name: str,
    rule: dict[str, str],
    accounts: dict[str, str],
    row_data: dict[str, Any],
) -> list[dict[str, Any]]:
    gross = amount(row_data["amount"])
    if gross == 0:
        return []

    section_account = text(row_data["section_account"]) or rule["contra_role"]
    nature = text(row_data["nature"])
    contra_code = account_code(accounts, section_account)
    nature_code = account_code(accounts, nature)
    reference = text(row_data["reference"])
    if not reference.startswith(rule["prefix"] + "-"):
        reference = f"{rule['prefix']}-{reference}" if reference else rule["prefix"]

    common = {
        "source": schedule_name,
        "date": row_data["date"],
        "reference": reference,
        "description": make_description(
            nature,
            text(row_data["customer"]),
            text(row_data["invoice_no"]),
            text(row_data["details"]),
        ),
        "gross_amount": gross,
        "settlement_1_bank": row_data["settlement_1_bank"],
        "settlement_1_date": row_data["settlement_1_date"],
        "settlement_1_amount": amount(row_data["settlement_1_amount"]),
        "settlement_2_bank": row_data["settlement_2_bank"],
        "settlement_2_date": row_data["settlement_2_date"],
        "settlement_2_amount": amount(row_data["settlement_2_amount"]),
        "outstanding_balance": amount(row_data["outstanding_balance"]),
    }

    roles = {
        "contra": {"account_name": section_account, "account_code": contra_code},
        "nature": {"account_name": nature, "account_code": nature_code},
    }
    debit_role = roles[rule["dr"]]
    credit_role = roles[rule["cr"]]

    return [
        {
            **common,
            "line_type": "Debit",
            "account_code": debit_role["account_code"],
            "account_name": debit_role["account_name"],
            "debit": abs(gross),
            "credit": 0.0,
            "signed_amount": abs(gross),
        },
        {
            **common,
            "line_type": "Credit",
            "account_code": credit_role["account_code"],
            "account_name": credit_role["account_name"],
            "debit": 0.0,
            "credit": abs(gross),
            "signed_amount": -abs(gross),
        },
    ]


def parse_schedule(wb, sheet_name: str, accounts: dict[str, str]) -> tuple[list[dict[str, Any]], list[str]]:
    if sheet_name not in wb.sheetnames:
        return [], [f"Missing sheet: {sheet_name}"]

    ws = wb[sheet_name]
    rule = SCHEDULES[sheet_name]
    section_account = text(ws["B4"].value) or rule["contra_role"]
    data_start = find_first_data_row(ws)
    lines: list[dict[str, Any]] = []
    warnings: list[str] = []

    for row in range(data_start, ws.max_row + 1):
        if row_is_total_or_check(ws, row):
            continue
        ref_no = text(ws.cell(row, 2).value)
        tx_date = ws.cell(row, 3).value
        nature = text(ws.cell(row, 4).value)
        gross = amount(ws.cell(row, 9).value)
        if not (ref_no or tx_date or nature or gross):
            continue
        if not (tx_date and nature and gross):
            warnings.append(f"{sheet_name} row {row}: skipped incomplete row")
            continue

        row_data = {
            "section_account": section_account,
            "reference": ref_no,
            "date": tx_date,
            "nature": nature,
            "customer": ws.cell(row, 5).value,
            "invoice_no": ws.cell(row, 6).value,
            "details": ws.cell(row, 7).value,
            "amount": gross,
            "settlement_1_bank": ws.cell(row, 11).value,
            "settlement_1_date": ws.cell(row, 12).value,
            "settlement_1_amount": ws.cell(row, 13).value,
            "settlement_2_bank": ws.cell(row, 15).value,
            "settlement_2_date": ws.cell(row, 16).value,
            "settlement_2_amount": ws.cell(row, 17).value,
            "outstanding_balance": ws.cell(row, 19).value,
        }
        row_lines = signed_lines_for_row(
            schedule_name=sheet_name,
            rule=rule,
            accounts=accounts,
            row_data=row_data,
        )
        for line in row_lines:
            if not line["account_code"]:
                warnings.append(
                    f"{sheet_name} row {row}: missing account code for {line['account_name']}"
                )
        lines.extend(row_lines)

    return lines, warnings


def summarize(lines: list[dict[str, Any]]) -> dict[str, Any]:
    by_reference: dict[str, float] = {}
    by_source: dict[str, dict[str, float]] = {}
    for line in lines:
        by_reference[line["reference"]] = by_reference.get(line["reference"], 0.0) + float(
            line["signed_amount"]
        )
        source = line["source"]
        bucket = by_source.setdefault(source, {"debit": 0.0, "credit": 0.0, "net": 0.0})
        bucket["debit"] += float(line["debit"])
        bucket["credit"] += float(line["credit"])
        bucket["net"] += float(line["signed_amount"])
    unbalanced = {
        ref: round(total, 2) for ref, total in by_reference.items() if round(total, 2) != 0
    }
    return {
        "journal_line_count": len(lines),
        "journal_reference_count": len(by_reference),
        "total_debit": round(sum(float(line["debit"]) for line in lines), 2),
        "total_credit": round(sum(float(line["credit"]) for line in lines), 2),
        "net": round(sum(float(line["signed_amount"]) for line in lines), 2),
        "by_source": {
            source: {key: round(value, 2) for key, value in values.items()}
            for source, values in by_source.items()
        },
        "unbalanced_references": unbalanced,
    }


def empty_payload() -> dict[str, Any]:
    return {
        "manual_journals_provided": False,
        "metadata": {},
        "journal_lines": [],
        "summary": {
            "journal_line_count": 0,
            "journal_reference_count": 0,
            "total_debit": 0.0,
            "total_credit": 0.0,
            "net": 0.0,
            "by_source": {},
            "unbalanced_references": {},
        },
        "warnings": [],
        "manual_entry_enabled": True,
        "manual_entry_note": "Manual Journals file was not supplied. User may enter JV lines manually later.",
    }


def parse_manual_journals(path: Path | None) -> dict[str, Any]:
    if path is None:
        return empty_payload()

    wb = load_workbook(path, data_only=True)
    accounts = read_chart_of_accounts(wb)
    lines: list[dict[str, Any]] = []
    warnings: list[str] = []
    for sheet_name in SCHEDULES:
        sheet_lines, sheet_warnings = parse_schedule(wb, sheet_name, accounts)
        lines.extend(sheet_lines)
        warnings.extend(sheet_warnings)
    return {
        "manual_journals_provided": True,
        "source_file": str(path),
        "metadata": read_master(wb),
        "journal_lines": lines,
        "summary": summarize(lines),
        "warnings": warnings,
        "manual_entry_enabled": True,
        "manual_entry_note": "User may still add manual JV lines after imported schedules.",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse optional Manual Journals schedules.")
    parser.add_argument("--manual-journals", type=Path)
    parser.add_argument("--out-json", type=Path, required=True)
    args = parser.parse_args()

    payload = parse_manual_journals(args.manual_journals)
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=json_default),
        encoding="utf-8",
    )
    print(json.dumps(payload["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
