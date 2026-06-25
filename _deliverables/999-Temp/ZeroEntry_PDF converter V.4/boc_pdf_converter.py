import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = ["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"]
DEFAULT_INPUT_PATH = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "generated_output"
DATE_RE = re.compile(r"^\d{4}/\d{2}/\d{2}$")
AMOUNT_RE = re.compile(r"^\(?\d{1,3}(?:,\d{3})*(?:\.\d{2})\)?$|^\(?\d+\.\d{2}\)?$")


def amount_value(text):
    sign = -1 if text.startswith("(") and text.endswith(")") else 1
    return sign * float(text.strip("()").replace(",", ""))


def parse_date(text):
    return datetime.strptime(text, "%Y/%m/%d")


def words_to_lines(words):
    rows = []
    for word in words:
        if word["top"] < 70 or word["top"] > 790:
            continue
        for row in rows:
            if abs(row["top"] - word["top"]) <= 2.2:
                row["words"].append(word)
                row["top"] = (row["top"] + word["top"]) / 2
                break
        else:
            rows.append({"top": word["top"], "words": [word]})

    rows.sort(key=lambda row: row["top"])
    for row in rows:
        row["words"].sort(key=lambda word: word["x0"])
    return rows


def classify_amount(word):
    x0, x1 = word["x0"], word["x1"]
    if 345 <= x1 <= 410:
        return "Deposit"
    if 430 <= x1 <= 490:
        return "Withdrawal"
    if x0 >= 505:
        return "Balance"
    return None


def sheet_title_for_account(account):
    return account[:31]


def append_description(row, extra):
    extra = extra.strip()
    if not extra:
        return
    row["Description"] = f"{row['Description']} {extra}".strip()


def extract_pdf(pdf_path):
    pdf_path = Path(pdf_path)
    accounts = defaultdict(list)
    current_account = None
    in_table = False
    last_row = None

    def add_transaction(date_value, description, amounts):
        nonlocal last_row
        if not current_account or not date_value:
            return
        row = {
            "Bank_Account": current_account,
            "Date": date_value,
            "Description": description.strip() or "BALANCE",
            "Deposit": amounts.get("Deposit"),
            "Withdrawal": amounts.get("Withdrawal"),
            "Balance": amounts.get("Balance"),
        }
        accounts[current_account].append(row)
        last_row = row

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            in_table = False
            lines = words_to_lines(page.extract_words(x_tolerance=1, y_tolerance=3) or [])
            for line in lines:
                words = line["words"]
                text = " ".join(word["text"] for word in words).strip()

                if "港元儲蓄" in text and "(" in text:
                    current_account = "BOC HKD Savings Account"
                    in_table = False
                    last_row = None
                    continue
                if "港元往來" in text and "(" in text:
                    current_account = "BOC HKD Current Account"
                    in_table = False
                    last_row = None
                    continue
                if text.startswith("交易日期"):
                    in_table = current_account is not None
                    continue
                if text.startswith("地址：") or text.startswith("中國銀行") or "企業客戶服務熱線" in text:
                    in_table = False
                    continue
                if not in_table:
                    continue

                date_words = [word for word in words if DATE_RE.match(word["text"]) and 55 <= word["x0"] <= 100]
                date_value = parse_date(date_words[0]["text"]) if date_words else None

                amounts = {}
                desc_words = []
                for word in words:
                    if DATE_RE.match(word["text"]):
                        continue
                    cls = classify_amount(word) if AMOUNT_RE.match(word["text"]) else None
                    if cls:
                        amounts[cls] = amount_value(word["text"])
                    elif 145 <= word["x0"] <= 335:
                        desc_words.append(word["text"])

                description = " ".join(desc_words).strip()
                if date_value and amounts:
                    add_transaction(date_value, description, amounts)
                elif date_value and description:
                    add_transaction(date_value, description, {})
                elif description and last_row is not None:
                    append_description(last_row, description)

    for rows in accounts.values():
        control = None
        for row in rows:
            if row["Deposit"] is None and row["Withdrawal"] is None and row["Balance"] is not None:
                control = row["Balance"]
            else:
                if control is None:
                    control = row["Balance"] or 0.0
                control += row["Deposit"] or 0.0
                control -= row["Withdrawal"] or 0.0
            row["Control"] = round(control, 2) if control is not None else None

    return dict(accounts)


def write_workbook(accounts, output_path):
    output_path = Path(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    control_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    grid_side = Side(style="thin", color="D9D9D9")
    grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)
    widths = [28, 14, 58, 16, 16, 16, 16]

    for account, rows in accounts.items():
        sheet = workbook.create_sheet(sheet_title_for_account(account))
        sheet.append(HEADERS)
        for row in rows:
            sheet.append(
                [
                    row["Bank_Account"],
                    row["Date"],
                    row["Description"],
                    row["Deposit"],
                    row["Withdrawal"],
                    row["Balance"],
                    None,
                ]
            )

        for cell in sheet[1]:
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = header_fill
            cell.border = grid_border

        for row_idx in range(2, sheet.max_row + 1):
            if row_idx == 2:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )
            else:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'G{row_idx - 1}+IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10, bold=cell.column == 7 or cell.row == 1)
                cell.border = grid_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if cell.column == 7:
                    cell.fill = control_fill

        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in sheet["B"][1:]:
            cell.number_format = "dd mmm yyyy"
        for column in ["D", "E"]:
            for cell in sheet[column][1:]:
                cell.number_format = '#,##0.00;(#,##0.00);"-"'
        for column in ["F", "G"]:
            for cell in sheet[column][1:]:
                cell.number_format = "#,##0.00;(#,##0.00);0.00"

        sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def validate_accounts(accounts):
    report = []
    for account, rows in accounts.items():
        deposit_total = round(sum(row["Deposit"] or 0 for row in rows), 2)
        withdrawal_total = round(sum(row["Withdrawal"] or 0 for row in rows), 2)
        final_balance = rows[-1]["Balance"] if rows else None
        final_control = rows[-1]["Control"] if rows else None
        balance_mismatches = [
            idx + 2
            for idx, row in enumerate(rows)
            if row["Balance"] is not None and abs(row["Balance"] - row["Control"]) > 0.01
        ]
        report.append(
            {
                "account": account,
                "rows": len(rows),
                "deposit_count": sum(1 for row in rows if row["Deposit"] is not None),
                "withdrawal_count": sum(1 for row in rows if row["Withdrawal"] is not None),
                "deposit_total": deposit_total,
                "withdrawal_total": withdrawal_total,
                "final_balance": final_balance,
                "final_control": final_control,
                "balance_mismatches": balance_mismatches,
            }
        )
    return report


def convert_pdf_to_excel(pdf_path, output_dir=None):
    pdf_path = Path(pdf_path)
    output_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
    accounts = extract_pdf(pdf_path)
    output_path = output_dir / f"{pdf_path.stem}.xlsx"
    return write_workbook(accounts, output_path)


def collect_pdf_paths(input_path):
    input_path = Path(input_path)
    if input_path.is_file():
        if input_path.suffix.lower() != ".pdf":
            raise ValueError(f"Input file is not a PDF: {input_path}")
        return [input_path]
    if input_path.is_dir():
        return sorted(path for path in input_path.iterdir() if path.is_file() and path.suffix.lower() == ".pdf")
    raise FileNotFoundError(f"Input path does not exist: {input_path}")


def build_arg_parser():
    parser = argparse.ArgumentParser(description="Convert BOC statement PDF to Excel.")
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        default=DEFAULT_INPUT_PATH,
        help=f"PDF file or folder of PDFs. Default: {DEFAULT_INPUT_PATH}",
    )
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help=f"Output directory. Default: {DEFAULT_OUTPUT_DIR}")
    return parser


def main(argv=None):
    args = build_arg_parser().parse_args(argv)
    pdf_paths = collect_pdf_paths(args.input)
    if not pdf_paths:
        print(f"No PDF files found in: {args.input}")
        return 1

    for pdf_path in pdf_paths:
        accounts = extract_pdf(pdf_path)
        output_path = write_workbook(accounts, args.output_dir / f"{pdf_path.stem}.xlsx")
        print(f"Saved: {output_path}")
        for item in validate_accounts(accounts):
            print(
                f"{item['account']}: rows={item['rows']}, "
                f"deposits={item['deposit_count']} amount={item['deposit_total']:,.2f}, "
                f"withdrawals={item['withdrawal_count']} amount={item['withdrawal_total']:,.2f}, "
                f"final_balance={item['final_balance']:,.2f}, final_control={item['final_control']:,.2f}, "
                f"balance_mismatches={len(item['balance_mismatches'])}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
