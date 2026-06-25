import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = ["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"]
DEFAULT_INPUT_PATH = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "generated_output"
DATE_RE = re.compile(r"^\d{4}/\d{2}/\d{2}$")
AMOUNT_RE = re.compile(r"^\(?\d{1,3}(?:,\d{3})*(?:\.\d{2})\)?$|^\(?\d+\.\d{2}\)?$")
ACCOUNT_RE = re.compile(r"(.+?)\((\d{8,})\)\s+([A-Z]{3})$")


def clean_text(text):
    return " ".join((text or "").replace("(cid:10)", " ").replace("(cid:13)", " ").split())


def amount_value(text):
    sign = -1 if text.startswith("(") and text.endswith(")") else 1
    return sign * float(text.strip("()").replace(",", ""))


def nonzero(value):
    if value is None or abs(value) < 0.005:
        return None
    return value


def parse_date(text):
    return datetime.strptime(text, "%Y/%m/%d")


def account_name(currency, account_type):
    normalized = " ".join(account_type.split())
    if "往來" in normalized or "CURRENT" in normalized.upper():
        normalized = "Current Account"
    elif "儲蓄" in normalized or "SAVINGS" in normalized.upper():
        normalized = "Savings Account"
    return f"NCB {currency} {normalized}"


def sheet_title_for_account(account):
    return account[:31]


def group_words_by_line(page):
    rows = defaultdict(list)
    words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False)
    for word in words:
        y = round(word["top"] / 3) * 3
        rows[y].append(word)
    return [(y, sorted(rows[y], key=lambda item: item["x0"])) for y in sorted(rows)]


def line_text(words):
    return clean_text(" ".join(word["text"] for word in words))


def words_in_range(words, start_x, end_x):
    return [word for word in words if word["x0"] >= start_x and word["x0"] < end_x]


def first_amount(words):
    for word in words:
        if AMOUNT_RE.match(word["text"]):
            return amount_value(word["text"])
    return None


def append_description(row, extra):
    extra = clean_text(extra)
    if extra:
        row["Description"] = f"{row['Description']} {extra}".strip()


def is_footer_or_notice(text):
    markers = [
        "重要事項",
        "請查看及核對",
        "南商提醒",
        "地址：",
        "客戶服務熱線",
        "網址：",
        "多謝閣下使用",
        "防騙",
    ]
    return any(marker in text for marker in markers)


def extract_pdf(pdf_path):
    accounts = defaultdict(list)
    current_account = None
    in_transactions = False
    last_row = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for _y, words in group_words_by_line(page):
                text = line_text(words)
                if not text:
                    continue

                if "賬戶交易詳情" in text:
                    in_transactions = True
                    last_row = None
                    continue
                if not in_transactions:
                    continue
                if is_footer_or_notice(text):
                    in_transactions = False
                    last_row = None
                    continue
                if text in {"往來賬戶"} or "交易日期" in text and "交易摘要" in text:
                    continue

                account_match = ACCOUNT_RE.match(text)
                if account_match:
                    account_type = account_match.group(1)
                    currency = account_match.group(3)
                    current_account = account_name(currency, account_type)
                    last_row = None
                    continue
                if not current_account:
                    continue

                parts = [word["text"] for word in words]
                if not parts:
                    continue
                if not DATE_RE.match(parts[0]):
                    if last_row is not None and min(word["x0"] for word in words) >= 190:
                        append_description(last_row, text)
                    continue

                date = parse_date(parts[0])
                desc = line_text(words_in_range(words, 190, 300))
                if not desc:
                    desc = "BALANCE"

                balance = first_amount(words_in_range(words, 500, 590))
                if balance is None:
                    continue

                if "今期結餘" in desc:
                    last_row = None
                    continue

                deposit = first_amount(words_in_range(words, 300, 365))
                withdrawal = first_amount(words_in_range(words, 365, 500))
                if "承前結餘" in desc:
                    deposit = None
                    withdrawal = None
                    desc = "B/F BALANCE"

                row = {
                    "Bank_Account": current_account,
                    "Date": date,
                    "Description": desc,
                    "Deposit": nonzero(deposit),
                    "Withdrawal": nonzero(withdrawal),
                    "Balance": balance,
                }
                accounts[current_account].append(row)
                last_row = row

    for rows in accounts.values():
        control = None
        for row in rows:
            if row["Deposit"] is None and row["Withdrawal"] is None and row["Balance"] is not None:
                control = row["Balance"]
            else:
                if control is None:
                    control = row["Balance"] or 0.0
                control += row["Deposit"] or 0.0
                control -= row["Withdrawal"] or 0.0
            row["Control"] = round(control, 2) if control is not None else None

    return dict(accounts)


def write_workbook(accounts, output_path):
    output_path = Path(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    control_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    grid_side = Side(style="thin", color="D9D9D9")
    grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)
    widths = [30, 14, 58, 16, 16, 16, 16]

    for account, rows in accounts.items():
        sheet = workbook.create_sheet(sheet_title_for_account(account))
        sheet.append(HEADERS)
        for row in rows:
            sheet.append([
                row["Bank_Account"],
                row["Date"],
                row["Description"],
                row["Deposit"],
                row["Withdrawal"],
                row["Balance"],
                None,
            ])
        for cell in sheet[1]:
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = header_fill
            cell.border = grid_border
        for row_idx in range(2, sheet.max_row + 1):
            if row_idx == 2:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )
            else:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'G{row_idx - 1}+IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10, bold=cell.column == 7 or cell.row == 1)
                cell.border = grid_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if cell.column == 7:
                    cell.fill = control_fill
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in sheet["B"][1:]:
            cell.number_format = "dd mmm yyyy"
        for cell in sheet["D"][1:]:
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        for cell in sheet["E"][1:]:
            cell.number_format = '(#,##0.00);(#,##0.00);"-"'
        for column in ["F", "G"]:
            for cell in sheet[column][1:]:
                cell.number_format = "#,##0.00;(#,##0.00);0.00"
        sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def validate_accounts(accounts):
    report = []
    for account, rows in accounts.items():
        deposit_total = round(sum(row["Deposit"] or 0 for row in rows), 2)
        withdrawal_total = round(sum(row["Withdrawal"] or 0 for row in rows), 2)
        final_balance = rows[-1]["Balance"] if rows else None
        final_control = rows[-1]["Control"] if rows else None
        balance_mismatches = [
            idx + 2
            for idx, row in enumerate(rows)
            if row["Balance"] is not None and abs(row["Balance"] - row["Control"]) > 0.01
        ]
        report.append({
            "account": account,
            "rows": len(rows),
            "deposit_count": sum(1 for row in rows if row["Deposit"] is not None),
            "withdrawal_count": sum(1 for row in rows if row["Withdrawal"] is not None),
            "deposit_total": deposit_total,
            "withdrawal_total": withdrawal_total,
            "final_balance": final_balance,
            "final_control": final_control,
            "balance_mismatches": balance_mismatches,
        })
    return report


def collect_pdf_paths(input_path):
    input_path = Path(input_path)
    if input_path.is_file():
        return [input_path]
    return sorted(path for path in input_path.iterdir() if path.is_file() and path.suffix.lower() == ".pdf")


def main(argv=None):
    parser = argparse.ArgumentParser(description="Convert NCB statement PDF to Excel.")
    parser.add_argument("input", type=Path, nargs="?", default=DEFAULT_INPUT_PATH)
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args(argv)
    for pdf_path in collect_pdf_paths(args.input):
        accounts = extract_pdf(pdf_path)
        output_path = write_workbook(accounts, args.output_dir / f"{pdf_path.stem}.xlsx")
        print(f"Saved: {output_path}")
        for item in validate_accounts(accounts):
            print(
                f"{item['account']}: rows={item['rows']}, deposits={item['deposit_count']} amount={item['deposit_total']:,.2f}, "
                f"withdrawals={item['withdrawal_count']} amount={item['withdrawal_total']:,.2f}, "
                f"final_balance={item['final_balance']:,.2f}, final_control={item['final_control']:,.2f}, "
                f"balance_mismatches={len(item['balance_mismatches'])}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
