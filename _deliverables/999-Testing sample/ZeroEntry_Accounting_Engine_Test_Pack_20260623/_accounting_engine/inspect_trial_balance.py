from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def fmt(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "date"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def used_bounds(ws) -> dict[str, int] | None:
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if min_row is None:
        return None
    return {"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col}


def nonempty_rows(ws, max_rows: int = 80, max_cols: int = 20) -> list[dict[str, Any]]:
    rows = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        cells = []
        for c in range(1, min(ws.max_column, max_cols) + 1):
            value = ws.cell(r, c).value
            if value not in (None, ""):
                cells.append({"cell": f"{get_column_letter(c)}{r}", "value": fmt(value)})
        if cells:
            rows.append({"row": r, "cells": cells})
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out-json", type=Path, required=True)
    parser.add_argument("--out-md", type=Path, required=True)
    args = parser.parse_args()

    wb_formula = load_workbook(args.workbook, data_only=False)
    wb_values = load_workbook(args.workbook, data_only=True)
    summary = {
        "path": str(args.workbook),
        "sheets": [],
    }
    md = ["# Trial Balance Inspection", ""]
    for ws in wb_formula.worksheets:
        value_ws = wb_values[ws.title]
        sheet = {
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "used_bounds": used_bounds(ws),
            "merged_ranges": [str(r) for r in list(ws.merged_cells.ranges)[:20]],
            "nonempty_rows_formula": nonempty_rows(ws),
            "nonempty_rows_values": nonempty_rows(value_ws),
        }
        summary["sheets"].append(sheet)
        md.extend([f"## {ws.title}", ""])
        md.append(f"- size: {ws.max_row} rows x {ws.max_column} columns")
        md.append(f"- used_bounds: {sheet['used_bounds']}")
        if sheet["merged_ranges"]:
            md.append(f"- merged_ranges: {'; '.join(sheet['merged_ranges'])}")
        md.append("- non-empty rows:")
        for row in sheet["nonempty_rows_values"][:40]:
            cells = "; ".join(f"{c['cell']}={c['value']}" for c in row["cells"])
            md.append(f"  - row {row['row']}: {cells}")
        md.append("")

    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    args.out_md.write_text("\n".join(md), encoding="utf-8")
    print(json.dumps({"sheets": [s["name"] for s in summary["sheets"]]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
