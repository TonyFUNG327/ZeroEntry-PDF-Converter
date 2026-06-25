from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HELPER_LABELS = {
    "date": {"date"},
    "reference": {"reference"},
    "description": {"description"},
    "bank_amount": {"bank (amount$)"},
    "nature_amount": {"nature (amount$)"},
}


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if not value:
            return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def read_chart_of_accounts(wb) -> tuple[dict[str, str], dict[str, str]]:
    if "Chart of Accounts" not in wb.sheetnames:
        return {}, {}
    ws = wb["Chart of Accounts"]
    name_to_code: dict[str, str] = {}
    code_to_name: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        name = text(ws.cell(row, 1).value)
        code = text(ws.cell(row, 2).value)
        if not name or not code:
            continue
        name_to_code[name.casefold()] = code
        code_to_name[code] = name
    return name_to_code, code_to_name


def account_code(name_to_code: dict[str, str], value: Any) -> str:
    raw = text(value)
    if not raw:
        return ""
    if raw.replace(".", "", 1).isdigit():
        return raw.split(".")[0]
    return name_to_code.get(raw.casefold(), "")


def find_header_row(ws) -> int | None:
    for row in range(1, min(ws.max_row, 40) + 1):
        labels = [text(ws.cell(row, col).value).casefold() for col in range(1, ws.max_column + 1)]
        joined = "|".join(labels)
        if "no." in joined and "date" in joined and "particular" in joined:
            return row
    return None


def find_helper_columns(ws, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    reference_col = None
    for col in range(1, ws.max_column + 1):
        label = text(ws.cell(header_row, col).value).casefold()
        sub_label = text(ws.cell(header_row + 1, col).value).casefold()
        if label == "reference" and col > 20:
            reference_col = col
            columns["reference"] = col
            columns["date"] = col - 1
        for key, candidates in HELPER_LABELS.items():
            if label in candidates and key not in columns and (key not in {"date", "reference"} or col > 20):
                columns[key] = col
        if sub_label in {"=c4", "=c5"}:
            columns.setdefault("bank_account", col)
        if sub_label == "(nature)":
            columns.setdefault("nature_account", col)
    if reference_col:
        gl_cols = [
            col
            for col in range(reference_col + 1, ws.max_column + 1)
            if text(ws.cell(header_row, col).value).casefold().startswith("g/l account")
        ]
        if gl_cols:
            columns.setdefault("bank_account", gl_cols[0])
        if len(gl_cols) > 1:
            columns.setdefault("nature_account", gl_cols[1])
    return columns


def make_description(nature: str, payer: str, invoice: str, details: str, month_year: Any) -> str:
    description = nature
    if payer:
        description += f":{payer}"
    for part in [invoice, details]:
        if part:
            description += f" - {part}"
    if month_year:
        if hasattr(month_year, "strftime"):
            description += f" - {month_year.strftime('%b-%y')}"
        else:
            description += f" - {month_year}"
    return description


def signed_line(
    *,
    source_sheet: str,
    tx_date: Any,
    reference: str,
    account_id: str,
    account_name: str,
    description: str,
    signed_amount: float,
    source_row: int,
) -> dict[str, Any]:
    amount = round(abs(float(signed_amount)), 2)
    return {
        "source_sheet": source_sheet,
        "source_row": source_row,
        "date": tx_date,
        "reference": reference,
        "account_id": account_id,
        "account_description": account_name,
        "description": description,
        "debit": amount if signed_amount > 0 else 0.0,
        "credit": amount if signed_amount < 0 else 0.0,
        "signed_amount": round(float(signed_amount), 2),
    }


def parse_bank_sheet(
    ws,
    *,
    name_to_code: dict[str, str],
    code_to_name: dict[str, str],
) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    header_row = find_header_row(ws)
    if not header_row:
        return [], [f"{ws.title}: no transaction header detected"], {}

    helper = find_helper_columns(ws, header_row)
    required = {"date", "reference", "description", "bank_amount", "nature_amount", "nature_account"}
    if not required.issubset(helper):
        missing = ", ".join(sorted(required - set(helper)))
        return [], [f"{ws.title}: missing helper columns: {missing}"], {}

    bank_name = text(ws["C4"].value) or text(ws["C5"].value) or ws.title
    bank_code = account_code(name_to_code, bank_name)
    lines: list[dict[str, Any]] = []
    warnings: list[str] = []
    transactions = 0
    bank_total = 0.0
    nature_total = 0.0

    for row in range(header_row + 2, ws.max_row + 1):
        tx_date = ws.cell(row, helper["date"]).value or ws.cell(row, 2).value
        nature = text(ws.cell(row, 3).value)
        if not tx_date and not nature:
            continue
        if nature.casefold() == "bank opening":
            continue
        bank_amount = number(ws.cell(row, helper["bank_amount"]).value)
        nature_amount = number(ws.cell(row, helper["nature_amount"]).value)
        if round(bank_amount + nature_amount, 2) != 0:
            warnings.append(
                f"{ws.title} row {row}: bank/nature amounts do not balance ({bank_amount} + {nature_amount})"
            )
        if bank_amount == 0 and nature_amount == 0:
            continue
        transactions += 1
        bank_total += bank_amount
        nature_total += nature_amount

        reference = text(ws.cell(row, helper["reference"]).value)
        if not reference:
            row_ref = text(ws.cell(row, 1).value)
            reference = f"{bank_name}-{row_ref}" if row_ref else f"{bank_name}-row{row}"
        description = text(ws.cell(row, helper["description"]).value)
        if not description:
            description = make_description(
                nature,
                text(ws.cell(row, 4).value),
                text(ws.cell(row, 5).value),
                text(ws.cell(row, 6).value),
                ws.cell(row, 7).value,
            )

        row_bank_code = account_code(name_to_code, ws.cell(row, helper["bank_account"]).value) if "bank_account" in helper else ""
        row_bank_code = row_bank_code or bank_code
        row_nature_code = account_code(name_to_code, ws.cell(row, helper["nature_account"]).value) or account_code(
            name_to_code, nature
        )
        if not row_bank_code:
            warnings.append(f"{ws.title} row {row}: missing bank account code for {bank_name}")
        if not row_nature_code:
            warnings.append(f"{ws.title} row {row}: missing nature account code for {nature}")

        lines.append(
            signed_line(
                source_sheet=ws.title,
                source_row=row,
                tx_date=tx_date,
                reference=reference,
                account_id=row_bank_code,
                account_name=code_to_name.get(row_bank_code, bank_name),
                description=description,
                signed_amount=bank_amount,
            )
        )
        lines.append(
            signed_line(
                source_sheet=ws.title,
                source_row=row,
                tx_date=tx_date,
                reference=reference,
                account_id=row_nature_code,
                account_name=code_to_name.get(row_nature_code, nature),
                description=description,
                signed_amount=nature_amount,
            )
        )

    return lines, warnings, {
        "source_sheet": ws.title,
        "bank_name": bank_name,
        "bank_account_id": bank_code,
        "transaction_count": transactions,
        "journal_line_count": len(lines),
        "bank_amount_total": round(bank_total, 2),
        "nature_amount_total": round(nature_total, 2),
        "net_check": round(bank_total + nature_total, 2),
    }


def summarize(lines: list[dict[str, Any]], sheet_summaries: list[dict[str, Any]], warnings: list[str]) -> dict[str, Any]:
    references: dict[str, float] = {}
    account_totals: dict[str, dict[str, Any]] = {}
    for line in lines:
        references[line["reference"]] = references.get(line["reference"], 0.0) + float(line["signed_amount"])
        account = account_totals.setdefault(
            line["account_id"],
            {
                "account_id": line["account_id"],
                "account_description": line["account_description"],
                "debit": 0.0,
                "credit": 0.0,
                "signed_amount": 0.0,
            },
        )
        account["debit"] += float(line["debit"])
        account["credit"] += float(line["credit"])
        account["signed_amount"] += float(line["signed_amount"])

    return {
        "sheet_count": len(sheet_summaries),
        "transaction_count": sum(int(item.get("transaction_count", 0)) for item in sheet_summaries),
        "journal_line_count": len(lines),
        "reference_count": len(references),
        "debit_total": round(sum(float(line["debit"]) for line in lines), 2),
        "credit_total": round(sum(float(line["credit"]) for line in lines), 2),
        "net_check": round(sum(float(line["signed_amount"]) for line in lines), 2),
        "unbalanced_references": {
            reference: round(amount, 2)
            for reference, amount in references.items()
            if round(amount, 2) != 0
        },
        "warning_count": len(warnings),
        "account_totals": [
            {
                **value,
                "debit": round(value["debit"], 2),
                "credit": round(value["credit"], 2),
                "signed_amount": round(value["signed_amount"], 2),
            }
            for _, value in sorted(account_totals.items(), key=lambda item: item[0])
        ],
    }


def parse_bank_record(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    name_to_code, code_to_name = read_chart_of_accounts(wb)
    lines: list[dict[str, Any]] = []
    warnings: list[str] = []
    sheet_summaries: list[dict[str, Any]] = []
    skipped_sheets: list[dict[str, str]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Chart of Accounts":
            continue
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            skipped_sheets.append(
                {
                    "source_sheet": sheet_name,
                    "reason": f"Hidden sheet ({ws.sheet_state}); excluded from Bank record journal output.",
                }
            )
            continue
        bank_name = text(ws["C4"].value) or text(ws["C5"].value) or sheet_name
        if sheet_name.casefold() == "finance costs" or bank_name.casefold() == "obligation under finance lease":
            skipped_sheets.append(
                {
                    "source_sheet": sheet_name,
                    "reason": "Non-bank schedule; excluded from Bank record journal output.",
                }
            )
            continue
        sheet_lines, sheet_warnings, sheet_summary = parse_bank_sheet(
            ws,
            name_to_code=name_to_code,
            code_to_name=code_to_name,
        )
        lines.extend(sheet_lines)
        warnings.extend(sheet_warnings)
        if sheet_summary:
            sheet_summaries.append(sheet_summary)

    return {
        "source_file": str(path),
        "journal_lines": lines,
        "sheet_summaries": sheet_summaries,
        "skipped_sheets": skipped_sheets,
        "warnings": warnings,
        "summary": summarize(lines, sheet_summaries, warnings),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse classified Bank record workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out-json", type=Path, required=True)
    args = parser.parse_args()

    payload = parse_bank_record(args.workbook)
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=json_default),
        encoding="utf-8",
    )
    print(json.dumps(payload["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
