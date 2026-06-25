from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RETAINED_EARNINGS_CODE = "3910"
RETAINED_EARNINGS_NAME = "Retained Earnings"


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if not value:
            return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def account_code_sort_key(code: str) -> tuple[int, str]:
    try:
        return int(code), code
    except ValueError:
        return 999999, code


def is_account_row(account_id: str, description: str) -> bool:
    if not account_id or not description:
        return False
    if account_id.casefold() in {"account id", "total:", "director:", "date:"}:
        return False
    return any(ch.isdigit() for ch in account_id)


def classify_account(account_id: str) -> str:
    try:
        code = int(account_id)
    except ValueError:
        return "Unknown"
    if code < 4000:
        return "Balance Sheet"
    return "Profit and Loss"


def find_tb_sheet(wb) -> str:
    for name in wb.sheetnames:
        if "tb" in name.casefold() or "trial" in name.casefold():
            return name
    return wb.sheetnames[0]


def expected_tb_sheet_name(closing_date: str) -> str:
    parsed = datetime.fromisoformat(closing_date).date()
    return f"TB {parsed.strftime('%d.%m.%y')}".lstrip("0").replace(".0", ".")


def parse_trial_balance(path: Path, closing_date: str, opening_date: str) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    sheet_name = find_tb_sheet(wb)
    ws = wb[sheet_name]

    tb_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    for row in range(1, ws.max_row + 1):
        account_id = text(ws.cell(row, 1).value)
        description = text(ws.cell(row, 2).value)
        if not is_account_row(account_id, description):
            continue
        debit = number(ws.cell(row, 3).value)
        credit = number(ws.cell(row, 4).value)
        signed = debit + credit
        account_class = classify_account(account_id)
        tb_rows.append(
            {
                "source_row": row,
                "account_id": account_id,
                "account_description": description,
                "closing_debit": debit,
                "closing_credit": credit,
                "closing_signed_amount": signed,
                "account_class": account_class,
            }
        )

    bs_rows = [row for row in tb_rows if row["account_class"] == "Balance Sheet"]
    pl_rows = [row for row in tb_rows if row["account_class"] == "Profit and Loss"]

    bs_by_code = {row["account_id"]: dict(row) for row in bs_rows}
    pl_signed = sum(float(row["closing_signed_amount"]) for row in pl_rows)
    retained = bs_by_code.get(RETAINED_EARNINGS_CODE)
    if retained is None:
        retained = {
            "source_row": None,
            "account_id": RETAINED_EARNINGS_CODE,
            "account_description": RETAINED_EARNINGS_NAME,
            "closing_debit": 0.0,
            "closing_credit": 0.0,
            "closing_signed_amount": 0.0,
            "account_class": "Balance Sheet",
        }
        bs_by_code[RETAINED_EARNINGS_CODE] = retained
        warnings.append("Retained Earnings account was not present; created account 3910.")

    retained["opening_adjustment_from_pl"] = round(pl_signed, 2)
    retained["opening_signed_amount"] = round(
        float(retained["closing_signed_amount"]) + pl_signed, 2
    )

    opening_rows: list[dict[str, Any]] = []
    for account_id, row in sorted(bs_by_code.items(), key=lambda item: account_code_sort_key(item[0])):
        opening_signed = row.get("opening_signed_amount", row["closing_signed_amount"])
        opening_rows.append(
            {
                "opening_date": opening_date,
                "account_id": account_id,
                "account_description": row["account_description"],
                "opening_signed_amount": round(float(opening_signed), 2),
                "opening_debit": round(float(opening_signed), 2) if opening_signed > 0 else 0.0,
                "opening_credit": round(float(opening_signed), 2) if opening_signed < 0 else 0.0,
                "source": "Prior-year closing TB",
                "note": (
                    "P&L closed into Retained Earnings"
                    if account_id == RETAINED_EARNINGS_CODE
                    else "Carried forward from Balance Sheet"
                ),
            }
        )

    summary = {
        "source_file": str(path),
        "source_sheet": sheet_name,
        "requested_closing_date": closing_date,
        "opening_date": opening_date,
        "tb_row_count": len(tb_rows),
        "balance_sheet_row_count": len(bs_rows),
        "profit_and_loss_row_count": len(pl_rows),
        "closing_debit_total": round(sum(float(row["closing_debit"]) for row in tb_rows), 2),
        "closing_credit_total": round(sum(float(row["closing_credit"]) for row in tb_rows), 2),
        "closing_net_check": round(sum(float(row["closing_signed_amount"]) for row in tb_rows), 2),
        "pl_signed_amount_closed_to_retained_earnings": round(pl_signed, 2),
        "opening_debit_total": round(sum(float(row["opening_debit"]) for row in opening_rows), 2),
        "opening_credit_total": round(sum(float(row["opening_credit"]) for row in opening_rows), 2),
        "opening_net_check": round(
            sum(float(row["opening_signed_amount"]) for row in opening_rows), 2
        ),
    }
    expected_sheet_name = expected_tb_sheet_name(closing_date)
    if sheet_name != expected_sheet_name:
        warnings.append(
            f"Workbook sheet name is '{sheet_name}', while expected sheet name is '{expected_sheet_name}' for requested closing date {closing_date}."
        )

    return {
        "summary": summary,
        "warnings": warnings,
        "trial_balance_rows": tb_rows,
        "opening_balance_rows": opening_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert prior-year TB to opening balances.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--closing-date", default="2024-03-31")
    parser.add_argument("--opening-date", default="2024-04-01")
    parser.add_argument("--out-json", type=Path, required=True)
    args = parser.parse_args()

    payload = parse_trial_balance(args.workbook, args.closing_date, args.opening_date)
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=json_default),
        encoding="utf-8",
    )
    print(json.dumps(payload["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
