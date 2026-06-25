from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def sort_key(account_id: str) -> tuple[int, str]:
    try:
        return int(account_id), account_id
    except ValueError:
        return 999999, account_id


def account_class(account_id: str) -> str:
    try:
        code = int(account_id)
    except ValueError:
        return "Unknown"
    if code < 4000:
        return "Balance Sheet"
    if 4000 <= code < 5000:
        return "Revenue"
    return "Expense"


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def signed_to_debit_credit(signed: float) -> tuple[float, float]:
    signed = round(float(signed), 2)
    if signed > 0:
        return signed, 0.0
    if signed < 0:
        return 0.0, abs(signed)
    return 0.0, 0.0


def parse_sample_tb(report_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(report_path, data_only=True)
    sheet_name = next(name for name in wb.sheetnames if name.startswith("TB"))
    ws = wb[sheet_name]
    rows: dict[str, dict[str, Any]] = {}
    for row in range(2, ws.max_row + 1):
        account_id = text(ws.cell(row, 1).value)
        name = text(ws.cell(row, 2).value)
        if not account_id or not name or account_id.casefold() == "total:":
            continue
        debit = number(ws.cell(row, 3).value)
        credit = number(ws.cell(row, 4).value)
        signed = round(debit + credit, 2)
        rows[account_id] = {
            "account_id": account_id,
            "account_description": name,
            "sample_debit": debit,
            "sample_credit": credit,
            "sample_signed_amount": signed,
        }
    return rows


def parse_sample_manual_jv(report_path: Path) -> dict[str, Any]:
    wb = load_workbook(report_path, data_only=True)
    sheet_name = next(name for name in wb.sheetnames if name.startswith("GL"))
    ws = wb[sheet_name]
    totals: dict[str, dict[str, Any]] = {}
    refs = set()
    for row in range(2, ws.max_row + 1):
        ref = text(ws.cell(row, 4).value)
        if not ref.startswith("JV"):
            continue
        account_id = text(ws.cell(row, 1).value)
        account_name = text(ws.cell(row, 2).value)
        debit = number(ws.cell(row, 7).value)
        credit = number(ws.cell(row, 8).value)
        signed = round(debit - credit, 2)
        refs.add(ref)
        bucket = totals.setdefault(
            account_id,
            {
                "account_id": account_id,
                "account_description": account_name,
                "manual_jv_signed_amount": 0.0,
                "manual_jv_debit": 0.0,
                "manual_jv_credit": 0.0,
                "references": set(),
            },
        )
        bucket["manual_jv_signed_amount"] += signed
        bucket["manual_jv_debit"] += debit
        bucket["manual_jv_credit"] += credit
        bucket["references"].add(ref)
    return {
        "references": sorted(refs),
        "totals": {
            account_id: {
                **value,
                "manual_jv_signed_amount": round(value["manual_jv_signed_amount"], 2),
                "manual_jv_debit": round(value["manual_jv_debit"], 2),
                "manual_jv_credit": round(value["manual_jv_credit"], 2),
                "references": sorted(value["references"]),
            }
            for account_id, value in totals.items()
        },
    }


def normalize_source_line(line: dict[str, Any], source_label: str) -> dict[str, Any]:
    account_id = text(line.get("account_id") or line.get("account_code"))
    account_description = text(line.get("account_description") or line.get("account_name"))
    return {
        "source": source_label,
        "date": line.get("date"),
        "reference": text(line.get("reference")),
        "jrnl": text(line.get("jrnl")) or "GENJ",
        "account_id": account_id,
        "account_description": account_description,
        "description": text(line.get("description")),
        "debit": number(line.get("debit")),
        "credit": number(line.get("credit")),
        "signed_amount": number(line.get("signed_amount")),
    }


def add_payload_lines(
    *,
    journal_lines: list[dict[str, Any]],
    account_names: dict[str, str],
    payloads: list[dict[str, Any]],
    source_label: str,
) -> None:
    for payload in payloads:
        for line in payload.get("journal_lines", []):
            normalized = normalize_source_line(line, source_label)
            if not normalized["account_id"]:
                continue
            account_names.setdefault(normalized["account_id"], normalized["account_description"])
            journal_lines.append(normalized)


def build_combined_payload(
    opening_payload: dict[str, Any],
    bank_payloads: list[dict[str, Any]],
    cash_payloads: list[dict[str, Any]],
    manual_payloads: list[dict[str, Any]],
    report_path: Path | None,
) -> dict[str, Any]:
    journal_lines: list[dict[str, Any]] = []
    account_names: dict[str, str] = {}

    for row in opening_payload["opening_balance_rows"]:
        account_id = text(row["account_id"])
        signed = number(row["opening_signed_amount"])
        debit, credit = signed_to_debit_credit(signed)
        account_names[account_id] = text(row["account_description"])
        journal_lines.append(
            {
                "source": "Opening",
                "date": row["opening_date"],
                "reference": "Opening balance",
                "jrnl": "GENJ",
                "account_id": account_id,
                "account_description": text(row["account_description"]),
                "description": f"Opening balance - {row['account_description']}",
                "debit": debit,
                "credit": credit,
                "signed_amount": signed,
            }
        )

    add_payload_lines(
        journal_lines=journal_lines,
        account_names=account_names,
        payloads=bank_payloads,
        source_label="Bank",
    )
    add_payload_lines(
        journal_lines=journal_lines,
        account_names=account_names,
        payloads=cash_payloads,
        source_label="Cash",
    )
    add_payload_lines(
        journal_lines=journal_lines,
        account_names=account_names,
        payloads=manual_payloads,
        source_label="Manual Journals",
    )

    totals = defaultdict(float)
    for line in journal_lines:
        totals[line["account_id"]] += number(line["signed_amount"])

    sample_tb = parse_sample_tb(report_path) if report_path else {}
    manual_jv = parse_sample_manual_jv(report_path) if report_path else {"references": [], "totals": {}}
    for account_id, row in sample_tb.items():
        account_names.setdefault(account_id, row["account_description"])
    for account_id, row in manual_jv["totals"].items():
        account_names.setdefault(account_id, row["account_description"])

    tb_rows = []
    for account_id in sorted(account_names, key=sort_key):
        signed = round(totals.get(account_id, 0.0), 2)
        if signed == 0 and account_id not in sample_tb:
            continue
        debit, credit = signed_to_debit_credit(signed)
        tb_rows.append(
            {
                "account_id": account_id,
                "account_description": account_names[account_id],
                "debit": debit,
                "credit": credit,
                "signed_amount": signed,
                "account_class": account_class(account_id),
            }
        )

    comparison = []
    all_accounts = sorted(set(account_names) | set(sample_tb), key=sort_key)
    all_diffs_match_jv = True
    material_non_jv = []
    if report_path:
        for account_id in all_accounts:
            generated = round(totals.get(account_id, 0.0), 2)
            sample = round(sample_tb.get(account_id, {}).get("sample_signed_amount", 0.0), 2)
            diff = round(sample - generated, 2)
            jv_diff = round(manual_jv["totals"].get(account_id, {}).get("manual_jv_signed_amount", 0.0), 2)
            unexplained = round(diff - jv_diff, 2)
            if abs(unexplained) > 0.01:
                all_diffs_match_jv = False
                material_non_jv.append(account_id)
            comparison.append(
                {
                    "account_id": account_id,
                    "account_description": account_names.get(account_id, sample_tb.get(account_id, {}).get("account_description", "")),
                    "generated_signed_amount": generated,
                    "sample_signed_amount": sample,
                    "difference_sample_less_generated": diff,
                    "sample_manual_jv_signed_amount": jv_diff,
                    "unexplained_difference": unexplained,
                    "matched_to_manual_jv": abs(unexplained) <= 0.01,
                    "manual_jv_references": ", ".join(manual_jv["totals"].get(account_id, {}).get("references", [])),
                }
            )
    else:
        all_diffs_match_jv = None

    gl_rows = sorted(
        journal_lines,
        key=lambda x: (
            sort_key(x["account_id"]),
            text(x.get("date")),
            text(x.get("reference")),
            number(x.get("credit")),
        ),
    )

    revenue_rows = [r for r in tb_rows if r["account_class"] == "Revenue"]
    expense_rows = [r for r in tb_rows if r["account_class"] == "Expense"]
    bs_rows = [r for r in tb_rows if r["account_class"] == "Balance Sheet"]
    revenue_total = round(sum(abs(r["signed_amount"]) for r in revenue_rows if r["signed_amount"] < 0), 2)
    expense_total = round(sum(r["signed_amount"] for r in expense_rows if r["signed_amount"] > 0), 2)
    net_income = round(revenue_total - expense_total, 2)

    assets = [r for r in bs_rows if sort_key(r["account_id"])[0] < 2000]
    liabilities_equity = [r for r in bs_rows if sort_key(r["account_id"])[0] >= 2000]
    asset_total = round(sum(r["signed_amount"] for r in assets), 2)
    liability_equity_total = round(sum(abs(r["signed_amount"]) for r in liabilities_equity if r["signed_amount"] < 0), 2)
    bs_check = round(asset_total - liability_equity_total, 2)

    return {
        "summary": {
            "source": "Opening balances + Bank/Cash/Manual Journals",
            "bank_payload_count": len(bank_payloads),
            "cash_payload_count": len(cash_payloads),
            "manual_payload_count": len(manual_payloads),
            "journal_line_count": len(gl_rows),
            "tb_debit_total": round(sum(r["debit"] for r in tb_rows), 2),
            "tb_credit_total": round(sum(r["credit"] for r in tb_rows), 2),
            "tb_net_check": round(sum(r["signed_amount"] for r in tb_rows), 2),
            "sample_manual_jv_references": manual_jv["references"],
            "differences_match_sample_manual_jv": all_diffs_match_jv,
            "non_jv_difference_accounts": material_non_jv,
            "revenue_total": revenue_total,
            "expense_total": expense_total,
            "net_income": net_income,
            "asset_total": asset_total,
            "liability_equity_total": liability_equity_total,
            "bs_check_before_current_income": bs_check,
        },
        "gl_rows": gl_rows,
        "tb_rows": tb_rows,
        "is_rows": {
            "revenue": revenue_rows,
            "expense": expense_rows,
            "revenue_total": revenue_total,
            "expense_total": expense_total,
            "net_income": net_income,
        },
        "bs_rows": {
            "assets": assets,
            "liabilities_equity": liabilities_equity,
            "asset_total": asset_total,
            "liability_equity_total": liability_equity_total,
            "check_before_current_income": bs_check,
        },
        "tb_comparison": comparison,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Combine opening balances plus Bank/Cash/Manual journal payloads for a full-period accounting run.")
    parser.add_argument("--opening-json", type=Path, required=True)
    parser.add_argument("--bank-json", type=Path, action="append", default=[])
    parser.add_argument("--cash-json", type=Path, action="append", default=[])
    parser.add_argument("--manual-json", type=Path, action="append", default=[])
    parser.add_argument("--sample-report", type=Path)
    parser.add_argument("--out-json", type=Path, required=True)
    args = parser.parse_args()

    payload = build_combined_payload(
        load_json(args.opening_json),
        [load_json(path) for path in args.bank_json],
        [load_json(path) for path in args.cash_json],
        [load_json(path) for path in args.manual_json],
        args.sample_report,
    )
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=json_default),
        encoding="utf-8",
    )
    print(json.dumps(payload["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
