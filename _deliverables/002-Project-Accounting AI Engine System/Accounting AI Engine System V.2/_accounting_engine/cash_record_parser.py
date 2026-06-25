from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if not value:
            return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def read_chart_of_accounts(wb) -> tuple[dict[str, str], dict[str, str]]:
    if "Chart of Accounts" not in wb.sheetnames:
        return {}, {}
    ws = wb["Chart of Accounts"]
    name_to_code: dict[str, str] = {}
    code_to_name: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        name = text(ws.cell(row, 1).value)
        code = text(ws.cell(row, 2).value)
        if not name or not code:
            continue
        name_to_code[name.casefold()] = code
        code_to_name[code] = name
    return name_to_code, code_to_name


def account_code(name_to_code: dict[str, str], value: Any) -> str:
    raw = text(value)
    if not raw:
        return ""
    if raw.replace(".", "", 1).isdigit():
        return raw.split(".")[0]
    return name_to_code.get(raw.casefold(), "")


def find_header_row(ws) -> int | None:
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [text(ws.cell(row, col).value).casefold() for col in range(1, ws.max_column + 1)]
        joined = "|".join(values)
        if "no." in joined and "date" in joined and "nature" in joined:
            return row
    return None


def find_amount_columns(ws, header_row: int) -> tuple[int, int, int]:
    labels = {text(ws.cell(header_row - 1, col).value).casefold(): col for col in range(1, ws.max_column + 1)}
    cash_in = labels.get("cash in") or labels.get("deposit") or 7
    cash_out = labels.get("cash out") or labels.get("withdrawal") or cash_in + 1
    control = labels.get("control") or cash_out + 3
    return cash_in, cash_out, control


def make_description(nature: str, payer: str, invoice: str, details: str, month_year: Any = None) -> str:
    description = nature
    if payer:
        description += f":{payer}"
    for part in [invoice, details]:
        if part:
            description += f" - {part}"
    if month_year:
        if hasattr(month_year, "strftime"):
            description += f" - {month_year.strftime('%b-%y')}"
        else:
            description += f" - {month_year}"
    return description


def signed_line(
    *,
    source_sheet: str,
    source_row: int,
    tx_date: Any,
    reference: str,
    account_id: str,
    account_name: str,
    description: str,
    signed_amount: float,
) -> dict[str, Any]:
    amount = round(abs(float(signed_amount)), 2)
    return {
        "source_sheet": source_sheet,
        "source_row": source_row,
        "date": tx_date,
        "reference": reference,
        "account_id": account_id,
        "account_description": account_name,
        "description": description,
        "debit": amount if signed_amount > 0 else 0.0,
        "credit": amount if signed_amount < 0 else 0.0,
        "signed_amount": round(float(signed_amount), 2),
    }


def parse_cash_sheet(ws, *, name_to_code: dict[str, str], code_to_name: dict[str, str]) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    header_row = find_header_row(ws)
    if not header_row:
        return [], [f"{ws.title}: no transaction header detected"], {}

    record_account_name = text(ws["C4"].value) or text(ws["C5"].value) or ws.title
    reference_name = text(ws["C5"].value) or ws.title
    record_account_id = account_code(name_to_code, record_account_name)
    cash_in_col, cash_out_col, control_col = find_amount_columns(ws, header_row)
    month_year_col = 7 if text(ws.cell(header_row, 7).value).casefold() == "month-year" else None

    warnings: list[str] = []
    lines: list[dict[str, Any]] = []
    transaction_count = 0
    record_total = 0.0
    nature_total = 0.0

    if not record_account_id:
        warnings.append(f"{ws.title}: missing record account code for {record_account_name}")

    for row in range(header_row + 1, ws.max_row + 1):
        tx_date = ws.cell(row, 2).value
        nature = text(ws.cell(row, 3).value)
        if not tx_date and not nature:
            continue
        if nature.casefold() in {"cash opening", "bank opening"}:
            continue

        cash_in = number(ws.cell(row, cash_in_col).value)
        cash_out = number(ws.cell(row, cash_out_col).value)
        amount = round(cash_in + cash_out, 2)
        if amount == 0:
            continue

        control = number(ws.cell(row, control_col).value)
        if round(control, 2) != 0:
            warnings.append(f"{ws.title} row {row}: control is not zero ({control})")

        nature_account_id = account_code(name_to_code, nature)
        if not nature_account_id:
            warnings.append(f"{ws.title} row {row}: missing nature account code for {nature}")

        row_ref = text(ws.cell(row, 1).value)
        reference = f"{reference_name}-{row_ref}" if row_ref else f"{reference_name}-row{row}"
        description = make_description(
            nature,
            text(ws.cell(row, 4).value),
            text(ws.cell(row, 5).value),
            text(ws.cell(row, 6).value),
            ws.cell(row, month_year_col).value if month_year_col else None,
        )
        transaction_count += 1
        record_total += amount
        nature_total -= amount

        lines.append(
            signed_line(
                source_sheet=ws.title,
                source_row=row,
                tx_date=tx_date,
                reference=reference,
                account_id=record_account_id,
                account_name=code_to_name.get(record_account_id, record_account_name),
                description=description,
                signed_amount=amount,
            )
        )
        lines.append(
            signed_line(
                source_sheet=ws.title,
                source_row=row,
                tx_date=tx_date,
                reference=reference,
                account_id=nature_account_id,
                account_name=code_to_name.get(nature_account_id, nature),
                description=description,
                signed_amount=-amount,
            )
        )

    return lines, warnings, {
        "source_sheet": ws.title,
        "record_account_name": record_account_name,
        "record_account_id": record_account_id,
        "transaction_count": transaction_count,
        "journal_line_count": len(lines),
        "record_amount_total": round(record_total, 2),
        "nature_amount_total": round(nature_total, 2),
        "net_check": round(record_total + nature_total, 2),
    }


def summarize(lines: list[dict[str, Any]], sheet_summaries: list[dict[str, Any]], warnings: list[str]) -> dict[str, Any]:
    references: dict[str, float] = {}
    for line in lines:
        references[line["reference"]] = references.get(line["reference"], 0.0) + float(line["signed_amount"])
    return {
        "sheet_count": len(sheet_summaries),
        "transaction_count": sum(int(item.get("transaction_count", 0)) for item in sheet_summaries),
        "journal_line_count": len(lines),
        "reference_count": len(references),
        "debit_total": round(sum(float(line["debit"]) for line in lines), 2),
        "credit_total": round(sum(float(line["credit"]) for line in lines), 2),
        "net_check": round(sum(float(line["signed_amount"]) for line in lines), 2),
        "unbalanced_references": {
            ref: round(total, 2) for ref, total in references.items() if round(total, 2) != 0
        },
        "warning_count": len(warnings),
    }


def parse_cash_record(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    name_to_code, code_to_name = read_chart_of_accounts(wb)
    lines: list[dict[str, Any]] = []
    warnings: list[str] = []
    sheet_summaries: list[dict[str, Any]] = []
    skipped_sheets: list[dict[str, str]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Chart of Accounts":
            continue
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            skipped_sheets.append({"source_sheet": sheet_name, "reason": f"Hidden sheet ({ws.sheet_state})"})
            continue
        sheet_lines, sheet_warnings, sheet_summary = parse_cash_sheet(
            ws, name_to_code=name_to_code, code_to_name=code_to_name
        )
        lines.extend(sheet_lines)
        warnings.extend(sheet_warnings)
        if sheet_summary:
            sheet_summaries.append(sheet_summary)

    return {
        "source_file": str(path),
        "journal_lines": lines,
        "sheet_summaries": sheet_summaries,
        "skipped_sheets": skipped_sheets,
        "warnings": warnings,
        "summary": summarize(lines, sheet_summaries, warnings),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse classified Cash/Credit Card record workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out-json", type=Path, required=True)
    args = parser.parse_args()

    payload = parse_cash_record(args.workbook)
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=json_default),
        encoding="utf-8",
    )
    print(json.dumps(payload["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
