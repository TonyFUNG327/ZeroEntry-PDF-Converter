from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_coa_from_workbook(path: Path, label: str) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    if "Chart of Accounts" not in wb.sheetnames:
        return []
    ws = wb["Chart of Accounts"]
    rows = []
    for row in range(1, ws.max_row + 1):
        name = text(ws.cell(row, 1).value)
        code = text(ws.cell(row, 2).value)
        if not name or not code or name.casefold() == "account description":
            continue
        rows.append({"source": label, "account_description": name, "account_id": code})
    return rows


def reconcile(sources: list[tuple[Path, str]]) -> dict[str, Any]:
    rows = []
    for path, label in sources:
        rows.extend(read_coa_from_workbook(path, label))

    by_name = defaultdict(set)
    by_id = defaultdict(set)
    for row in rows:
        by_name[row["account_description"].casefold()].add(row["account_id"])
        by_id[row["account_id"]].add(row["account_description"])

    issues = []
    for name_key, ids in sorted(by_name.items()):
        if len(ids) > 1:
            examples = [row for row in rows if row["account_description"].casefold() == name_key]
            issues.append(
                {
                    "issue_type": "Same account description has different IDs",
                    "account_description": examples[0]["account_description"],
                    "account_ids": sorted(ids),
                    "sources": sorted({row["source"] for row in examples}),
                }
            )
    for account_id, names in sorted(by_id.items()):
        if len(names) > 1:
            examples = [row for row in rows if row["account_id"] == account_id]
            issues.append(
                {
                    "issue_type": "Same account ID has different descriptions",
                    "account_id": account_id,
                    "account_descriptions": sorted(names),
                    "sources": sorted({row["source"] for row in examples}),
                }
            )

    return {
        "source_count": len(sources),
        "account_row_count": len(rows),
        "issue_count": len(issues),
        "accounts": rows,
        "issues": issues,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Reconcile Chart of Accounts across input workbooks.")
    parser.add_argument("--source", action="append", nargs=2, metavar=("LABEL", "WORKBOOK"), required=True)
    parser.add_argument("--out-json", type=Path, required=True)
    args = parser.parse_args()

    sources = [(Path(path), label) for label, path in args.source]
    payload = reconcile(sources)
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({k: payload[k] for k in ("source_count", "account_row_count", "issue_count")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
