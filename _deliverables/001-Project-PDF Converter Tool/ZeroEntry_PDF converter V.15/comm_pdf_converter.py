import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = ["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"]
DEFAULT_INPUT_PATH = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "generated_output"
DATE_RE = re.compile(r"^\d{4}/\d{2}/\d{2}$")
AMOUNT_RE = re.compile(r"^\(?\d{1,3}(?:,\d{3})*(?:\.\d{2})\)?$|^\(?\d+\.\d{2}\)?$")
CURRENCIES = {"HKD", "USD", "CNY", "EUR", "GBP", "AUD", "CAD", "CHF", "JPY", "NOK", "NZD", "SGD"}


def amount_value(text):
    sign = -1 if text.startswith("(") and text.endswith(")") else 1
    return sign * float(text.strip("()").replace(",", ""))


def parse_date(text):
    return datetime.strptime(text, "%Y/%m/%d")


def words_to_lines(words):
    rows = []
    for word in words:
        if word["top"] < 75 or word["top"] > 765:
            continue
        for row in rows:
            if abs(row["top"] - word["top"]) <= 2.2:
                row["words"].append(word)
                row["top"] = (row["top"] + word["top"]) / 2
                break
        else:
            rows.append({"top": word["top"], "words": [word]})
    rows.sort(key=lambda row: row["top"])
    for row in rows:
        row["words"].sort(key=lambda word: word["x0"])
    return rows


def classify_amount(word):
    x0, x1 = word["x0"], word["x1"]
    if 365 <= x1 <= 425:
        return "Withdrawal"
    if 430 <= x1 <= 500:
        return "Deposit"
    if x0 >= 505:
        return "Balance"
    return None


def append_description(row, extra):
    extra = extra.strip()
    if extra:
        row["Description"] = f"{row['Description']} {extra}".strip()


def is_header_or_footer_line(text):
    markers = [
        "ACCOUNT SUMMARY",
        "BALANCE IN",
        "CONSOLIDATED STATEMENT",
        "CST NO.",
        "DEPOSITS IN",
        "END OF STATEMENT",
        "EXCHANGE RATE",
        "IN CASE OF ERROR",
        "ORIGINAL CURRENCY",
        "PAGE",
        "PERIOD",
        "PRINTED ON",
        "REFERENCE",
        "SAVINGS/CURRENT DEPOSITS ACTIVITIES",
        "TOTAL DEPOSITS",
        "TO BE CONTINUED",
        "TRANSACTION DETAILS",
        "VALUE DATE",
        "WITHDRAWALS IN",
        "交易摘要",
        "會計日期",
        "原幣",
        "貨幣",
    ]
    return any(marker in text for marker in markers)


def sheet_title_for_account(account):
    if account == "Bank of Communications HKD Savings":
        return "COMM HKD Savings"
    if account == "Bank of Communications HKD Current":
        return "COMM HKD Current"
    match = re.match(r"^Bank of Communications ([A-Z]{3}) (Savings|Current)$", account)
    if match:
        return f"COMM {match.group(1)} {match.group(2)}"
    return account[:31]


def account_name(currency, account_type):
    return f"Bank of Communications {currency} {account_type}"


def extract_consolidated_pdf(pdf_path):
    accounts = defaultdict(list)
    account_type = None
    current_account = None
    in_activities = False
    pending_desc = []
    last_row = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            lines = words_to_lines(page.extract_words(x_tolerance=1, y_tolerance=3) or [])
            for line in lines:
                words = line["words"]
                text = " ".join(word["text"] for word in words).strip()

                if "SAVINGS/CURRENT DEPOSITS ACTIVITIES" in text:
                    in_activities = True
                    continue
                if not in_activities:
                    continue
                if "TO BE CONTINUED" in text or "END OF STATEMENT" in text:
                    in_activities = False
                    pending_desc = []
                    last_row = None
                    continue
                if "DEPOSIT INTEREST RATE INFORMATION" in text:
                    in_activities = False
                    pending_desc = []
                    last_row = None
                    continue
                if "TOTAL TRANSACTION AMOUNT" in text or "NO.OF TRANSACTION" in text:
                    pending_desc = []
                    continue
                if is_header_or_footer_line(text):
                    continue

                if "SAVINGS" in text and "：" in text:
                    account_type = "Savings"
                    current_account = None
                    pending_desc = []
                    last_row = None
                    continue
                if "CURRENT" in text and "：" in text:
                    account_type = "Current"
                    current_account = None
                    pending_desc = []
                    last_row = None
                    continue

                date_words = [word for word in words if DATE_RE.match(word["text"])]
                date_value = parse_date(date_words[0]["text"]) if date_words else None
                currency_words = [word for word in words if word["text"] in CURRENCIES]
                currency = currency_words[0]["text"] if currency_words else None

                amounts = {}
                desc_words = []
                for word in words:
                    if DATE_RE.match(word["text"]) or word["text"] in CURRENCIES:
                        continue
                    cls = classify_amount(word) if AMOUNT_RE.match(word["text"]) else None
                    if cls:
                        amounts[cls] = amount_value(word["text"])
                    elif 160 <= word["x0"] <= 360:
                        desc_words.append(word["text"])

                description = " ".join(desc_words).strip()
                if not date_value and "BAL B/F" in text and account_type and currency:
                    current_account = account_name(currency, account_type)
                    balance = amounts.get("Balance")
                    if balance is not None:
                        row = {
                            "Bank_Account": current_account,
                            "Date": None,
                            "Description": "BAL B/F",
                            "Deposit": None,
                            "Withdrawal": None,
                            "Balance": balance,
                        }
                        accounts[current_account].append(row)
                        last_row = row
                    pending_desc = []
                    continue

                if date_value and account_type and currency:
                    current_account = account_name(currency, account_type)
                if date_value and current_account and amounts:
                    desc_parts = pending_desc[:]
                    if description:
                        desc_parts.append(description)
                    row = {
                        "Bank_Account": current_account,
                        "Date": date_value,
                        "Description": " ".join(desc_parts).strip() or "BALANCE",
                        "Deposit": amounts.get("Deposit"),
                        "Withdrawal": amounts.get("Withdrawal"),
                        "Balance": amounts.get("Balance"),
                    }
                    accounts[current_account].append(row)
                    last_row = row
                    pending_desc = []
                    continue

                if not date_value and description:
                    if last_row and last_row.get("Date") is not None and not pending_desc:
                        append_description(last_row, description)
                    else:
                        pending_desc.append(description)

    for rows in accounts.values():
        control = None
        for row in rows:
            if row["Deposit"] is None and row["Withdrawal"] is None and row["Balance"] is not None:
                control = row["Balance"]
            else:
                if control is None:
                    control = row["Balance"] or 0.0
                control += row["Deposit"] or 0.0
                control -= row["Withdrawal"] or 0.0
            row["Control"] = round(control, 2) if control is not None else None

    return dict(accounts)


def extract_pdf(pdf_path):
    accounts = defaultdict(list)
    current_account = None
    in_table = False
    pending_desc = []
    last_row = None

    def account_from_line(text):
        if "SAVINGS" in text and "382" in text:
            return "Bank of Communications HKD Savings"
        if "CURRENT" in text and "382" in text:
            return "Bank of Communications HKD Current"
        return None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            lines = words_to_lines(page.extract_words(x_tolerance=1, y_tolerance=3) or [])
            for line in lines:
                words = line["words"]
                text = " ".join(word["text"] for word in words).strip()

                detected_account = account_from_line(text)
                if detected_account:
                    current_account = detected_account
                    in_table = True
                    pending_desc = []
                    last_row = None
                    continue
                if "TOTAL TRANSACTION AMOUNT" in text:
                    in_table = False
                    pending_desc = []
                    continue
                if is_header_or_footer_line(text):
                    continue
                if not in_table or not current_account:
                    continue

                date_words = [word for word in words if DATE_RE.match(word["text"])]
                date_value = parse_date(date_words[0]["text"]) if date_words else None

                amounts = {}
                desc_words = []
                for word in words:
                    if DATE_RE.match(word["text"]) or word["text"] == "HKD":
                        continue
                    cls = classify_amount(word) if AMOUNT_RE.match(word["text"]) else None
                    if cls:
                        amounts[cls] = amount_value(word["text"])
                    elif 165 <= word["x0"] <= 360:
                        desc_words.append(word["text"])

                description = " ".join(desc_words).strip()
                if not date_value and "BAL B/F" in text:
                    balance = amounts.get("Balance")
                    if balance is not None and current_account:
                        row = {
                            "Bank_Account": current_account,
                            "Date": None,
                            "Description": "BAL B/F",
                            "Deposit": None,
                            "Withdrawal": None,
                            "Balance": balance,
                        }
                        accounts[current_account].append(row)
                        last_row = row
                    pending_desc = []
                    continue
                if not date_value and description:
                    pending_desc.append(description)
                    continue
                if not date_value and last_row and pending_desc:
                    append_description(last_row, " ".join(pending_desc))
                    pending_desc = []
                    continue
                if not date_value:
                    continue
                if not amounts:
                    pending_desc = []
                    continue

                desc_parts = pending_desc[:]
                if description:
                    desc_parts.append(description)
                pending_desc = []
                row = {
                    "Bank_Account": current_account,
                    "Date": date_value,
                    "Description": " ".join(desc_parts).strip() or "BALANCE",
                    "Deposit": amounts.get("Deposit"),
                    "Withdrawal": amounts.get("Withdrawal"),
                    "Balance": amounts.get("Balance"),
                }
                accounts[current_account].append(row)
                last_row = row

    for rows in accounts.values():
        control = None
        for row in rows:
            if row["Deposit"] is None and row["Withdrawal"] is None and row["Balance"] is not None:
                control = row["Balance"]
            else:
                if control is None:
                    control = row["Balance"] or 0.0
                control += row["Deposit"] or 0.0
                control -= row["Withdrawal"] or 0.0
            row["Control"] = round(control, 2) if control is not None else None

    result = dict(accounts)
    if not result:
        return extract_consolidated_pdf(pdf_path)
    return result


def write_workbook(accounts, output_path):
    output_path = Path(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    control_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    grid_side = Side(style="thin", color="D9D9D9")
    grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)
    widths = [34, 14, 58, 16, 16, 16, 16]

    for account, rows in accounts.items():
        sheet = workbook.create_sheet(sheet_title_for_account(account))
        sheet.append(HEADERS)
        for row in rows:
            sheet.append([row["Bank_Account"], row["Date"], row["Description"], row["Deposit"], row["Withdrawal"], row["Balance"], None])
        for cell in sheet[1]:
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = header_fill
            cell.border = grid_border
        for row_idx in range(2, sheet.max_row + 1):
            if row_idx == 2:
                sheet.cell(row_idx, 7).value = f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
            else:
                sheet.cell(row_idx, 7).value = f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},G{row_idx - 1}+IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10, bold=cell.column == 7 or cell.row == 1)
                cell.border = grid_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if cell.column == 7:
                    cell.fill = control_fill
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in sheet["B"][1:]:
            cell.number_format = "dd mmm yyyy"
        for cell in sheet["D"][1:]:
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        for cell in sheet["E"][1:]:
            cell.number_format = '(#,##0.00);(#,##0.00);"-"'
        for column in ["F", "G"]:
            for cell in sheet[column][1:]:
                cell.number_format = "#,##0.00;(#,##0.00);0.00"
        sheet.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def validate_accounts(accounts):
    report = []
    for account, rows in accounts.items():
        deposit_total = round(sum(row["Deposit"] or 0 for row in rows), 2)
        withdrawal_total = round(sum(row["Withdrawal"] or 0 for row in rows), 2)
        final_balance = rows[-1]["Balance"] if rows else None
        final_control = rows[-1]["Control"] if rows else None
        balance_mismatches = [idx + 2 for idx, row in enumerate(rows) if row["Balance"] is not None and abs(row["Balance"] - row["Control"]) > 0.01]
        report.append({
            "account": account,
            "rows": len(rows),
            "deposit_count": sum(1 for row in rows if row["Deposit"] is not None),
            "withdrawal_count": sum(1 for row in rows if row["Withdrawal"] is not None),
            "deposit_total": deposit_total,
            "withdrawal_total": withdrawal_total,
            "final_balance": final_balance,
            "final_control": final_control,
            "balance_mismatches": balance_mismatches,
        })
    return report


def collect_pdf_paths(input_path):
    input_path = Path(input_path)
    if input_path.is_file():
        return [input_path]
    return sorted(path for path in input_path.iterdir() if path.is_file() and path.suffix.lower() == ".pdf")


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path, nargs="?", default=DEFAULT_INPUT_PATH)
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args(argv)
    for pdf_path in collect_pdf_paths(args.input):
        accounts = extract_pdf(pdf_path)
        output_path = write_workbook(accounts, args.output_dir / f"{pdf_path.stem}.xlsx")
        print(f"Saved: {output_path}")
        for item in validate_accounts(accounts):
            print(
                f"{item['account']}: rows={item['rows']}, deposits={item['deposit_count']} amount={item['deposit_total']:,.2f}, "
                f"withdrawals={item['withdrawal_count']} amount={item['withdrawal_total']:,.2f}, "
                f"final_balance={item['final_balance']:,.2f}, final_control={item['final_control']:,.2f}, "
                f"balance_mismatches={len(item['balance_mismatches'])}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
