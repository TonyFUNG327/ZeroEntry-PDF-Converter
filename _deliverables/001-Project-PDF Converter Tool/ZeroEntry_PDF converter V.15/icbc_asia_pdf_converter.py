import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = ["Bank_Account", "Date", "Description", "Deposit", "Withdrawal", "Balance", "Control"]
DEFAULT_INPUT_PATH = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "generated_output"
DATE_RE = re.compile(r"^\d{4}/\d{2}/\d{2}$")
AMOUNT_RE = re.compile(r"^\(?\d{1,3}(?:,\d{3})*(?:\.\d{2})\)?$|^\(?\d+\.\d{2}\)?$")
ACCOUNT_RE = re.compile(r"^(.+?)\s+账户号码:\s*([0-9-]+)$")
CURRENCIES = {"HKD", "USD", "CNY", "EUR", "GBP", "AUD", "CAD", "CHF", "JPY", "SGD"}


def clean_text(text):
    return " ".join((text or "").replace("\ufeff", " ").replace("\u200b", " ").split())


def amount_value(text):
    sign = -1 if text.startswith("(") and text.endswith(")") else 1
    return sign * float(text.strip("()").replace(",", ""))


def nonzero(value):
    if value is None or abs(value) < 0.005:
        return None
    return value


def parse_date(text):
    return datetime.strptime(text, "%Y/%m/%d")


def account_name(currency, account_type):
    normalized = clean_text(account_type)
    if "往來" in normalized or "往来" in normalized:
        account_kind = "Current Account"
    elif "儲蓄" in normalized or "储蓄" in normalized:
        account_kind = "Savings Account"
    else:
        account_kind = normalized
    return f"ICBC Asia {currency} {account_kind}"


def sheet_title_for_account(account):
    return account.replace("Current Account", "Current")[:31]


def append_description(row, extra):
    extra = clean_text(extra)
    if extra:
        row["Description"] = f"{row['Description']} {extra}".strip()


def is_noise_line(text):
    markers = [
        "更改地址通知书",
        "存户核对收支项目对账表",
        "各种交易代码",
        "客户服务热线",
        "中國工商銀行",
        "INDUSTRIAL AND COMMERCIAL BANK",
        "Customer Service Hotline",
        "www.icbcasia.com",
        "备注：",
        "由于市场",
        "保险资料",
        "请详细核对",
        "# - 已作抵押",
        "OD -",
        "DR -",
        "CR -",
        "资产组合总览",
        "户口摘要",
        "定期存款",
    ]
    return any(marker in text for marker in markers)


def split_transaction_line(text):
    parts = text.split()
    if not parts or not DATE_RE.match(parts[0]):
        return None
    currency_idx = None
    for idx, part in enumerate(parts[1:], start=1):
        if part in CURRENCIES:
            currency_idx = idx
            break
    if currency_idx is None:
        return None
    amounts = []
    trailing_notes = []
    for token in parts[currency_idx + 1:]:
        if AMOUNT_RE.match(token):
            amounts.append(amount_value(token))
        else:
            trailing_notes.append(token)
    if not amounts:
        return None
    balance = amounts[-1]
    if trailing_notes and trailing_notes[-1].upper() == "DR":
        balance = -abs(balance)
    return {
        "date": parse_date(parts[0]),
        "description": clean_text(" ".join(parts[1:currency_idx])),
        "currency": parts[currency_idx],
        "amounts": amounts,
        "balance": balance,
    }


def classify_amount(amount, balance, previous_balance, description):
    if amount is None:
        return None, None
    if previous_balance is not None:
        if abs(round(previous_balance + amount, 2) - round(balance, 2)) <= 0.01:
            return amount, None
        if abs(round(previous_balance - amount, 2) - round(balance, 2)) <= 0.01:
            return None, amount

    deposit_markers = ["存入", "存款", "FPS Transfer Reject", "FPS Transfer", "匯款存入", "汇款存入", "利息存入"]
    withdrawal_markers = ["提款", "FPS至", "繳費", "缴费", "銀行收費", "银行收费", "利息支出", "CQW"]
    if any(marker in description for marker in deposit_markers):
        return amount, None
    if any(marker in description for marker in withdrawal_markers):
        return None, amount
    return None, amount


def extract_pdf(pdf_path):
    accounts = defaultdict(list)
    current_account = None
    current_balance = {}
    last_row = None
    in_transactions = False

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for raw_line in text.splitlines():
                line = clean_text(raw_line)
                if not line:
                    continue

                if "户口进支" in line:
                    in_transactions = True
                    last_row = None
                    continue
                if not in_transactions:
                    continue
                if is_noise_line(line):
                    last_row = None
                    continue
                if "日期" in line and "摘要" in line and "结余" in line:
                    last_row = None
                    continue
                if line.startswith("进支摘要"):
                    last_row = None
                    continue

                account_match = ACCOUNT_RE.match(line)
                if account_match:
                    account_type = account_match.group(1)
                    current_account = account_name("", account_type).replace("ICBC Asia  ", "ICBC Asia ")
                    last_row = None
                    continue

                parsed = split_transaction_line(line)
                if parsed is None:
                    if last_row is not None and line.startswith("("):
                        append_description(last_row, line)
                    continue

                currency = parsed["currency"]
                if current_account is None:
                    continue
                if not current_account.startswith(f"ICBC Asia {currency} "):
                    account_kind = current_account.replace("ICBC Asia ", "").strip()
                    current_account = f"ICBC Asia {currency} {account_kind}"

                description = parsed["description"] or "BALANCE"
                balance = parsed["balance"]

                if "户口结余" in description:
                    last_row = None
                    current_balance[current_account] = balance
                    continue

                if "承上结余" in description:
                    row = {
                        "Bank_Account": current_account,
                        "Date": parsed["date"],
                        "Description": "B/F BALANCE",
                        "Deposit": None,
                        "Withdrawal": None,
                        "Balance": balance,
                    }
                    accounts[current_account].append(row)
                    current_balance[current_account] = balance
                    last_row = row
                    continue

                if len(parsed["amounts"]) < 2:
                    continue
                amount = parsed["amounts"][-2]
                previous_balance = current_balance.get(current_account)
                deposit, withdrawal = classify_amount(amount, balance, previous_balance, description)
                row = {
                    "Bank_Account": current_account,
                    "Date": parsed["date"],
                    "Description": description,
                    "Deposit": nonzero(deposit),
                    "Withdrawal": nonzero(withdrawal),
                    "Balance": balance,
                }
                accounts[current_account].append(row)
                current_balance[current_account] = balance
                last_row = row

    for rows in accounts.values():
        control = None
        for row in rows:
            if row["Deposit"] is None and row["Withdrawal"] is None and row["Balance"] is not None:
                control = row["Balance"]
            else:
                if control is None:
                    control = row["Balance"] or 0.0
                control += row["Deposit"] or 0.0
                control -= row["Withdrawal"] or 0.0
            row["Control"] = round(control, 2) if control is not None else None

    return dict(accounts)


def write_workbook(accounts, output_path):
    output_path = Path(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    control_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    grid_side = Side(style="thin", color="D9D9D9")
    grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)
    widths = [30, 14, 58, 16, 16, 16, 16]

    for account, rows in accounts.items():
        sheet = workbook.create_sheet(sheet_title_for_account(account))
        sheet.append(HEADERS)
        for row in rows:
            sheet.append([
                row["Bank_Account"],
                row["Date"],
                row["Description"],
                row["Deposit"],
                row["Withdrawal"],
                row["Balance"],
                None,
            ])
        for cell in sheet[1]:
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = header_fill
            cell.border = grid_border
        for row_idx in range(2, sheet.max_row + 1):
            if row_idx == 2:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )
            else:
                sheet.cell(row_idx, 7).value = (
                    f'=IF(AND(D{row_idx}="",E{row_idx}=""),F{row_idx},'
                    f'G{row_idx - 1}+IF(D{row_idx}="",0,D{row_idx})-IF(E{row_idx}="",0,E{row_idx}))'
                )
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10, bold=cell.column == 7 or cell.row == 1)
                cell.border = grid_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if cell.column == 7:
                    cell.fill = control_fill
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in sheet["B"][1:]:
            cell.number_format = "dd mmm yyyy"
        for cell in sheet["D"][1:]:
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        for cell in sheet["E"][1:]:
            cell.number_format = '(#,##0.00);(#,##0.00);"-"'
        for column in ["F", "G"]:
            for cell in sheet[column][1:]:
                cell.number_format = "#,##0.00;(#,##0.00);0.00"
        sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def validate_accounts(accounts):
    report = []
    for account, rows in accounts.items():
        deposit_total = round(sum(row["Deposit"] or 0 for row in rows), 2)
        withdrawal_total = round(sum(row["Withdrawal"] or 0 for row in rows), 2)
        final_balance = rows[-1]["Balance"] if rows else None
        final_control = rows[-1]["Control"] if rows else None
        balance_mismatches = [
            idx + 2
            for idx, row in enumerate(rows)
            if row["Balance"] is not None and abs(row["Balance"] - row["Control"]) > 0.01
        ]
        report.append({
            "account": account,
            "rows": len(rows),
            "deposit_count": sum(1 for row in rows if row["Deposit"] is not None),
            "withdrawal_count": sum(1 for row in rows if row["Withdrawal"] is not None),
            "deposit_total": deposit_total,
            "withdrawal_total": withdrawal_total,
            "final_balance": final_balance,
            "final_control": final_control,
            "balance_mismatches": balance_mismatches,
        })
    return report


def collect_pdf_paths(input_path):
    input_path = Path(input_path)
    if input_path.is_file():
        return [input_path]
    return sorted(path for path in input_path.iterdir() if path.is_file() and path.suffix.lower() == ".pdf")


def main(argv=None):
    parser = argparse.ArgumentParser(description="Convert ICBC Asia statement PDF to Excel.")
    parser.add_argument("input", type=Path, nargs="?", default=DEFAULT_INPUT_PATH)
    parser.add_argument("-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args(argv)
    for pdf_path in collect_pdf_paths(args.input):
        accounts = extract_pdf(pdf_path)
        output_path = write_workbook(accounts, args.output_dir / f"{pdf_path.stem}.xlsx")
        print(f"Saved: {output_path}")
        for item in validate_accounts(accounts):
            print(
                f"{item['account']}: rows={item['rows']}, deposits={item['deposit_count']} amount={item['deposit_total']:,.2f}, "
                f"withdrawals={item['withdrawal_count']} amount={item['withdrawal_total']:,.2f}, "
                f"final_balance={item['final_balance']:,.2f}, final_control={item['final_control']:,.2f}, "
                f"balance_mismatches={len(item['balance_mismatches'])}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
