from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
BANK_PATH = BASE / "001-I026-Bank record-31.12.25.converted.xlsx"
REPORT_PATH = BASE / "002-I026-TB, GL, PL, BS-31.12.25.xlsx"
OUT = BASE / "i026_process_upgrade_notes.md"


def fmt(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        joined = "|".join(vals)
        if "no." in joined and "date" in joined and "particular" in joined:
            return r
    return None


def get_labels(ws, header_row):
    return {c: str(ws.cell(header_row, c).value or "").strip() for c in range(1, ws.max_column + 1)}


def bank_summaries():
    wbf = load_workbook(BANK_PATH, data_only=False)
    wbv = load_workbook(BANK_PATH, data_only=True)
    sections = []
    for ws in wbf.worksheets:
        if ws.title == "Chart of Accounts":
            continue
        vws = wbv[ws.title]
        header = find_header_row(ws)
        labels = get_labels(ws, header) if header else {}
        totals = defaultdict(float)
        nonblank_transactions = []
        control_nonzero = []
        helper_cols = {}
        if header:
            for c, label in labels.items():
                sub = str(ws.cell(header + 1, c).value or "").strip()
                if label in {"Reference", "Description", "Number of Distributions", "Bank (Amount$)", "Nature (Amount$)"} or sub in {"=C4", "=C5", "(Nature)", "(Amount$)"}:
                    helper_cols[c] = label or sub
            for r in range(header + 1, ws.max_row + 1):
                nature = vws.cell(r, 3).value
                date = vws.cell(r, 2).value
                if nature not in (None, "") and date not in (None, ""):
                    nonblank_transactions.append(r)
                for c, label in labels.items():
                    v = vws.cell(r, c).value
                    if isinstance(v, (int, float)) and abs(v) > 0.00001 and label:
                        totals[label] += v
                for c, label in labels.items():
                    if label == "Control":
                        v = vws.cell(r, c).value
                        if isinstance(v, (int, float)) and abs(v) > 0.00001:
                            control_nonzero.append((r, v))
        sections.append(f"## Input Sheet: {ws.title}")
        sections.append(f"- Account/record label: {fmt(vws['C3'].value)}")
        sections.append(f"- COA name: {fmt(vws['C4'].value)}")
        sections.append(f"- Detected header row: {header}")
        sections.append(f"- Nonblank dated transaction rows: {len(nonblank_transactions)}")
        if nonblank_transactions[:8]:
            sections.append("- First transaction rows:")
            for r in nonblank_transactions[:8]:
                sections.append(
                    f"  - row {r}: date={fmt(vws.cell(r,2).value)}, nature={fmt(vws.cell(r,3).value)}, "
                    f"payer/payee={fmt(vws.cell(r,4).value)}, invoice={fmt(vws.cell(r,5).value)}"
                )
        if helper_cols:
            sections.append("- Helper columns detected:")
            sections.append("  " + "; ".join(f"{c}:{label}" for c, label in sorted(helper_cols.items())))
        if control_nonzero[:10]:
            sections.append("- Non-zero control rows observed:")
            sections.append("  " + "; ".join(f"row {r}={v}" for r, v in control_nonzero[:10]))
        important = [
            "Deposit", "Withdrawal", "Control", "Interbank", "Receipt in advance",
            "Trade receivable", "Prepayment", "Accruals", "Membership fee income",
            "Sales", "Bank interest income", "Bank charges", "Web hosting fee",
            "Bank (Amount$)", "Nature (Amount$)",
        ]
        sections.append("- Key totals:")
        for label in important:
            if abs(totals.get(label, 0)) > 0.00001:
                sections.append(f"  - {label}: {totals[label]}")
        sections.append("")
    return sections


def coa_summary():
    wb = load_workbook(BANK_PATH, data_only=True)
    ws = wb["Chart of Accounts"]
    lines = ["## Chart of Accounts Observations"]
    groups = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        name, aid = ws.cell(r, 1).value, ws.cell(r, 2).value
        if name in (None, "") or aid in (None, ""):
            continue
        try:
            aid_int = int(aid)
        except Exception:
            aid_int = 0
        bucket = f"{aid_int // 1000}xxx"
        groups[bucket].append((aid, name))
    for bucket in sorted(groups):
        sample = "; ".join(f"{aid} {name}" for aid, name in groups[bucket][:10])
        lines.append(f"- {bucket}: {sample}")
    lines.append("")
    return lines


def report_summary():
    wb = load_workbook(REPORT_PATH, data_only=True)
    lines = ["## Output Report Snapshot"]
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"### {sheet}")
        for r in range(1, ws.max_row + 1):
            vals = [fmt(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 9) + 1)]
            if any(vals):
                if sheet.startswith("GL") and r > 35:
                    continue
                lines.append(f"- row {r}: " + " | ".join(vals))
        lines.append("")
    return lines


def manual_jv_summary():
    wb = load_workbook(REPORT_PATH, data_only=True)
    ws = wb["GL 31.12.25"]
    refs = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(r, 4).value
        if isinstance(ref, str) and ref.startswith("JV"):
            refs[ref].append((ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 6).value, ws.cell(r, 7).value, ws.cell(r, 8).value))
    lines = ["## Manual JV References Found in GL"]
    for ref, rows in sorted(refs.items()):
        lines.append(f"- {ref}:")
        for account_id, account, desc, debit, credit in rows[:12]:
            lines.append(f"  - {account_id} {account}: Dr={fmt(debit)} Cr={fmt(credit)} Desc={fmt(desc)}")
    lines.append("")
    return lines


def main():
    sections = [
        "# I026 Process Upgrade Notes",
        "",
        "I026 confirms the H026 advanced structure but adds deferred income, AR reclassification, PPE/depreciation, finance-cost style non-bank schedules, and broader report sections.",
        "",
    ]
    sections += bank_summaries()
    sections += coa_summary()
    sections += manual_jv_summary()
    sections += report_summary()
    OUT.write_text("\n".join(sections), encoding="utf-8")
    print(OUT)


if __name__ == "__main__":
    main()
