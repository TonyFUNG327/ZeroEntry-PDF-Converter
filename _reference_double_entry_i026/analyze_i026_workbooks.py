from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
BANK_PATH = BASE / "001-I026-Bank record-31.12.25.converted.xlsx"
REPORT_PATH = BASE / "002-I026-TB, GL, PL, BS-31.12.25.xlsx"
JSON_OUT = BASE / "i026_structure_summary.json"
MD_OUT = BASE / "i026_accounting_mapping_notes.md"


def fmt(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def used_bounds(ws):
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if min_row is None:
        return None
    return {"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col}


def formula_samples(ws, limit=25):
    samples = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                samples.append({"cell": cell.coordinate, "formula": cell.value})
                if len(samples) >= limit:
                    return samples
    return samples


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        joined = "|".join(vals).lower()
        if "no." in joined and "date" in joined and "particular" in joined:
            return r
    return None


def nonempty_row(ws, row):
    vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if v not in (None, ""):
            vals.append(f"{get_column_letter(c)}={fmt(v)}")
    return vals


def workbook_summary(path):
    wb = load_workbook(path, data_only=False)
    result = {"path": str(path), "sheets": []}
    for ws in wb.worksheets:
        result["sheets"].append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "used_bounds": used_bounds(ws),
                "header_row": find_header_row(ws),
                "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:20]],
                "formula_samples": formula_samples(ws),
            }
        )
    return result


def bank_sheet_notes(ws_formula, ws_values):
    header_row = find_header_row(ws_formula)
    lines = [f"## Bank/Cash record sheet: {ws_formula.title}", ""]
    lines.append(f"- Company: {fmt(ws_values['C1'].value)}")
    lines.append(f"- Period: {fmt(ws_values['C2'].value)}")
    lines.append(f"- Account label: {fmt(ws_values['C3'].value)}")
    lines.append(f"- Detected transaction header row: {header_row}")
    for r in range(1, min(ws_formula.max_row, 15) + 1):
        vals = nonempty_row(ws_formula, r)
        if vals:
            lines.append(f"- header row {r}: " + "; ".join(vals[:45]))

    if not header_row:
        return "\n".join(lines)

    labels = []
    for c in range(1, ws_formula.max_column + 1):
        label = ws_formula.cell(header_row, c).value
        if label not in (None, ""):
            labels.append((c, str(label)))
    lines.append("- Header labels:")
    lines.append("  " + "; ".join(f"{get_column_letter(c)}={label}" for c, label in labels[:80]))

    helper_names = {
        "Reference",
        "G/L Account (Bank)",
        "G/L Account (Nature)",
        "Description",
        "Bank (Amount$)",
        "Nature (Amount$)",
        "Number of Distributions",
    }
    helper_cols = [(c, label) for c, label in labels if label in helper_names]
    lines.append("- Peachtree/helper columns:")
    lines.append("  " + "; ".join(f"{get_column_letter(c)}={label}" for c, label in helper_cols))

    active_by_label = defaultdict(float)
    active_rows = []
    for r in range(header_row + 1, ws_values.max_row + 1):
        active = []
        for c, label in labels:
            v = ws_values.cell(r, c).value
            if isinstance(v, (int, float)) and abs(v) > 0.00001:
                active.append((c, label, v))
                active_by_label[label] += v
        if active:
            active_rows.append((r, active))

    lines.append("- First active rows:")
    for r, active in active_rows[:25]:
        context = " | ".join(fmt(ws_values.cell(r, c).value) for c in range(1, min(ws_values.max_column, 9) + 1))
        active_bits = "; ".join(f"{get_column_letter(c)}:{label}={value}" for c, label, value in active[:16])
        lines.append(f"  - row {r}: {context} || {active_bits}")

    lines.append("- Non-zero totals by header label:")
    for label, total in sorted(active_by_label.items()):
        if abs(total) > 0.00001:
            lines.append(f"  - {label}: {total}")
    return "\n".join(lines)


def report_sheet_notes(ws_formula, ws_values, max_rows=140):
    lines = [f"## Report sheet: {ws_formula.title}", ""]
    for r in range(1, min(ws_formula.max_row, max_rows) + 1):
        fvals, vvals = [], []
        for c in range(1, min(ws_formula.max_column, 12) + 1):
            fv = ws_formula.cell(r, c).value
            vv = ws_values.cell(r, c).value
            if fv not in (None, "") or vv not in (None, ""):
                fvals.append(fmt(fv))
                vvals.append(fmt(vv))
        if fvals:
            if fvals == vvals:
                lines.append(f"- row {r}: " + " | ".join(fvals))
            else:
                lines.append(f"- row {r}: formula=[" + " | ".join(fvals) + "]")
                lines.append(f"          values =[" + " | ".join(vvals) + "]")
    return "\n".join(lines)


def main():
    summaries = {
        "bank_record": workbook_summary(BANK_PATH),
        "reports": workbook_summary(REPORT_PATH),
    }
    JSON_OUT.write_text(json.dumps(summaries, indent=2, ensure_ascii=False), encoding="utf-8")

    bank_formula = load_workbook(BANK_PATH, data_only=False)
    bank_values = load_workbook(BANK_PATH, data_only=True)
    report_formula = load_workbook(REPORT_PATH, data_only=False)
    report_values = load_workbook(REPORT_PATH, data_only=True)

    sections = ["# I026 Accounting Mapping Notes", "", "## Workbook Overview"]
    for key, summary in summaries.items():
        sections.append(f"- {key}: {Path(summary['path']).name}")
        for sheet in summary["sheets"]:
            sections.append(
                f"  - {sheet['name']}: rows={sheet['max_row']}, cols={sheet['max_col']}, "
                f"used={sheet['used_bounds']}, header_row={sheet['header_row']}, "
                f"formulas_sampled={len(sheet['formula_samples'])}"
            )
    sections.append("")

    for ws in bank_formula.worksheets:
        if ws.title.lower() not in {"instruction"}:
            sections.append(bank_sheet_notes(ws, bank_values[ws.title]))
            sections.append("")

    for ws in report_formula.worksheets:
        sections.append(report_sheet_notes(ws, report_values[ws.title]))
        sections.append("")

    MD_OUT.write_text("\n".join(sections), encoding="utf-8")
    print(JSON_OUT)
    print(MD_OUT)


if __name__ == "__main__":
    main()
